# -*- coding: utf-8 -*-
"""
Created on SEP 23 2018\n
LDA with Collapsed Gibbs sampling + fixed-point iteration\n
@author: takeda masaki
"""


import numpy as np
from scipy.special import digamma
import sys
import random
import math
import csv
import pickle
import openpyxl
import gensim
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBar, FormatObject
from openpyxl.formatting.rule import Rule

from LDA_GIBBS import LDA_util


class LDA:
    """
    K         : number of topics\n
    alpha     : initial parameter of theta\n
                same value between documents\n
    beta      : initial parameter of phi\n
                same value between topics\n
    docs      : documents\n
                [
                    [
                        [
                            [word], [word], ...
                        ], [sentence], [sentence], ...
                    ], [document], [document], ...
                ]
    arrangeInSentece: If true, consider sentences as one independent document
    randomInit      : If true, latent topics will be set random topic
    scoreCalcRate   : freqency of history recording
    testDocs        : test documetns, used to calc perplexity(test)
    w2vModel        : word2vec model from wikipedia, used to calc coherence
    """

    def __init__(self, K, alpha, beta, docs, *,
                 arrangeInSentece=False, randomInit=False, scoreCalcRate=10, testDocs=None, w2vModel=None):
        """ 0. init variables """
        self.countLearned = 0
        self.docs = docs
        self.randomInit = randomInit

        # init learing materials
        self.wordcounts = {}
        docCount = {}
        for document in docs:
            f0 = []
            for sentence in document:
                for word in sentence:
                    if(word not in self.wordcounts.keys()):
                        self.wordcounts[word] = 0
                        docCount[word] = 0
                    self.wordcounts[word] += 1
                    if(word not in f0):
                        docCount[word] += 1
                        f0.append(word)
        self.words = []
        # ["word", "word", ...]
        for k, v in sorted(self.wordcounts.items(), key=lambda x: -x[1]):
            self.words.append(k)

        self.documents = []
        # [
        #     [0, 1, 2, ..(word id)..],
        #     [document], [document], ...
        # ]
        if(not arrangeInSentece):
            for document in docs:
                l0 = []
                for sentence in document:
                    for word in sentence:
                        l0.append(self.words.index(word))
                self.documents.append(l0)
        else:
            for document in docs:
                for sentence in document:
                    l0 = []
                    for word in sentence:
                        l0.append(self.words.index(word))
                    self.documents.append(l0)
        D = len(self.documents)
        V = len(self.words)
        self.arrangeInSentece = arrangeInSentece

        self.IDF = [0] * V
        for v, word in enumerate(self.words):
            self.IDF[v] = np.log(D / docCount[word])
        self.meanTFIDF = [0] * V
        for document in self.documents:
            for v in document:
                self.meanTFIDF[v] += np.log(D / docCount[self.words[v]]) / self.wordcounts[self.words[v]] / docCount[word]

        self.w2vModel = w2vModel  # word2vec model from wikipedia
        # d0 = []
        # for document in docs:
        #     l0 = []
        #     for sentence in document:
        #         l0 += sentence
        #     d0.append(l0)
        # self.w2vModel2 = gensim.models.Word2Vec(sentences=d0)  # word2vec model from this corpus

        cofreq = np.zeros((V, V))
        for document in self.documents:
            a0 = np.sort(np.unique(np.array(document)))
            for i in range(len(a0)):
                for j in range(i + 1, len(a0)):
                    cofreq[a0[i], a0[j]] += 1
        self.similarity = cofreq + cofreq.T + 1  # log( (D[v1, v2] + 1) / D[v2] )
        for v in range(V):
            self.similarity[:, v] /= docCount[self.words[v]]
        self.similarity = np.log(self.similarity)
        for v in range(V):
            self.similarity[v, v] = 0

        # init fields
        self.D = D  # number of documents
        self.V = V  # number of wordtype
        self.K = K  # number of topics

        # init parameter of each probability
        self.alpha = np.full(K, alpha)  # parameter of theta
        self.beta = np.full(V, beta)  # parameter of phi

        # init probability
        self.theta = np.zeros((D, K))  # word's topic probability for each document
        self.phi = np.zeros((K, V))  # word probability for each topic
        self.theta2 = np.zeros((D, K))
        self.theta3 = np.zeros((D, K))

        # init latent variables
        self.z = []  # word's topic, generated by theta
        for document in self.documents:
            self.z.append([0] * len(document))

        # init counts
        self.nWord_d = np.zeros(D)  # number of words, given document, (D)
        self.nWord_k = np.zeros(K)  # number of words, given topic, (K)
        self.nWord_dk = np.zeros((D, K))  # number of words, given document and topic, (D*K)
        self.nWord_kv = np.zeros((K, V))  # number of words, given topics and wordtype, (K*V)
        for d, document in enumerate(self.documents):
            self.nWord_d[d] = len(document)

        """ 1. sampleing initial values of z """
        self.sampleInitZ(randomInit)

        """ 2. update alpha and beta """
        self.updateAlphaAndBeta()
        self.calcThetaAndPhi()

        # init history
        self.scoreCalcRate = scoreCalcRate
        self.testDocs = testDocs
        self.loopsOnRecord = []
        self.perplexityHistory = []
        self.testPerplexityHistory = []
        self.ceherenceHistory = []
        self.ceherenceHistory2 = []
        self.addHistory()

    def sampleInitZ(self, randomInit):
        if(not randomInit):
            for d, document in enumerate(self.documents):
                for n, v in enumerate(document):
                    p0 = self.nWord_dk[d] + self.alpha
                    p1 = self.nWord_kv[:, v] + self.beta[v]
                    p2 = self.nWord_k + self.beta.sum()
                    p = p0 * p1 / p2
                    p = p / p.sum()
                    z0 = np.random.multinomial(1, p).argmax()
                    self.z[d][n] = z0
                    self.nWord_k[z0] += 1
                    self.nWord_dk[d, z0] += 1
                    self.nWord_kv[z0, v] += 1
        else:
            for d, document in enumerate(self.documents):
                for n, v in enumerate(document):
                    z0 = np.random.randint(0, self.K)
                    self.z[d][n] = z0
                    self.nWord_k[z0] += 1
                    self.nWord_dk[d, z0] += 1
                    self.nWord_kv[z0, v] += 1

    def learn(self):
        """
        Execute single learning loop\n
        Make sure to use #calcThetaAndPhi after learning loop
        """

        """ 1. sampleing z (Collapsed Gibbs sampling) (bottleneck)"""
        LDA_util.updateZ(self.documents,
                         self.alpha, self.beta,
                         self.nWord_dk, self.nWord_kv, self.nWord_k,
                         self.z)

        """ 2. update alpha and beta (fixed-point iteration) """
        self.updateAlphaAndBeta()

        self.countLearned += 1

    def startLearning(self, loop):
        """ Start learning loops """
        print("Start learning")
        print("Document: " + str(self.D))
        if(self.arrangeInSentece):
            print("(Consider sentences as one document)")
        print("Wordtypes: " + str(self.V))
        print("Topics: " + str(self.K))
        for i in range(loop):
            sys.stdout.write("\r{:>4}/{:>4}".format(i + 1, loop))
            sys.stdout.flush()
            self.learn()
            if(self.countLearned % self.scoreCalcRate == 0):
                sys.stdout.write(" ...History Recording")
                sys.stdout.flush()
                self.addHistory()
                sys.stdout.write("\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b")
                sys.stdout.write("                     ")
                sys.stdout.flush()
        if(self.countLearned != self.loopsOnRecord[len(self.loopsOnRecord) - 1]):
            sys.stdout.write(" ...History Recording")
            sys.stdout.flush()
            self.addHistory()
            sys.stdout.write("\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b")
            sys.stdout.write("                     ")
            sys.stdout.flush()
        self.calcThetaAndPhi()
        print("")

    def updateAlphaAndBeta(self):
        a0 = digamma(self.nWord_dk + np.resize(self.alpha, (self.D, self.K))).sum(axis=0)
        a0 -= self.D * digamma(self.alpha)
        a1 = digamma(self.nWord_d + np.full(self.D, self.alpha.sum())).sum()
        a1 -= self.D * digamma(self.alpha.sum())
        self.alpha *= a0 / a1
        b0 = digamma(self.nWord_kv + np.resize(self.beta, (self.K, self.V))).sum(axis=0)
        b0 -= self.K * digamma(self.beta)
        b1 = digamma(self.nWord_k + np.full(self.K, self.beta.sum())).sum()
        b1 -= self.K * digamma(self.beta.sum())
        self.beta *= b0 / b1

    def calcThetaAndPhi(self):
        t0 = self.nWord_dk + np.resize(self.alpha, (self.D, self.K))
        t1 = np.resize(self.nWord_d + self.alpha.sum(), (self.K, self.D)).T
        self.theta = t0 / t1
        p0 = self.nWord_kv + np.resize(self.beta, (self.K, self.V))
        p1 = np.resize(self.nWord_k + self.beta.sum(), (self.V, self.K)).T
        self.phi = p0 / p1

        self.theta2 = np.zeros((self.D, self.K))
        self.theta3 = np.zeros((self.D, self.K))
        phi_ = self.phi / self.phi.sum(axis=0, keepdims=True)
        for d, document in enumerate(self.documents):
            for n, v in enumerate(document):
                self.theta2[d] += phi_[:, v]
                self.theta3[d] += self.phi[:, v]

    def addHistory(self):
        """
        This method should be called by #startLearning, once per #scoreCalcRate
        """
        self.loopsOnRecord.append(self.countLearned)
        # self.perplexityHistory.append(self.calcPerplexity(self.docs, 20, 0.2))
        self.perplexityHistory.append(self.calcPerplexity2(self.docs))
        if(self.testDocs is not None):
            # self.testPerplexityHistory.append(self.calcPerplexity(self.testDocs, 20, 0.2))
            self.testPerplexityHistory.append(self.calcPerplexity2(self.testDocs))
        self.ceherenceHistory.append(self.calcCoherence(w2v=self.w2vModel))
        # if(self.w2vModel is not None):
        #     self.ceherenceHistory2.append(self.calcCoherence(w2v=self.w2vModel))

    # def calcPerplexity(self, testdocs, loop, testRatio):
    #     """
    #     Calculate perplexity(the value showing generalization ability)\n
    #     First, this splits testdocs into 2 parts "trainset" and "testset"\n
    #     Trainset is used to estimate theta and phi for given testdocs by using this LDA model
    #     (Model should be already learned with another docs)\n
    #     Testset is used to calculate perplexity by using calculated theta and phi\n
    #     loop: number of learning loop by trainset\n
    #     testRate: ratio of testset to trainset \n
    #               nTestWord_d = int(nWord_d * testRatio)\n
    #     """

    #     self.updateAlphaAndBeta()

    #     # init trainset
    #     trainset = []
    #     nTestWord = 0
    #     for document in testdocs:
    #         l0 = []
    #         if(not self.arrangeInSentece):
    #             for sentence in document:
    #                 for word in sentence:
    #                     if(word in self.words):
    #                         l0.append(self.words.index(word))
    #             trainset.append(l0)
    #         else:
    #             for sentence in document:
    #                 l0 = []
    #                 for word in sentence:
    #                     if(word in self.words):
    #                         l0.append(self.words.index(word))
    #                 trainset.append(l0)

    #     # init testset by spliting trainset
    #     testset = []
    #     for document in trainset:
    #         l1 = []
    #         i0 = int(len(document) * testRatio)
    #         if(i0 > 0):
    #             for i1 in range(i0):
    #                 l1.append(document.pop(random.randrange(len(document))))
    #         nTestWord += i0
    #         testset.append(l1)
    #     testD = len(trainset)

    #     # init counts that will be sampled by trainset
    #     nWord_d = []
    #     for document in trainset:
    #         nWord_d.append(len(document))
    #     nWord_dk = np.zeros((testD, self.K))
    #     z = []
    #     for document in trainset:
    #         z.append([0] * len(document))

    #     # sampling z for trainset
    #     for d, document in enumerate(trainset):
    #         for n, v in enumerate(document):
    #             p0 = nWord_dk[d] + self.alpha
    #             p1 = self.nWord_kv[:, v] + self.beta[v]
    #             p2 = self.nWord_k + self.beta.sum()
    #             p = p0 * p1 / p2
    #             p = p / p.sum()
    #             z0 = np.random.multinomial(1, p).argmax()
    #             z[d][n] = z0
    #             nWord_dk[d, z0] += 1
    #     for lp in range(loop):
    #         LDA_util.updateZ_dk(trainset,
    #                             self.alpha, self.beta,
    #                             nWord_dk, self.nWord_kv, self.nWord_k,
    #                             z)

    #     # calc perplexity with testset
    #     p = 0
    #     for d, document in enumerate(testset):
    #         for n, v in enumerate(document):
    #             p0 = nWord_dk[d] + self.alpha
    #             p1 = nWord_d[d] + self.alpha.sum()
    #             p2 = self.nWord_kv[:, v] + self.beta[v]
    #             p3 = self.nWord_k + self.beta.sum()
    #             p += np.log((p0 / p1 * p2 / p3).sum())
    #     perplexity = math.exp(-1 * p / nTestWord)

    #     return perplexity

    def calcPerplexity2(self, testdocs, testRatio=0.2):

        self.updateAlphaAndBeta()
        self.calcThetaAndPhi()

        # init trainset
        trainset = []
        for document in testdocs:
            l0 = []
            if(not self.arrangeInSentece):
                for sentence in document:
                    for word in sentence:
                        if(word in self.words):
                            l0.append(self.words.index(word))
                trainset.append(l0)
            else:
                for sentence in document:
                    l0 = []
                    for word in sentence:
                        if(word in self.words):
                            l0.append(self.words.index(word))
                    trainset.append(l0)
        testD = len(trainset)

        # init testset by spliting trainset
        testset = []
        for document in trainset:
            l1 = []
            i0 = int(len(document) * testRatio)
            if(i0 > 0):
                for i1 in range(i0):
                    l1.append(document.pop(random.randrange(len(document))))
            testset.append(l1)

        nWord_d = []
        for document in trainset:
            nWord_d.append(len(document))
        nWord_dk = np.zeros((testD, self.K))
        phi_ = self.phi / self.phi.sum(axis=0, keepdims=True)
        for d, document in enumerate(trainset):
            for n, v in enumerate(document):
                nWord_dk[d] += phi_[:, v]

        # calc perplexity
        p = 0
        nTestWord = 0
        for d, document in enumerate(testset):
            nTestWord += len(document)
            for n, v in enumerate(document):
                p0 = nWord_dk[d] + self.alpha
                p1 = nWord_d[d] + self.alpha.sum()
                p2 = self.nWord_kv[:, v] + self.beta[v]
                p3 = self.nWord_k + self.beta.sum()
                p += np.log((p0 / p1 * p2 / p3).sum())
        perplexity = math.exp(-1 * p / nTestWord)

        return perplexity

    def calcCoherence(self, *, c=10, k=None, w2v=None):
        """
        Calculate ceherence.
        Similarity will be calculated with mimno's method\n
        or with word2vec model if w2v is not None
        """

        if(k is not None):
            if(w2v is None):
                sum0 = 0
                argSorted = np.argsort(self.phi[k, :])[::-1]
                words = argSorted[0:c]
                for i in range(c):
                    for j in range(i + 1, c):
                        sum0 += self.similarity[words[i], words[j]]
                return sum0 / c / (c - 1) * 2
            else:
                sum0 = 0
                argSorted = np.argsort(self.phi[k, :])[::-1]
                words = []
                for i in range(c):
                    if(self.words[argSorted[i]] in w2v.wv.vocab):
                        words.append(self.words[argSorted[i]])
                c = len(words)
                if(c <= 1):
                    return None
                for i in range(c):
                    for j in range(i + 1, c):
                        sum0 += w2v.wv.similarity(words[i], words[j])
                return sum0 / c / (c - 1) * 2
        else:
            sum1 = 0
            n0 = 0
            for k in range(self.K):
                a = self.calcCoherence(c=c, k=k, w2v=w2v)
                if(a is None):
                    n0 += 1
                    continue
                sum1 += a
            if(self.K == n0):
                return None
            return sum1 / (self.K - n0)

    def getTopicName(self, k):
        return "topic{}".format(k)

    def writeModelToFolder(self, pathResultFolder):
        """
        Make files about model under param folder\n
        created files:
            model.pickle    : pickle of this model
            theta.csv       : theta
            phi.csv         : phi and historys
            phi.xlsx        : phi
            z.csv           : z
            alpha-beta.csv  : alpha and beta
        """

        pathResultFolder.mkdir(parents=True, exist_ok=True)
        with open(str(pathResultFolder.joinpath("model.pickle")), mode='wb') as f0:
            pickle.dump(self, f0)
        with open(str(pathResultFolder.joinpath("theta.csv")), "w", encoding="utf_8_sig") as f1:
            writer = csv.writer(f1, lineterminator='\n')
            writer.writerows(self.theta)

        with open(str(pathResultFolder.joinpath("phi.csv")), "w", encoding="utf_8_sig") as f2:
            writer = csv.writer(f2, lineterminator='\n')
            writer.writerow(self.words)
            writer.writerows(self.phi)
            writer.writerow(["freqency"])
            writer.writerow(self.nWord_kv.sum(axis=0))
            writer.writerow(["IDF"])
            writer.writerow(self.IDF)
            writer.writerow(["mean TF-IDF"])
            writer.writerow(self.meanTFIDF)
            writer.writerow("")
            writer.writerow(["TOP 100 words in each topic"])
            writer.writerow([None] + list(range(1, 101)))
            top100 = []
            top100_p = []
            for k in range(self.K):
                top100_k = []
                ppp = []
                words = self.words.copy()
                p0 = self.phi[k, :]
                argSorted = np.argsort(p0)[::-1]
                for index in range(100):
                    top100_k.append(words[argSorted[index]])
                    ppp.append(float(self.phi[k, argSorted[index]]))
                top100.append([self.getTopicName(k)] + top100_k)
                top100_p.append([None] + ppp)
            writer.writerows(top100)
            writer.writerows(top100_p)
            writer.writerow("")
            writer.writerow(["countLearned", self.countLearned])
            writer.writerow(["random init", self.randomInit])
            writer.writerow("")
            writer.writerow(["loops on recording history"])
            writer.writerow(self.loopsOnRecord)
            writer.writerow(["perplexity history"])
            writer.writerow(self.perplexityHistory)
            if(len(self.testPerplexityHistory) != 0):
                writer.writerow(["test perplexity history"])
                writer.writerow(self.testPerplexityHistory)
            if(len(self.ceherenceHistory) != 0):
                writer.writerow(["coherence history"])
                writer.writerow(self.ceherenceHistory)
            # if(len(self.ceherenceHistory2) != 0):
            #     writer.writerow(["coherence history with word2vec from this corpus"])
            #     writer.writerow(self.ceherenceHistory2)

        # def appendRow(ws1, row):
        #     ws1.append(list(row))
        # def appendRows(ws1, rows):
        #     for row in rows:
        #         appendRow(ws1, row)
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "summary"
        # appendRow(ws, self.words)
        # appendRows(ws, self.phi)
        # appendRow(ws, ["freqency"])
        # appendRow(ws, self.nWord_kv.sum(axis=0))
        # appendRow(ws, "")
        # appendRow(ws, ["TOP 100 words in each topic"])
        # top100 = []
        # top100_p = []
        # for k in range(self.K):
        #     top100_k = []
        #     ppp = []
        #     words = self.words.copy()
        #     p0 = self.phi[k, :]
        #     argSorted = np.argsort(p0)[::-1]
        #     for index in range(100):
        #         top100_k.append(words[argSorted[index]])
        #         ppp.append(float(self.phi[k, argSorted[index]]))
        #     top100.append(top100_k)
        #     top100_p.append(ppp)
        # appendRows(ws, top100)
        # appendRows(ws, top100_p)
        # appendRow(ws, "")
        # appendRow(ws, ["countLearned", self.countLearned])
        # appendRow(ws, "")
        # appendRow(ws, ["loops on recording history"])
        # appendRow(ws, self.loopsOnRecord)
        # appendRow(ws, ["perplexity history"])
        # appendRow(ws, self.perplexityHistory)
        # if(len(self.testPerplexityHistory) != 0):
        #     appendRow(ws, ["test perplexity history"])
        #     appendRow(ws, self.testPerplexityHistory)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "phi"
        # ws.sheet_properties.syncHorizontal = True
        for k in range(self.K):
            ws.cell(row=1, column=2 + k, value="topic{}".format(k)).font = Font(color="dc143c")
            ws.cell(row=2, column=2 + k, value=self.calcCoherence(k=k, w2v=self.w2vModel)).font = Font(color="dc143c")
            ws.cell(row=1, column=6 + self.K + k, value="topic{}".format(k)).font = Font(color="dc143c")
        ws.cell(row=2, column=1, value="coherence").font = Font(color="dc143c")
        ws.cell(row=1, column=self.K + 2, value="count").font = Font(color="dc143c")
        ws.cell(row=1, column=self.K + 3, value="sum(phi)").font = Font(color="dc143c")
        ws.cell(row=2, column=self.K + 5, value="sum(phi)").font = Font(color="dc143c")
        phi = self.phi
        phi2 = self.phi / self.phi.sum(axis=0, keepdims=True)
        for v, word in enumerate(self.words):
            ws.cell(row=4 + v, column=1, value=word)
            ws.cell(row=4 + v, column=self.K + 5, value=word)
            for k in range(self.K):
                ws.cell(row=4 + v, column=2 + k, value=phi[k, v])
                ws.cell(row=4 + v, column=self.K + 6 + k, value=phi2[k, v])
            ws.cell(row=4 + v, column=self.K + 2, value=self.wordcounts[word])
            ws.cell(row=4 + v, column=self.K + 3, value=phi[:, v].sum())
        for k in range(self.K):
            ws.cell(row=2, column=self.K + 6 + k, value=phi2[k, :].sum()).font = Font(color="dc143c")
        p1 = openpyxl.utils.cell.get_column_letter(2) + "4:" + openpyxl.utils.cell.get_column_letter(1 + self.K) + str(self.V + 3)
        data_bar1 = DataBar(cfvo=[FormatObject(type='min'), FormatObject(type='max')], color="00bfff", showValue=None, minLength=None, maxLength=None)
        ws.conditional_formatting.add(p1, Rule(type='dataBar', dataBar=data_bar1))
        p2 = openpyxl.utils.cell.get_column_letter(6 + self.K) + "4:" + openpyxl.utils.cell.get_column_letter(5 + self.K * 2) + str(self.V + 3)
        data_bar2 = DataBar(cfvo=[FormatObject(type='min'), FormatObject(type='max')], color="00bfff", showValue=None, minLength=None, maxLength=None)
        ws.conditional_formatting.add(p2, Rule(type='dataBar', dataBar=data_bar2))
        wb.save(str(pathResultFolder.joinpath("phi.xlsx")))

        with open(str(pathResultFolder.joinpath("z.csv")), "w", encoding="utf_8_sig") as f3:
            writer = csv.writer(f3, lineterminator='\n')
            writer.writerows(self.z)
        with open(str(pathResultFolder.joinpath("alpha-beta.csv")), "w", encoding="utf_8_sig") as f4:
            writer = csv.writer(f4, lineterminator='\n')
            writer.writerow(["alpha"])
            writer.writerow(self.alpha)
            writer.writerow(["beta"])
            writer.writerow(self.beta)
