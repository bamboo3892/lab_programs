import numpy as np
from sklearn.decomposition import PCA
from sklearn.decomposition import NMF
import pickle
import csv
import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBar, FormatObject
from openpyxl.formatting.rule import Rule


def doMatrixDecomp(pathTensor, pathWords, pathResultFolder, nComponent, method="pca"):
    with open(str(pathTensor), 'rb') as f:
        tensor = pickle.load(f)
    with open(str(pathWords), "r", encoding="utf_8_sig") as f:
        words = f.readline().split("\t")

    if(method == "pca"):
        model = PCA(n_components=nComponent)
        t = tensor - np.sum(tensor, axis=0, keepdims=True) / tensor.shape[0]
        # t = tensor - np.sum(tensor, axis=1, keepdims=True) / tensor.shape[1]
        theta = model.fit_transform(t)
    elif(method == "nmf"):
        model = NMF(n_components=nComponent)
        theta = model.fit_transform(tensor)

    LDAlike = type('', (), {})()
    LDAlike.theta = theta
    LDAlike.phi = model.components_
    LDAlike.K = LDAlike.phi.shape[0]
    for k in range(LDAlike.K):
        if(np.max(LDAlike.phi[k, :]) < -np.min(LDAlike.phi[k, :])):
            LDAlike.phi[k, :] *= -1
    LDAlike.V = tensor.shape[1]
    LDAlike.words = words

    def getTopicName(i):
        return f"Topic{i}"
    LDAlike.getTopicName = getTopicName
    writeModelToFolder(LDAlike, pathResultFolder)


def writeModelToFolder(LDAlike, pathResultFolder):
    """
    Make files about model under param folder\n
    created files:
        theta.csv       : theta
        phi.csv         : phi and historys
        phi.xlsx        : phi
    """

    pathResultFolder.mkdir(parents=True, exist_ok=True)
    with open(str(pathResultFolder.joinpath("theta.csv")), "w", encoding="utf_8_sig") as f1:
        writer = csv.writer(f1, lineterminator='\n')
        writer.writerows(LDAlike.theta)

    with open(str(pathResultFolder.joinpath("phi.csv")), "w", encoding="utf_8_sig") as f2:
        writer = csv.writer(f2, lineterminator='\n')
        writer.writerow(LDAlike.words)
        writer.writerows(LDAlike.phi)
        writer.writerow("")
        writer.writerow(["TOP 100 words in each topic"])
        writer.writerow([None] + list(range(1, 101)))
        top100 = []
        top100_p = []
        for k in range(LDAlike.K):
            top100_k = []
            ppp = []
            words = LDAlike.words.copy()
            p0 = LDAlike.phi[k, :]
            argSorted = np.argsort(p0)[::-1]
            for index in range(min([100, LDAlike.V])):
                top100_k.append(words[argSorted[index]])
                ppp.append(float(LDAlike.phi[k, argSorted[index]]))
            top100.append([LDAlike.getTopicName(k)] + top100_k)
            top100_p.append([None] + ppp)
        writer.writerows(top100)
        writer.writerows(top100_p)
        writer.writerow("")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "phi"
    # ws.sheet_properties.syncHorizontal = True
    for k in range(LDAlike.K):
        ws.cell(row=1, column=2 + k, value="topic{}".format(k)).font = Font(color="dc143c")
        # ws.cell(row=2, column=2 + k, value=LDAlike.calcCoherence(k=k, w2v=LDAlike.w2vModel)).font = Font(color="dc143c")
        ws.cell(row=1, column=6 + LDAlike.K + k, value="topic{}".format(k)).font = Font(color="dc143c")
    ws.cell(row=2, column=1, value="coherence").font = Font(color="dc143c")
    ws.cell(row=1, column=LDAlike.K + 2, value="count").font = Font(color="dc143c")
    ws.cell(row=1, column=LDAlike.K + 3, value="sum(phi)").font = Font(color="dc143c")
    ws.cell(row=2, column=LDAlike.K + 5, value="sum(phi)").font = Font(color="dc143c")
    phi = LDAlike.phi
    phi2 = LDAlike.phi / LDAlike.phi.sum(axis=0, keepdims=True)
    for v, word in enumerate(LDAlike.words):
        ws.cell(row=4 + v, column=1, value=word)
        ws.cell(row=4 + v, column=LDAlike.K + 5, value=word)
        for k in range(LDAlike.K):
            ws.cell(row=4 + v, column=2 + k, value=phi[k, v])
            ws.cell(row=4 + v, column=LDAlike.K + 6 + k, value=phi2[k, v])
        # ws.cell(row=4 + v, column=LDAlike.K + 2, value=LDAlike.wordcounts[word])
        ws.cell(row=4 + v, column=LDAlike.K + 3, value=phi[:, v].sum())
    for k in range(LDAlike.K):
        ws.cell(row=2, column=LDAlike.K + 6 + k, value=phi2[k, :].sum()).font = Font(color="dc143c")
    p1 = openpyxl.utils.cell.get_column_letter(2) + "4:" + openpyxl.utils.cell.get_column_letter(1 + LDAlike.K) + str(LDAlike.V + 3)
    data_bar1 = DataBar(cfvo=[FormatObject(type='min'), FormatObject(type='max')], color="00bfff", showValue=None, minLength=None, maxLength=None)
    ws.conditional_formatting.add(p1, Rule(type='dataBar', dataBar=data_bar1))
    p2 = openpyxl.utils.cell.get_column_letter(6 + LDAlike.K) + "4:" + openpyxl.utils.cell.get_column_letter(5 + LDAlike.K * 2) + str(LDAlike.V + 3)
    data_bar2 = DataBar(cfvo=[FormatObject(type='min'), FormatObject(type='max')], color="00bfff", showValue=None, minLength=None, maxLength=None)
    ws.conditional_formatting.add(p2, Rule(type='dataBar', dataBar=data_bar2))
    wb.save(str(pathResultFolder.joinpath("phi.xlsx")))
