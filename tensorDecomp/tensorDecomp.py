import torch
from torchvision import datasets, transforms
from torch.autograd import Variable
import torch.nn as nn
import torch.optim as optim
import numpy as np
import matplotlib.pyplot as plt
import pickle
import csv
import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBar, FormatObject
from openpyxl.formatting.rule import Rule
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix
import sktensor
import tensorly
from tensorly.decomposition import tucker
from tensorly.decomposition import non_negative_tucker
import logging


def doCPDecomp_torch(pathTensor3, pathWords, pathMIDs, pathResultFolder):
    rank = 10
    fromPickle = False

    with open(str(pathWords), "r", encoding="utf_8_sig") as f:
        words = f.readline().split("\t")
    with open(str(pathMIDs), "r", encoding="utf_8_sig") as f:
        mIDs = f.readline().split("\t")

    if(not fromPickle):
        with open(str(pathTensor3), 'rb') as f:
            tensor = pickle.load(f)
        if(pathResultFolder.joinpath("loss.csv").exists()):
            pathResultFolder.joinpath("loss.csv").unlink()

        valid = np.logical_not(np.any(np.sum(tensor[:, :, :], axis=1) == 0, axis=1))
        tensor = tensor[valid, :, :]
        mIDs = np.array(mIDs)[valid].tolist()
        with open(str(pathResultFolder.joinpath("mIDs.dat")), "w", encoding="utf_8_sig") as f:
            f.write("\t".join(map(str, mIDs)))

        model = Model(tensor.shape, rank)
        optimizer = optim.Adagrad(model.parameters(), lr=0.1, lr_decay=0, weight_decay=0)

        X = Variable(torch.from_numpy(tensor))
        last_loss = []
        dLoss = None
        batch_idx = -1
        while(True):
            batch_idx += 1

            optimizer.zero_grad()
            output = model.forward()
            loss_out = my_mseloss(X, output)
            loss_out.backward()
            optimizer.step()
            loss = loss_out.item()
            last_loss.append(loss)

            if(len(last_loss) >= 10):
                dLoss = (last_loss[-10] - loss) / loss
                if(dLoss <= 0.0001):
                    print(f"\rindex : {batch_idx}, Loss: {loss}, dLoss: {dLoss}")
                    break
            print(f"\rindex : {batch_idx}, Loss: {loss}, dLoss: {dLoss}")

            if(batch_idx % 10 == 0):
                with open(str(pathResultFolder.joinpath("loss.csv")), "a", encoding="utf_8_sig") as f:
                    f.write(f"{batch_idx}, {loss}, {dLoss}\n")
            if(batch_idx % 100 == 0):
                with open(str(pathResultFolder.joinpath("tmp", f"model_{batch_idx}.pickle")), 'wb') as f:
                    pickle.dump(model, f)

        with open(str(pathResultFolder.joinpath("model.pickle")), 'wb') as f:
            pickle.dump(model, f)

    else:
        with open(str(pathResultFolder.joinpath("model.pickle")), 'rb') as f:
            model = pickle.load(f)

    LDAlike = type('', (), {})()
    LDAlike.theta = model.U.data.numpy().T
    LDAlike.phi = np.abs(model.V.data.numpy())
    LDAlike.K = rank
    LDAlike.V = len(words)
    LDAlike.words = words

    def getTopicName(i):
        return "Topic{}".format(i)
    LDAlike.getTopicName = getTopicName

    writeModelToFolder(LDAlike, pathResultFolder)
    with open(str(pathResultFolder.joinpath("W.csv")), "w", encoding="utf_8_sig") as f1:
        writer = csv.writer(f1, lineterminator='\n')
        writer.writerows(model.W.data.numpy())


def doTensorDecomp_sktensor(pathTensor3, pathWords, pathMIDs, pathResultFolder, method):

    logging.basicConfig(level=logging.DEBUG)

    with open(str(pathWords), "r", encoding="utf_8_sig") as f:
        words = f.readline().split("\t")
    with open(str(pathMIDs), "r", encoding="utf_8_sig") as f:
        mIDs = f.readline().split("\t")
    with open(str(pathTensor3), 'rb') as f:
        tensor = pickle.load(f)

    X = sktensor.dtensor(tensor)
    if(method == "cp_als"):
        P, fit, itr = sktensor.cp_als(X, rank=5, init='random', max_iter=500)
        U = P.U
        with open(str(pathResultFolder.joinpath("P.pickle")), 'wb') as f:
            pickle.dump(P, f)
    elif(method == "hooi"):
        core, U = sktensor.tucker_hooi(X, rank=[10, 5, 4], init='random', maxIter=10)
        with open(str(pathResultFolder.joinpath("core.pickle")), 'wb') as f:
            pickle.dump(core, f)
        with open(str(pathResultFolder.joinpath("U.pickle")), 'wb') as f:
            pickle.dump(U, f)

    LDAlike = type('', (), {})()
    LDAlike.theta = np.array(U[0])
    LDAlike.phi = np.array(U[1].T)
    LDAlike.K = LDAlike.phi.shape[0]
    LDAlike.V = len(words)
    LDAlike.words = words

    def getTopicName(i):
        return "Topic{}".format(i)
    LDAlike.getTopicName = getTopicName

    writeModelToFolder(LDAlike, pathResultFolder)
    with open(str(pathResultFolder.joinpath("W.csv")), "w", encoding="utf_8_sig") as f1:
        writer = csv.writer(f1, lineterminator='\n')
        writer.writerows(U[2].T)


def doTensorDecomp_tensorly(pathTensor3, pathWords, pathResultFolder, nComponent, nonnegative):

    fromPickle = False
    pathResultFolder.mkdir(parents=True, exist_ok=True)

    with open(str(pathWords), "r", encoding="utf_8_sig") as f:
        words = f.readline().split("\t")

    if(not fromPickle):
        with open(str(pathTensor3), 'rb') as f:
            tensor = pickle.load(f)
        tensor = tensor.astype("float32")
        tensor = tensor - np.sum(tensor, axis=0, keepdims=True) / tensor.shape[0]
        # tensor = tensor - np.sum(tensor, axis=1, keepdims=True) / tensor.shape[1]
        X = tensorly.tensor(tensor)

        if(not nonnegative):
            # core, U = tucker(X, rank=[5, 5, 4], init='random', svd="numpy_svd", n_iter_max=1, tol=10e-10, random_state=0)
            core, U = tucker(X, rank=nComponent, init='svd', svd="numpy_svd", n_iter_max=100, tol=10e-10, random_state=0)
        else:
            core, U = non_negative_tucker(X, rank=nComponent, init='random', n_iter_max=100, tol=10e-10, random_state=0)

        with open(str(pathResultFolder.joinpath("core.pickle")), 'wb') as f:
            pickle.dump(core, f)
        with open(str(pathResultFolder.joinpath("U.pickle")), 'wb') as f:
            pickle.dump(U, f)
    else:
        with open(str(pathResultFolder.joinpath("core.pickle")), 'rb') as f:
            core = pickle.load(f)
        with open(str(pathResultFolder.joinpath("U.pickle")), 'rb') as f:
            U = pickle.load(f)

    # 正負補正
    if(not nonnegative):
        for m in range(core.shape[0]):
            # if(np.sum(U[0][:, m] < 0) > U[0].shape[0]):
            if(np.max(np.array(U[0][:, m])) < -np.min(U[0][:, m])):
                U[0][:, m] *= -1
                core[m, :, :] *= -1
        for k in range(core.shape[1]):
            if(np.max(np.array(U[1][:, k])) < -np.min(U[1][:, k])):
                U[1][:, k] *= -1
                core[:, k, :] *= -1

    # LDAlike = type('', (), {})()
    # LDAlike.theta = np.array(U[0])
    # LDAlike.phi = np.array(U[1].T)
    # LDAlike.K = LDAlike.phi.shape[0]
    # LDAlike.V = len(words)
    # LDAlike.words = words

    # def getTopicName(i):
    #     return "Topic{}".format(i)
    # LDAlike.getTopicName = getTopicName

    # writeModelToFolder(LDAlike, pathResultFolder)
    # with open(str(pathResultFolder.joinpath("A3.csv")), "w", encoding="utf_8_sig") as f1:
    #     writer = csv.writer(f1, lineterminator='\n')
    #     writer.writerows(U[2].T)
    # with open(str(pathResultFolder.joinpath("M.csv")), "w", encoding="utf_8_sig") as f1:
    #     writer = csv.writer(f1, lineterminator='\n')
    #     for nnn in range(core.shape[0]):
    #         writer.writerow([f"core {nnn}"])
    #         writer.writerows(core[nnn])
    #         writer.writerow([""])
    #         for n in range(core.shape[1]):
    #             writer.writerow([np.dot(core[nnn, n, :], U[2][i, :]) for i in range(core.shape[2])])
    #         writer.writerow([""])

    B = np.tensordot(core, U[1], axes=(1, 1))
    B = np.transpose(B, (0, 2, 1))
    B = np.tensordot(B, U[2], axes=(2, 1))

    wb = openpyxl.Workbook()
    tmp_ws = wb.get_active_sheet()

    ws = wb.create_sheet("core")
    for i in range(core.shape[0]):
        writeMatrix(ws, core[i], (core.shape[1] + 3) * i + 1, 1,
                    row_names=[f"topic{k+1}" for k in range(nComponent[1])],
                    column_names=[f"attr{s+1}" for s in range(nComponent[2])])

    ws = wb.create_sheet("A^(0)")
    writeMatrix(ws, U[0], 1, 1,
                row_names=[f"docs{d+1}" for d in range(tensor.shape[0])],
                column_names=[f"topic{k+1}" for k in range(nComponent[0])])

    ws = wb.create_sheet("A^(1)")
    writeMatrix(ws, U[1], 1, 1,
                row_names=words,
                column_names=[f"topic{k+1}" for k in range(nComponent[1])])

    ws = wb.create_sheet("A^(2)")
    writeMatrix(ws, U[2], 1, 1,
                row_names=[f"attr{s+1}" for s in range(nComponent[2])],
                column_names=[f"attr{s+1}" for s in range(nComponent[2])])

    ws = wb.create_sheet("A^(1)_sorted")
    writeSortedMatrix(ws, U[1], axis=0, row=1, column=1,
                      row_names=words, column_names=[f"topic{k+1}" for k in range(nComponent[1])],
                      maxwrite=100, order="higher",)

    ws = wb.create_sheet("A^(1)_value_sorted")
    writeSortedMatrix(ws, U[1], axis=0, row=1, column=1,
                      row_names=None, column_names=[f"topic{k+1}" for k in range(nComponent[1])],
                      maxwrite=100, order="higher",)

    for i in range(nComponent[2]):
        ws = wb.create_sheet(f"B_{i+1}_sorted")
        writeSortedMatrix(ws, B[:, :, i].T, axis=0, row=1, column=1,
                          row_names=words, column_names=[f"topic{k+1}" for k in range(nComponent[0])],
                          maxwrite=100, order="higher",)

    wb.remove_sheet(tmp_ws)
    wb.save(pathResultFolder.joinpath("result.xlsx"))


def writeModelToFolder(LDAlike, pathResultFolder):
    """
    Make files about model under param folder\n
    created files:
        theta.csv       : theta
        phi.csv         : phi and historys
        phi.xlsx        : phi
    """

    pathResultFolder.mkdir(parents=True, exist_ok=True)
    with open(str(pathResultFolder.joinpath("theta.csv")), "w", encoding="utf_8_sig") as f1:
        writer = csv.writer(f1, lineterminator='\n')
        writer.writerows(LDAlike.theta)

    with open(str(pathResultFolder.joinpath("phi.csv")), "w", encoding="utf_8_sig") as f2:
        writer = csv.writer(f2, lineterminator='\n')
        writer.writerow(LDAlike.words)
        writer.writerows(LDAlike.phi)
        writer.writerow("")
        writer.writerow(["TOP 100 words in each topic"])
        writer.writerow([None] + list(range(1, 101)))
        top100 = []
        top100_p = []
        for k in range(LDAlike.K):
            top100_k = []
            ppp = []
            words = LDAlike.words.copy()
            p0 = LDAlike.phi[k, :]
            argSorted = np.argsort(p0)[::-1]
            for index in range(100):
                top100_k.append(words[argSorted[index]])
                ppp.append(float(LDAlike.phi[k, argSorted[index]]))
            top100.append([LDAlike.getTopicName(k)] + top100_k)
            top100_p.append([None] + ppp)
        writer.writerows(top100)
        writer.writerows(top100_p)
        writer.writerow("")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "phi"
    # ws.sheet_properties.syncHorizontal = True
    for k in range(LDAlike.K):
        ws.cell(row=1, column=2 + k, value="topic{}".format(k)).font = Font(color="dc143c")
        # ws.cell(row=2, column=2 + k, value=LDAlike.calcCoherence(k=k, w2v=LDAlike.w2vModel)).font = Font(color="dc143c")
        ws.cell(row=1, column=6 + LDAlike.K + k, value="topic{}".format(k)).font = Font(color="dc143c")
    ws.cell(row=2, column=1, value="coherence").font = Font(color="dc143c")
    ws.cell(row=1, column=LDAlike.K + 2, value="count").font = Font(color="dc143c")
    ws.cell(row=1, column=LDAlike.K + 3, value="sum(phi)").font = Font(color="dc143c")
    ws.cell(row=2, column=LDAlike.K + 5, value="sum(phi)").font = Font(color="dc143c")
    phi = LDAlike.phi
    phi2 = LDAlike.phi / LDAlike.phi.sum(axis=0, keepdims=True)
    for v, word in enumerate(LDAlike.words):
        ws.cell(row=4 + v, column=1, value=word)
        ws.cell(row=4 + v, column=LDAlike.K + 5, value=word)
        for k in range(LDAlike.K):
            ws.cell(row=4 + v, column=2 + k, value=phi[k, v])
            ws.cell(row=4 + v, column=LDAlike.K + 6 + k, value=phi2[k, v])
        # ws.cell(row=4 + v, column=LDAlike.K + 2, value=LDAlike.wordcounts[word])
        ws.cell(row=4 + v, column=LDAlike.K + 3, value=phi[:, v].sum())
    for k in range(LDAlike.K):
        ws.cell(row=2, column=LDAlike.K + 6 + k, value=phi2[k, :].sum()).font = Font(color="dc143c")
    p1 = openpyxl.utils.cell.get_column_letter(2) + "4:" + openpyxl.utils.cell.get_column_letter(1 + LDAlike.K) + str(LDAlike.V + 3)
    data_bar1 = DataBar(cfvo=[FormatObject(type='min'), FormatObject(type='max')], color="00bfff", showValue=None, minLength=None, maxLength=None)
    ws.conditional_formatting.add(p1, Rule(type='dataBar', dataBar=data_bar1))
    p2 = openpyxl.utils.cell.get_column_letter(6 + LDAlike.K) + "4:" + openpyxl.utils.cell.get_column_letter(5 + LDAlike.K * 2) + str(LDAlike.V + 3)
    data_bar2 = DataBar(cfvo=[FormatObject(type='min'), FormatObject(type='max')], color="00bfff", showValue=None, minLength=None, maxLength=None)
    ws.conditional_formatting.add(p2, Rule(type='dataBar', dataBar=data_bar2))
    wb.save(str(pathResultFolder.joinpath("phi.xlsx")))


class Model(nn.Module):
    def __init__(self, input_shape, rank):
        """
        input_shape : tuple
        Ex. (3,28,28)
        """
        super(Model, self).__init__()

        l, m, n = input_shape
        self.input_shape = input_shape
        self.rank = rank

        self.U = torch.nn.Parameter(torch.randn(rank, l) / 100., requires_grad=True)
        self.V = torch.nn.Parameter(torch.randn(rank, m) / 100., requires_grad=True)
        self.W = torch.nn.Parameter(torch.randn(rank, n) / 100., requires_grad=True)

    def forward_one_rank(self, u, v, w):
        """
        input
            u : torch.FloatTensor of size l
            v : torch.FloatTensor of size m
            w : torch.FloatTensor of size n
        output
            outputs : torch.FloatTensor of size lxmxn
        """
        l, m, n = self.input_shape
        UV = torch.ger(u, v)
        UV2 = UV.unsqueeze(2).repeat(1, 1, n)
        W2 = w.unsqueeze(0).unsqueeze(1).repeat(l, m, 1)
        outputs = UV2 * W2
        return outputs

    def forward(self):
        l, m, n = self.input_shape
        output = self.forward_one_rank(self.U[0], self.V[0], self.W[0])

        for i in np.arange(1, self.rank):
            one_rank = self.forward_one_rank(self.U[i], self.V[i], self.W[i])
            output = output + one_rank
        return output


def my_mseloss(data, output):
    """
    input
        data : torch.autograd.variable.Variable
        output : torch.autograd.variable.Variable
    output
        mse_loss : torch.autograd.variable.Variable
    """
    mse_loss = (data - output).pow(2).sum()
    return mse_loss
