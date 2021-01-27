import random
import numpy as np
from scipy import stats
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import pickle
import openpyxl
from sklearn.cross_decomposition import CCA
from openpyxl.styles import Font, Color

from utils.openpyxl_util import writeMatrix, writeVector, addColorScaleRules, addBorderToMaxCell, paintCells, writePaintedText, drawImage, addDataBarRules, addComments
from utils.matching_util import nn_matching, nn_matching_by_ps


def CCA_between_thetas(path_theta1, path_theta2, pathTensors, pathResult, matching_type):
    """
    matching_type: "ps" or "cca_score" or "each_measure", "random"
    """
    font_path = '/usr/share/fonts/truetype/takao-gothic/TakaoPGothic.ttf'
    font_prop = FontProperties(fname=font_path)
    plt.rcParams['font.family'] = font_prop.get_name()

    n = 7

    data1 = pd.read_csv(path_theta1)
    data2 = pd.read_csv(path_theta2)

    with open(str(pathTensors), 'rb') as f:
        tensors = pickle.load(f)
    measurementKeysHC = tensors["measurement_keys"]
    habitKeysHC = tensors["habit_keys"]
    tensors_before = pd.DataFrame({key: tensors[key] for key in measurementKeysHC})
    tensors_after = pd.DataFrame({key: tensors[key + "_after"] for key in measurementKeysHC})
    tensors_outcome = tensors_after - tensors_before
    aaa = ~tensors_outcome.isnull().all(axis=1)
    tensors_before = tensors_before[aaa]
    tensors_after = tensors_after[aaa]
    tensors_outcome = tensors_outcome[aaa]

    pseqs, idx1, idx2 = np.intersect1d(data1.iloc[:, 0].values, data2.iloc[:, 0].values, assume_unique=True, return_indices=True)
    pseqs, idx3, idxToPID = np.intersect1d(pseqs, np.array(tensors["ids"])[aaa], assume_unique=True, return_indices=True)
    theta1 = data1.iloc[idx1[idx3], 1:].values  # [n_samples, n_features1]
    theta2 = data2.iloc[idx2[idx3], 1:].values  # [n_samples, n_features2]
    names1 = data1.columns[1:]
    names2 = data2.columns[1:]
    corr = np.corrcoef(theta1.T, theta2.T)
    corr11 = corr[0:theta1.shape[1], 0:theta1.shape[1]]  # [n_features1, n_features1]
    corr12 = corr[0:theta1.shape[1], theta1.shape[1]:]  # [n_features1, n_features2]
    corr22 = corr[theta1.shape[1]:, theta1.shape[1]:]  # [n_features2, n_features2]
    type_names = [f"type{t+1}" for t in range(n)]
    colors = [Color('FF0000'), Color('FFFFFF'), Color('0000FF')]

    cca = CCA(n_components=n, scale=True)
    cca.fit(theta1, theta2)

    wb = openpyxl.Workbook()
    ws_config = wb[wb.get_sheet_names()[0]]
    ws_config.title = "config"
    ws = wb.create_sheet("構造相関係数")
    writeMatrix(ws, cca.x_weights_.T, 0 * n + 1, 1, row_names=type_names, column_names=names1, rule="colorscale", rule_color=colors, ruleBoundary=[-1., 0., 1.])
    writeMatrix(ws, cca.y_weights_.T, 1 * n + 3, 1, row_names=type_names, column_names=names2, rule="colorscale", rule_color=colors, ruleBoundary=[-1., 0., 1.])

    ws = wb.create_sheet("正準変数間の相関係数")
    aaa = np.corrcoef(cca.x_scores_.T, cca.y_scores_.T)
    c_corr = np.array([aaa[i, i + n] for i in range(n)])
    writeMatrix(ws, c_corr[:, None], 1, 1, row_names=type_names, rule="databar", ruleBoundary=[0, 1])
    # aaa = np.corrcoef(np.dot(theta1, cca.x_weights_).T, np.dot(theta2, cca.y_weights_).T)
    # c_corr = np.array([aaa[i, i + n] for i in range(n)])
    # writeMatrix(ws, c_corr[:, None], 1, 5, row_names=type_names, rule="databar", ruleBoundary=[0, 1])

    ws = wb.create_sheet("正準負荷量")
    cx = (cca.x_loadings_ ** 2).mean(axis=0)  # 内部で標準化されている
    cy = (cca.y_loadings_ ** 2).mean(axis=0)
    writeMatrix(ws, cx[:, None], 0 * n + 1, 1, row_names=type_names, column_names=["寄与率"])
    writeMatrix(ws, cca.x_loadings_.T, 0 * n + 1, 3, column_names=names1, rule="colorscale", rule_color=colors, ruleBoundary=[-1, 0, 1])
    writeMatrix(ws, cy[:, None], 1 * n + 3, 1, row_names=type_names, column_names=["寄与率"])
    writeMatrix(ws, cca.y_loadings_.T, 1 * n + 3, 3, column_names=names2, rule="colorscale", rule_color=colors, ruleBoundary=[-1, 0, 1])

    ws = wb.create_sheet("交差負荷量")
    x_cross = np.dot(cca.y_weights_.T, corr12.T).T
    y_cross = np.dot(cca.x_weights_.T, corr12).T
    rx = (x_cross ** 2).mean(axis=0)
    ry = (y_cross ** 2).mean(axis=0)
    writeMatrix(ws, rx[:, None], 0 * n + 1, 1, row_names=type_names, column_names=["冗長性係数"])
    writeMatrix(ws, x_cross.T, 0 * n + 1, 3, column_names=names1, rule="colorscale", rule_color=colors, ruleBoundary=[-1, 0, 1])
    writeMatrix(ws, ry[:, None], 1 * n + 3, 1, row_names=type_names, column_names=["冗長性係数"])
    writeMatrix(ws, y_cross.T, 1 * n + 3, 3, column_names=names2, rule="colorscale", rule_color=colors, ruleBoundary=[-1, 0, 1])

    ws = wb.create_sheet("rotation")
    writeMatrix(ws, cca.x_rotations_.T, 0 * n + 1, 1, row_names=type_names, column_names=names1, rule="colorscale", rule_color=colors, ruleBoundary=[-1, 0, 1])
    writeMatrix(ws, cca.y_rotations_.T, 1 * n + 3, 1, row_names=type_names, column_names=names2, rule="colorscale", rule_color=colors, ruleBoundary=[-1, 0, 1])

    # matching
    n_sample = len(cca.x_scores_[:, 0])
    ratio_high = 0.1
    num_high = int(n_sample * ratio_high)
    pathResult.joinpath("figs").mkdir(exist_ok=True, parents=True)
    if(matching_type == "cca_score"):
        min_range = 0.1
        ws_config.append(["ratio_high", ratio_high])
        ws_config.append(["min_range", min_range])

    def _save_ws(ws, t, idx1, idx2):
        fig = plt.figure(figsize=(4.2, 4.2))
        ax = fig.add_subplot(111)
        plt.scatter(cca.x_scores_[:, t], cca.y_scores_[:, t], c="tab:gray", marker=".", s=9)
        plt.scatter(cca.x_scores_[idx1, t], cca.y_scores_[idx1, t], c="tab:orange", marker=".", s=9)
        plt.scatter(cca.x_scores_[idx2, t], cca.y_scores_[idx2, t], c="tab:blue", marker=".", s=9)
        plt.xlim(-8.1, 8.1)
        plt.ylim(-8.1, 8.1)
        plt.xlabel("第1変数グループの正準得点", fontsize=12)
        plt.ylabel("第2変数グループの正準得点", fontsize=12)
        ax.set_xticks([-7.5, -5, -2.5, 0, 2.5, 5, 7.5])
        ax.set_yticks([-7.5, -5, -2.5, 0, 2.5, 5, 7.5])
        plt.subplots_adjust(left=0.17, bottom=0.12)
        plt.savefig(pathResult.joinpath("figs", f"{ws.title}.png"))
        plt.close()

        writePaintedText(ws, 1, f"各指導方法が測定項目にどのような影響を与えたのか(n={len(idx1)})", paintColor="000000")
        drawImage(ws, pathResult.joinpath("figs", f"{ws.title}.png"), 200, 200, "B3")
        row = 15
        writePaintedText(ws, row, "指導記録全体")
        writeMatrix(ws, [tensors_before.mean(), tensors_before.std(), tensors_outcome.mean(), stats.ttest_1samp(tensors_outcome, 0)[1]],
                    row + 1, 1, row_names=["平均", "標準偏差", "改善量平均", "pvalue"], column_names=measurementKeysHC)
        addDataBarRules(ws, row + 5, row + 5, 2, len(measurementKeysHC) + 2, boundary=[0, 1], axis="column")
        row += 7
        writePaintedText(ws, row, "指導前の検診データの比較(保健指導から一年前以内)")
        _print_compare_two_group(ws, row + 1, tensors_before.iloc[idxToPID[idx1]], tensors_before.iloc[idxToPID[idx2]])
        row += 7
        writePaintedText(ws, row, "指導後の検診データの比較(保健指導から一年後以内)")
        _print_compare_two_group(ws, row + 1, tensors_after.iloc[idxToPID[idx1]], tensors_after.iloc[idxToPID[idx2]])
        row += 7
        writePaintedText(ws, row, "指導前後の検査データの変化量の比較")
        _print_compare_two_group(ws, row + 1, tensors_outcome.iloc[idxToPID[idx1]], tensors_outcome.iloc[idxToPID[idx2]])

        # if(matching_type == "cca_score"):
        #     r = stats.ttest_ind(tensors_outcome.iloc[idxToPID[idx1]], tensors_outcome.iloc[idxToPID[idx2]])
        #     for i in range(len(r[1])):
        #         if(r[1][i] < 0.1):
        #             fig = plt.figure(figsize=[6, 4])
        #             plt.boxplot([tensors_before.values[idxToPID[idx1], i], tensors_before.values[idxToPID[idx2], i],
        #                          tensors_after.values[idxToPID[idx1], i], tensors_after.values[idxToPID[idx2], i]],
        #                         labels=["指導あり", "他の指導", "指導あり", "他の指導"])
        #             plt.ylabel(tensors["measurement_keys"][i])
        #             fig.text(0.31, 0.02, "指導前", ha="center")
        #             fig.text(0.705, 0.02, "指導後", ha="center")
        #             plt.grid()
        #             plt.savefig(pathResult.joinpath("figs", f"{ws.title}_{i}.png"))
        #             plt.close()

    def _print_compare_two_group(ws, row, values1, values2):
        r = stats.ttest_ind(values1, values2)
        writeMatrix(ws, [values1.mean()], row, 1, row_names=["指導あり"], column_names=measurementKeysHC)
        writeMatrix(ws, [values2.mean()], row + 2, 1, row_names=["他の指導"])
        writeMatrix(ws, [values1.mean() - values2.mean()], row + 3, 1, row_names=["diff"])
        writeMatrix(ws, [r[1]], row + 4, 1, row_names=["pvalue"], rule="databar", ruleBoundary=[0, 1])
        # addComments(ws, [values1.std()], row + 1, 2)
        # addComments(ws, [values2.std()], row + 2, 2)
        # addComments(ws, [(values1.values - values2.values).std(axis=0)], row + 3, 2)

    def _save_ws_each_measure(ws, t, idx_high_group, idx_middle_group):
        # calc matrix
        n_samples = np.zeros([tensors_before.shape[1]])
        before = np.zeros([tensors_before.shape[1], 4])
        after = np.zeros([tensors_before.shape[1], 4])
        outcome = np.zeros([tensors_before.shape[1], 4])
        for i in range(tensors_before.shape[1]):
            idx1, idx2 = nn_matching(tensors_before.values[idx_high_group, i], tensors_before.values[idx_middle_group, i],
                                     min_range=0)
            #  min_range=tensors_before.values[:, i].std() / 100)
            idx1 = idx_high_group[idx1]
            idx2 = idx_middle_group[idx2]
            n_samples[i] = len(idx1)
            before[i, 0] = tensors_before.values[idx1, i].mean()
            before[i, 1] = tensors_before.values[idx2, i].mean()
            before[i, 2] = tensors_before.values[idx1, i].mean() - tensors_before.values[idx2, i].mean()
            before[i, 3] = stats.ttest_ind(tensors_before.values[idx1, i], tensors_before.values[idx2, i])[1]
            after[i, 0] = tensors_after.values[idx1, i].mean()
            after[i, 1] = tensors_after.values[idx2, i].mean()
            after[i, 2] = tensors_after.values[idx1, i].mean() - tensors_after.values[idx2, i].mean()
            after[i, 3] = stats.ttest_ind(tensors_after.values[idx1, i], tensors_after.values[idx2, i])[1]
            outcome[i, 0] = tensors_outcome.values[idx1, i].mean()
            outcome[i, 1] = tensors_outcome.values[idx2, i].mean()
            outcome[i, 2] = tensors_outcome.values[idx1, i].mean() - tensors_outcome.values[idx2, i].mean()
            outcome[i, 3] = stats.ttest_ind(tensors_outcome.values[idx1, i], tensors_outcome.values[idx2, i])[1]

        writePaintedText(ws, 1, f"各指導方法が測定項目にどのような影響を与えたのか", paintColor="000000")
        writeMatrix(ws, n_samples[None, :], 3, 1, row_names=["n_samples"], column_names=measurementKeysHC)
        row = 6
        writePaintedText(ws, row, "指導前の検診データの比較(保健指導から一年前以内)")
        writeMatrix(ws, before.T, row + 1, 1, row_names=["指導あり", "他の指導", "diff", "pvalue"], column_names=measurementKeysHC)
        addDataBarRules(ws, row + 5, row + 5, 2, len(measurementKeysHC) + 2, boundary=[0, 1], axis="column")
        row += 7
        writePaintedText(ws, row, "指導後の検診データの比較(保健指導から一年後以内)")
        writeMatrix(ws, after.T, row + 1, 1, row_names=["指導あり", "他の指導", "diff", "pvalue"], column_names=measurementKeysHC)
        addDataBarRules(ws, row + 5, row + 5, 2, len(measurementKeysHC) + 2, boundary=[0, 1], axis="column")
        row += 7
        writePaintedText(ws, row, "指導前後の検査データの変化量の比較")
        writeMatrix(ws, outcome.T, row + 1, 1, row_names=["指導あり", "他の指導", "diff", "pvalue"], column_names=measurementKeysHC)
        addDataBarRules(ws, row + 5, row + 5, 2, len(measurementKeysHC) + 2, boundary=[0, 1], axis="column")

    for t in range(n):
        # grouping
        score_xy = cca.x_scores_[:, t] * cca.y_scores_[:, t]
        idx_pos = np.array(np.where((score_xy > 0) & (cca.x_scores_[:, t] > 0))).ravel()  # 第１象限
        idx_high_group = idx_pos[np.argsort(score_xy[idx_pos])[-num_high:]]  # 第１象限でxyが大きいやつ

        idx_neg = np.array(np.where((score_xy > 0) & (cca.x_scores_[:, t] < 0))).ravel()  # 第３象限
        idx_low_group = idx_neg[np.argsort(score_xy[idx_neg])[-num_high:]]  # 第３象限でxyが大きいやつ

        idx_other = np.full_like(score_xy, True, dtype="bool")
        idx_other[idx_high_group] = False
        idx_other[idx_low_group] = False
        idx_middle_group = np.array(np.where(idx_other)).ravel()  # それ以外

        plt.figure(figsize=(16, 12))
        plt.scatter(cca.x_scores_[idx_high_group, t], cca.y_scores_[idx_high_group, t], c="tab:blue", marker=".")
        plt.scatter(cca.x_scores_[idx_low_group, t], cca.y_scores_[idx_low_group, t], c="tab:orange", marker=".")
        plt.scatter(cca.x_scores_[idx_middle_group, t], cca.y_scores_[idx_middle_group, t], c="tab:green", marker=".")
        plt.xlim(-10, 10)
        plt.ylim(-10, 10)
        plt.savefig(pathResult.joinpath(f"scatter_type{t+1}.png"))
        plt.close()

        # matching
        if(matching_type == "ps"):
            idx1, idx2 = nn_matching_by_ps(tensors_before.values[idx_high_group], tensors_before.values[idx_middle_group], min_range=0.00001, model_name="LR")
            _save_ws(wb.create_sheet(f"matching_type{t+1}_pos"), t, idx_high_group[idx1], idx_middle_group[idx2])
            idx1, idx2 = nn_matching_by_ps(tensors_before.values[idx_low_group], tensors_before.values[idx_middle_group], min_range=0.00001, model_name="LR")
            _save_ws(wb.create_sheet(f"matching_type{t+1}_neg"), t, idx_low_group[idx1], idx_middle_group[idx2])
        elif(matching_type == "cca_score"):
            idx1, idx2 = nn_matching(cca.x_scores_[idx_high_group, t], cca.x_scores_[idx_middle_group, t], min_range=min_range)
            _save_ws(wb.create_sheet(f"matching_type{t+1}_pos"), t, idx_high_group[idx1], idx_middle_group[idx2])
            idx1, idx2 = nn_matching(cca.x_scores_[idx_low_group, t], cca.x_scores_[idx_middle_group, t], min_range=min_range)
            _save_ws(wb.create_sheet(f"matching_type{t+1}_neg"), t, idx_low_group[idx1], idx_middle_group[idx2])
        elif(matching_type == "each_measure"):
            _save_ws_each_measure(wb.create_sheet(f"matching_type{t+1}_pos"), t, idx_high_group, idx_middle_group)
            _save_ws_each_measure(wb.create_sheet(f"matching_type{t+1}_neg"), t, idx_low_group, idx_middle_group)
        elif(matching_type == "random"):
            idx2 = random.sample(idx_middle_group.tolist(), len(idx_high_group))
            _save_ws(wb.create_sheet(f"matching_type{t+1}_pos"), t, idx_high_group, idx2)
            idx2 = random.sample(idx_middle_group.tolist(), len(idx_low_group))
            _save_ws(wb.create_sheet(f"matching_type{t+1}_neg"), t, idx_low_group, idx2)

    pathResult.mkdir(exist_ok=True, parents=True)
    wb.save(pathResult.joinpath("cca_result.xlsx"))
