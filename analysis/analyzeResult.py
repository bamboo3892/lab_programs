# -*- coding: utf-8 -*-
"""
Created on DEC 13 2018\n
@author: takeda masaki
"""

import os
import json
import sys
import csv
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Color
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
import warnings
from matplotlib.font_manager import FontProperties
import pickle
import sklearn
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import plotly.graph_objs as go
import plotly.offline as po
import shutil


# warnings.filterwarnings("ignore")

# logger

# loaded object
LDA_MODEL = None
DICT_REVIEWS = None
DF_ANOVA = None

# constants
THETA_KEY = "theta_d"
MIN_DIRECTION_TIMES = 4

featureLabels = ["age", "pNum", "pTotalTime", "pAveTime", "startWaist", "startWeight", "cNum", "nLabel"]  # term, other score
HCLabel = ["waist", "weight", "BMI", "HbA1c", "DBP", "SBP", "HDL", "LDL", "gammaGT", "MetaboDecision", "MD_estimeted"]
abnormalValues = [85, None, None, 5.6, 85, 130, 40, 150, 100, None, None]
abnormalFlags = [True, None, None, True, True, True, False, True, True, None, None]
outcomeLabels = ["dWaist", "dWeight"]
outcomeLabels2 = ["dWaist", "dWeight", "dBMI", "dHbA1c", "dDBP", "dSBP", "dHDL", "dLDL", "dGammaGT", "dMetaboDecision", "dMD_estimated"]
csvLabels = ["腹囲", "体重", "ＢＭＩ", "ＨｂＡ１ｃ（ＮＧＳＰ）", "拡張期血圧", "収縮期血圧", "ＨＤＬコレステロール", "ＬＤＬコレステロール", "γ－ＧＴ（γ－ＧＴＰ）", "メタボリックシンドローム判定"]
thetaLabels = None

DICT0 = {"1200": "モニタリング", "1210": "正しい病識", "1100": "適正受診", "1220": "適正服薬", "1230": "疾病の自己管理",
         "1300": "活動量",
         "1400": "飲料習慣", "1500": "間食習慣",
         "1720": "栄養バランス", "1730": "飽和脂肪酸の量", "1740": "主菜バランス", "1750": "野菜,海藻,きのこ類の量", "1760": "塩分量",
         "1600": "飲酒習慣",
         "1700": "夕食の時間と量", "1710": "食事リズム",
         "1800": "喫煙",
         "1900": "疲労回復,ストレス解消",
         "2100": "（就寝前の）歯磨き", "2200": "フッ素入り歯磨き剤", "2300": "歯間ブラシ・フロス", "2400": "ゆっくり噛む・食べる", "2500": "歯磨き指導", "2600": "定期健診（歯科）",
         "2000": "その他"}
DICT1 = {"1100": "適正受診",
         "1200": "モニタリング",
         "1300": "運動",
         "1400": "飲料",
         "1500": "間食",
         "1600": "飲酒",
         "1700": "生活リズム", "1710": "生活リズム"}

relation_habit_direction = {"３０分以上の運動習慣": "活動量",
                            "歩行又は身体活動": "活動量",
                            "歩行速度": "活動量",
                            "食べ方２（就寝前）": "夕食の時間と量",
                            "食べ方３（夜食／間食）": "間食習慣",
                            "食習慣": "食事リズム",
                            "食べ方１（早食い等）": "夕食の時間と量",
                            "飲酒": "飲酒習慣",
                            "飲酒量": "飲酒習慣",
                            "喫煙": "喫煙",
                            "睡眠": "疲労回復,ストレス解消"}

# variables
targets = {}
advisors = {}
features = []
before = []
after = []
outcomes = []
outcomes2 = []
outcomes_fixed = []
theta1 = []  # numTopic * numTarget (first direction)
thetas = [[] for i in range(MIN_DIRECTION_TIMES)]  # MIN_DIRECTION_TIMES * numTopic * numTarget (whole direction)
# theta2 = []
# theta3 = []
# theta4 = []
fkr = []
idxFull = []  # indexes
idxCNum0 = []  # indexes advised by skilled advisor
idxCNum1 = []  # non-skilled
idxD0 = []  # indexes with good outcome
idxD1 = []  # bad
idxHCValid = []  # indexes with valid health check data
idxMedicineTake = []  # is taking medicine or not
idxMetaboD = [[[] for j in range(3)] for i in range(3)]  # indexes to each metabo decision ([before][after])
idxGammaGTAbnorm = []  # indexes to targets with abnormal GammaGT (<100)
idxToTarget = []  # indexes to target data
idxToMID = []
idxToCID = []
idxToSDATE = []
idxToEDATE = []
idxToHCDATE1 = []  # health check date before direction
idxToHCDATE2 = []  # health check date after direction
idxAbnormals = []
aveCNum = None
aveWaistD = None

numTarget = None

nowRow = [1]
fill1 = openpyxl.styles.PatternFill(fill_type='solid', fgColor="ff7f50")
fill2 = openpyxl.styles.PatternFill(fill_type='solid', fgColor="ffdab9")
fill3 = openpyxl.styles.PatternFill(fill_type='solid', fgColor="0040FF")
fill4 = openpyxl.styles.PatternFill(fill_type='solid', fgColor="0080FF")
fill5 = openpyxl.styles.PatternFill(fill_type='solid', fgColor="0B610B")

temp = None


def analyze(pathLDAResultFolder, folderToWriteResultIn, pathTemplate, pathHealthCheck, pathAnova=None):

    font_path = '/usr/share/fonts/truetype/takao-gothic/TakaoPGothic.ttf'
    font_prop = FontProperties(fname=font_path)
    plt.rcParams['font.family'] = font_prop.get_name()
    plt.style.use('ggplot')

    print("Start analyzing")
    loadFiles(pathLDAResultFolder, pathAnova)

    # with open(str(pathLDAResultFolder.joinpath("phi_sim.csv")), "w", encoding="utf_8_sig") as f:
    #     writer = csv.writer(f, lineterminator='\n')
    #     l = [LDA_MODEL.getTopicName(k) for k in range(LDA_MODEL.K)]
    #     writer.writerow(l)
    #     for k1 in range(LDA_MODEL.K):
    #         l = []
    #         for k2 in range(LDA_MODEL.K):
    #             l.append(np.dot(LDA_MODEL.phi[k1], LDA_MODEL.phi[k2]) / np.linalg.norm(LDA_MODEL.phi[k1]) / np.linalg.norm(LDA_MODEL.phi[k2]))
    #         writer.writerow(l)

    # with open(str(pathLDAResultFolder.joinpath("phi_pca.csv")), "w", encoding="utf_8_sig") as f2:
    #     writer = csv.writer(f2, lineterminator='\n')
    #     l = [LDA_MODEL.getTopicName(k) for k in range(LDA_MODEL.K)]
    #     writer.writerow(l)
    #     pca = PCA(n_components=2)
    #     pca.fit(LDA_MODEL.phi.T)
    #     writer.writerows(pca.components_)

    print("Processing LDA result")
    initVariables(pathHealthCheck)

    # saveDocVectors(pathLDAResultFolder.joinpath("docvectors.csv"))

    # # save fig hist2d
    # for i in range(len(outcomeLabels2)):
    #     # plt.figure(figsize=(8, 8))
    #     # plt.hist(before[i][idxHCValid].tolist(), bins=20)
    #     # plt.title(HCLabel[i])
    #     # plt.savefig(str(folderToWriteResultIn.joinpath("aaa", HCLabel[i] + '.png')))
    #     # plt.close()

    #     # plt.figure(figsize=(8, 8))
    #     # plt.hist2d(before[i][idxHCValid], outcomes2[i][idxHCValid], bins=20)
    #     # plt.title("before-outcome ({})".format(HCLabel[i]))
    #     # plt.savefig(str(folderToWriteResultIn.joinpath("aaa", HCLabel[i] + '.png')))
    #     # plt.close()

    #     # plt.figure(figsize=(8, 8))
    #     # plt.scatter(before[i][idxHCValid], outcomes2[i][idxHCValid])
    #     # plt.title("before-outcome ({})".format(HCLabel[i]))
    #     # plt.savefig(str(folderToWriteResultIn.joinpath("aaa", HCLabel[i] + '.png')))
    #     # plt.close()

    #     # plt.figure(figsize=(8, 8))
    #     # plt.scatter(before[i][idxHCValid], outcomes_fixed[i][idxHCValid])
    #     # plt.title("before-outcome_fixed ({})".format(HCLabel[i]))
    #     # plt.savefig(str(folderToWriteResultIn.joinpath("aaa", HCLabel[i] + '.png')))
    #     # plt.close()

    # # idx = np.logical_or(theta1[3] != 0, theta1[12] != 0)
    # idx = idxAbnormals[4]
    # plt.figure(figsize=(8, 8))
    # # plt.scatter(theta2[3][idx], theta2[12][idx])
    # plt.scatter(theta2[3, idxHCValid][idx], theta2[12, idxHCValid][idx])
    # plt.title("scatter_{}-{}".format(thetaLabels[3], thetaLabels[12]))
    # plt.savefig(str(folderToWriteResultIn.joinpath("scatter_{}-{}".format(thetaLabels[3], thetaLabels[12]) + '.png')))
    # plt.close()

    # # save fig confusion matrix
    # for i in range(len(outcomeLabels2)):
    #     plt.figure(figsize=(8, 8))
    #     plt.hist2d(before[i][idxHCValid], after[i][idxHCValid], bins=20, normed=True)
    #     plt.title(HCLabel[i])
    #     plt.savefig(str(folderToWriteResultIn.joinpath("bbb", HCLabel[i] + '_cm_normed.png')))
    #     plt.close()

    # # metabo dicision
    # plt.figure(figsize=(8, 8))
    # h, xedges, yedges, image = plt.hist2d(before[9][idxHCValid], after[9][idxHCValid], bins=3)
    # plt.title(HCLabel[9])
    # plt.savefig(str(folderToWriteResultIn.joinpath("aaa", HCLabel[9] + '_cm.png')))
    # plt.close()
    # with open(str(pathLDAResultFolder.joinpath("metabo_ba.csv")), "w", encoding="utf_8_sig") as f2:
    #     writer = csv.writer(f2, lineterminator='\n')
    #     writer.writerows(h)
    #     writer.writerow(xedges)
    #     writer.writerow(yedges)

    # for i in range(len(HCLabel)):
    #     for j in range(len(thetaLabels)):
    #         name = "{}-{}".format(HCLabel[i], thetaLabels[j])
    #         indexes = theta1[j, idxHCValid] != 0
    #         if(idxAbnormals[i] is None):
    #             y = outcomes2[i, idxHCValid][indexes]
    #             x = theta1[j, idxHCValid][indexes]
    #         else:
    #             idx = np.logical_and(indexes, idxAbnormals[i])
    #             y = outcomes2[i, idxHCValid][idx]
    #             x = theta1[j, idxHCValid][idx]
    #         plt.figure(figsize=(8, 8))
    #         plt.scatter(x, y)
    #         plt.title(name)
    #         plt.savefig(str(folderToWriteResultIn.joinpath("direction-outcome_scatter", name + '.png')))
    #         plt.close()

    # for i in range(len(HCLabel)):
    #     for j in range(len(thetaLabels)):
    #         name = "{}-{}".format(HCLabel[i], thetaLabels[j])
    #         indexes = theta1[j, idxHCValid] != 0
    #         if(idxAbnormals[i] is None):
    #             y = outcomes_fixed[i, idxHCValid][indexes]
    #             x = theta1[j, idxHCValid][indexes]
    #         else:
    #             idx = np.logical_and(indexes, idxAbnormals[i])
    #             y = outcomes_fixed[i, idxHCValid][idx]
    #             x = theta1[j, idxHCValid][idx]
    #         plt.figure(figsize=(8, 8))
    #         plt.scatter(x, y)
    #         plt.title(name)
    #         plt.savefig(str(folderToWriteResultIn.joinpath("direction-outcome_fixed_scatter", name + '.png')))
    #         plt.close()

    # # save 3d figs
    # i = 4
    # j = 3
    # k = 4
    # trace = go.Scatter3d(
    #     x=before[i, idxHCValid],
    #     y=theta1[j, idxHCValid],
    #     # z=outcomes2[k, idxHCValid],
    #     z=outcomes_fixed[k, idxHCValid],
    #     mode='markers',
    #     marker=dict(
    #         color='rgb(127, 127, 127)',
    #         size=12,
    #         symbol='circle',
    #         line=dict(
    #             color='rgb(204, 204, 204)',
    #             width=1
    #         ),
    #         opacity=0.9
    #     )
    # )
    # data = [trace]
    # layout = go.Layout(
    #     xaxis=dict(title=HCLabel[i]),
    #     yaxis=dict(title=thetaLabels[j]),
    #     # zaxis=dict(title=outcomeLabes2[k]),
    #     margin=dict(
    #         l=0,
    #         r=0,
    #         b=0,
    #         t=0
    #     ),
    #     legend=dict(
    #         x=0
    #     )
    # )
    # fig = go.Figure(data=data, layout=layout)
    # po.plot(fig, filename=str(pathLDAResultFolder.joinpath("scatter_{}-{}-{}_fixed.html".format(HCLabel[i], thetaLabels[j], outcomeLabels2[k]))))

    # for i in range(len(HCLabel)):
    #     for j in range(len(thetaLabels)):
    #         if(abnormalValues[i] is not None):
    #             name = "{}-{}".format(HCLabel[i], thetaLabels[j])
    #             indexes = theta1[j, idxHCValid] != 0
    #             idx = np.logical_and(indexes, idxAbnormals[i])
    #             with open(str(folderToWriteResultIn.joinpath("text", name + '.json')), "w", encoding="utf_8_sig") as output:
    #                 a0 = []
    #                 for iii in range(len(idx)):
    #                     if(idx[iii]):
    #                         a1 = idxToTarget[idxHCValid[iii]][0]
    #                         a2 = idxToTarget[idxHCValid[iii]][1]
    #                         a3 = idxToTarget[idxHCValid[iii]][2]
    #                         a4 = idxToTarget[idxHCValid[iii]][3]
    #                         d0 = {}
    #                         d0["p_r_tgtset_explan1"] = a1["p_r_tgtset_explan"]
    #                         d0["p_r_tgtset_explan2"] = a2["p_r_tgtset_explan"]
    #                         d0["p_r_tgtset_explan3"] = a3["p_r_tgtset_explan"]
    #                         d0["p_r_tgtset_explan4"] = a4["p_r_tgtset_explan"]
    #                         d0["__before"] = before[i, idxHCValid[iii]]
    #                         d0["__after"] = after[i, idxHCValid[iii]]
    #                         d0["__direction1"] = theta1[j, idxHCValid[iii]]
    #                         d0["__direction2"] = theta2[j, idxHCValid[iii]]
    #                         d0["__direction3"] = theta3[j, idxHCValid[iii]]
    #                         d0["__direction4"] = theta4[j, idxHCValid[iii]]
    #                         d0["__outcome"] = outcomes2[i, idxHCValid[iii]]
    #                         a0.append(d0)
    #                 json.dump(a0, output, ensure_ascii=False, indent=2)

    # for j in range(len(thetaLabels)):
    #     name = "{}".format(thetaLabels[j])
    #     indexes = np.argsort(theta1[j])[::-1]
    #     with open(str(folderToWriteResultIn.joinpath("text_direction", name + '.json')), "w", encoding="utf_8_sig") as output:
    #         a0 = []
    #         for i in range(10):
    #             a1 = idxToTarget[indexes[i]][0]
    #             d0 = {}
    #             d0["p_r_tgtset_explan"] = a1["p_r_tgtset_explan"]
    #             d0["__direction"] = theta1[j, indexes[i]]
    #             a0.append(d0)
    #         json.dump(a0, output, ensure_ascii=False, indent=2)

    # sys.exit(0)

    # write analysis
    wb = openpyxl.load_workbook(str(pathTemplate), read_only=False, keep_vba=True)
    global temp
    temp = folderToWriteResultIn.joinpath(".temp")
    temp.mkdir(exist_ok=True)

    flag13 = True  # stats sheet
    flag0 = False  # feature sheet
    flag1 = False  # direction sheet
    flag2 = True  # direction2 sheet
    flag3 = False  # outcome sheet
    flag4 = False  # outcome2 sheet
    flag5 = True  # hist sheet
    flag11 = False  # PCA
    flag6 = True  # average sheet
    flag10 = True  # average2 shhet
    flag7 = True  # coef sheets
    flag7_2 = False
    flag9 = True  # coef2 sheet
    flag8 = False  # 2d sheets
    flag12 = False  # anova sheet
    flag14 = True  # lee sheet

    if(flag13):
        print("Making stats sheet")
        makeStatsSheet(wb)
    if(flag0):
        print("Making feature sheet")
        makeFeatureSheet(wb)
    if(flag1):
        print("Making direction sheet")
        makeDirectionSheet(wb)
    if(flag2):
        print("Making direction2 sheet")
        makeDirection2Sheet(wb)
    makePhi_simSheet(wb)
    if(flag3):
        print("Making outcome sheet")
        makeOutcomeSheet(wb)
    if(flag4):
        print("Making outcome2 sheet")
        makeOutcome2Sheet(wb)
    if(flag5):
        print("Making hist sheet")
        makeHistSheet(wb)
    if(flag11):
        makePCA_phiSheet(wb)
        makePCA_thetaSheet(wb, folderToWriteResultIn)
    if(flag6):
        print("Making average sheet")
        makeAverageSheet(wb)
    if(flag7):
        print("Making coef sheet")
        makeCoefSheet(wb, flag7_2)
    if(flag9):
        print("Making coef2 sheet")
        makeCoef2Sheet(wb)
    if(flag8):
        print("Making 2d sheet")
        make2dSheet(wb)
    if(flag12 and DF_ANOVA is not None):
        print("Making Anova sheet")
        makeAnovaSheet(wb)
    if(flag14):
        print("Making Lee sheet")
        makeLeeSheet(wb)

    print("Saving analysis")
    wb.save(str(folderToWriteResultIn.joinpath("analysis.xlsm")))
    print("Analyze finished")


def loadFiles(pathLDAResultFolder, pathAnova):
    with open(str(pathLDAResultFolder.joinpath("model.pickle")), mode='rb') as f:
        global LDA_MODEL
        LDA_MODEL = pickle.load(f)

    with open(str(pathLDAResultFolder.joinpath("documents.json")), "r", encoding="utf_8_sig") as f0:
        global DICT_REVIEWS
        DICT_REVIEWS = json.load(f0)

    if(pathAnova is not None):
        global DF_ANOVA
        DF_ANOVA = pd.read_csv(str(pathAnova))
        DF_ANOVA = DF_ANOVA.set_index(DF_ANOVA.columns[0])
        DF_ANOVA = DF_ANOVA.rename(columns={csvLabels[i]: HCLabel[i] for i in range(len(csvLabels))})


def saveDocVectors(path):
    with open(str(path), "w", encoding="utf_8_sig") as f:
        writer = csv.writer(f, lineterminator='\n')
        writer.writerow(["p_seq"] + LDA_MODEL.labelTypes)
        for review in DICT_REVIEWS:
            writer.writerow([review["p_seq"]] + review["theta_d"])


def initVariables(pathHealthCheck):

    global targets
    global advisors
    global features  # numfeature * numTarget
    global before
    global after
    global outcomes
    global outcomes2
    global outcomes_fixed
    global theta1
    global thetas
    # global theta2
    # global theta3
    # global theta4
    global fkr
    global idxFull
    global idxCNum0
    global idxCNum1
    global idxD0
    global idxD1
    global idxHCValid
    global idxMedicineTake
    global idxMetaboD
    global idxGammaGTAbnorm
    global idxToTarget
    global idxToMID
    global idxToCID
    global idxToSDATE
    global idxToEDATE
    global idxToHCDATE1
    global idxToHCDATE2
    global idxAbnormals
    global numTarget
    global thetaLabels
    global aveCNum
    global aveWaistD

    for review in DICT_REVIEWS:
        if(review["m_id"] not in targets.keys()):
            targets[review["m_id"]] = []
        targets[review["m_id"]].append(review)
        if(review["c_id"] not in advisors.keys()):
            advisors[review["c_id"]] = 0
        advisors[review["c_id"]] += 1

    thetaLabels = LDA_MODEL.labelTypes + ["topic{}".format(k) for k in range(len(LDA_MODEL.labelTypes), LDA_MODEL.K)]

    for i in range(len(featureLabels)):
        features.append([])
    for i in range(len(HCLabel) - 1):
        before.append([])
        after.append([])
    for i in range(len(outcomeLabels)):
        outcomes.append([])
    for i in range(len(outcomeLabels2) - 1):
        outcomes2.append([])

    for target in targets.values():
        if(len(target) < MIN_DIRECTION_TIMES):
            continue
        # fukuyaku = False
        # for pn in range(len(target)):
        #     if("p_r_fukuyaku" not in target[pn].keys() or target[pn]["p_r_fukuyaku"] == "2.0"):
        #         fukuyaku = True
        # if(fukuyaku):
        #     continue
        targetSorted = sorted(target, key=lambda x: int(x["p_num"]))
        m_id = targetSorted[0]["m_id"]
        c_id = targetSorted[0]["c_id"]
        bday = datetime.datetime.strptime(targetSorted[0]["m_bday"], '%Y/%m/%d')
        psday = datetime.datetime.strptime(targetSorted[0]["p_s_date"], '%Y-%m-%d')
        peday = datetime.datetime.strptime(targetSorted[len(targetSorted) - 1]["p_s_date"], '%Y-%m-%d')
        features[0].append((psday - bday).days / 365.25)
        features[1].append(len(targetSorted))
        p_time = 0
        for review in targetSorted:
            start = datetime.datetime.strptime(review["p_s_time"], '%H:%M:%S')
            end = datetime.datetime.strptime(review["p_e_time"], '%H:%M:%S')
            p_time += (end - start).seconds / 60
        features[2].append(p_time)
        features[3].append(p_time / len(targetSorted))

        def func0(dict, a):
            b = a + "_gu"
            if(a in dict.keys()):
                return float(dict[a])
            elif(b in dict.keys()):
                return float(dict[b])
            else:
                return None
        startWaist = func0(targetSorted[0], "p_r_waist")
        startWeight = func0(targetSorted[0], "p_r_weight")
        endWaist = func0(targetSorted[len(targetSorted) - 1], "p_r_waist")
        endWeight = func0(targetSorted[len(targetSorted) - 1], "p_r_weight")
        features[4].append(startWaist)
        features[5].append(startWeight)
        outcomes[0].append(endWaist - startWaist)
        outcomes[1].append(endWeight - startWeight)

        t = targetSorted[0][THETA_KEY]
        theta1.append(t / np.sum(t))
        for i in range(MIN_DIRECTION_TIMES):
            t = targetSorted[i][THETA_KEY]
            thetas[i].append(t / np.sum(t))

        l0 = [0] * len(thetaLabels)
        if("p_r_tgtset_explan_seqs_id" in targetSorted[0].keys()):
            l1 = targetSorted[0]["p_r_tgtset_explan_seqs_id"].split(",")
            nLabel = len(l1)
            for l11 in l1:
                if(l11 in DICT0.keys() and DICT0[l11] in thetaLabels):
                    l0[thetaLabels.index(DICT0[l11])] += 1
        fkr.append(l0)
        idx = len(idxFull)
        idxFull.append(idx)
        features[7].append(nLabel)
        idxToTarget.append(targetSorted)
        idxToMID.append(m_id)
        idxToCID.append(c_id)
        idxToSDATE.append(psday)
        idxToEDATE.append(peday)

    # health check
    print("Processing health check data")
    healthChecks = {}
    with open(str(pathHealthCheck), "r", encoding="utf_8_sig") as reader:
        label = reader.readline().split(",")
        checkIdx = [label.index("腹囲"), label.index("ＨｂＡ１ｃ（ＮＧＳＰ）"), label.index("γ－ＧＴ（γ－ＧＴＰ）")]
        idxID = label.index("m_id\n")
        idxDate = label.index("健診受診日")
        csvreader = csv.reader(reader, delimiter=',')
        for row in csvreader:
            f0 = False
            for iii in checkIdx:
                try:
                    float(row[iii])
                except:
                    f0 = True
                    break
            if(f0):
                continue
            mId = row[idxID]
            date = datetime.datetime.strptime(row[idxDate], '4%y%m%d.0')
            date = date.replace(year=date.year - 12)
            row[idxDate] = date
            if(mId not in healthChecks):
                healthChecks[mId] = []
            healthChecks[mId].append(row)
    for idx in idxFull:
        mId = idxToMID[idx]
        psday = idxToSDATE[idx]
        peday = idxToEDATE[idx]
        if(mId in healthChecks):
            i0 = -1
            i1 = -1
            rows = healthChecks[mId]
            for i, row in enumerate(rows):
                if(row[idxDate] <= psday):
                    if(i0 < 0 or rows[i0][idxDate] < row[idxDate]):
                        i0 = i
                elif(peday <= row[idxDate]):
                    if(i1 < 0 or row[idxDate] < rows[i1][idxDate]):
                        i1 = i
            if(i0 is not -1 and i1 is not -1):
                idxHCValid.append(idx)
                idxToHCDATE1.append(rows[i0][idxDate])
                idxToHCDATE2.append(rows[i1][idxDate])

                for i in range(len(outcomeLabels2) - 2):
                    iii = label.index(csvLabels[i])
                    outcomes2[i].append(float(rows[i1][iii]) - float(rows[i0][iii]))
                    before[i].append(float(rows[i0][iii]))
                    after[i].append(float(rows[i1][iii]))
                iii = label.index("メタボリックシンドローム判定")
                score0 = 0 if rows[i0][iii] == "非該当" else (1 if rows[i0][iii] == "予備群該当" else 2)
                score1 = 0 if rows[i1][iii] == "非該当" else (1 if rows[i1][iii] == "予備群該当" else 2)
                outcomes2[9].append(score1 - score0)
                before[9].append(score0)
                after[9].append(score1)
                idxMetaboD[score0][score1].append(idx)
                if(before[8][idx] > 100):
                    idxGammaGTAbnorm.append(idx)
            else:
                idxToHCDATE1.append(None)
                idxToHCDATE2.append(None)
                for j in range(len(outcomeLabels2) - 1):
                    outcomes2[j].append(None)
                    before[j].append(None)
                    after[j].append(None)
        else:
            idxToHCDATE1.append(None)
            idxToHCDATE2.append(None)
            for j in range(len(outcomeLabels2) - 1):
                outcomes2[j].append(None)
                before[j].append(None)
                after[j].append(None)
    for idx in idxFull:
        features[6].append(advisors[idxToCID[idx]])
    aveCNum = np.average(features[6])
    aveWaistD = np.average(outcomes[0])
    for idx in idxFull:
        if(advisors[idxToCID[idx]] >= aveCNum):
            idxCNum0.append(idx)
        else:
            idxCNum1.append(idx)
        waistD = outcomes[0][idx]
        if(waistD < aveWaistD):
            idxD0.append(idx)
        else:
            idxD1.append(idx)

    features = np.array(features).astype("float64")
    before = np.array(before).astype("float64")
    after = np.array(after).astype("float64")
    outcomes = np.array(outcomes).astype("float64")
    outcomes2 = np.array(outcomes2).astype("float64")
    theta1 = np.array(theta1).astype("float64").T  # K * M
    fkr = np.array(fkr).astype("float64").T
    thetas = np.array(thetas).transpose([0, 2, 1])
    # theta2 = np.array(theta2).T
    # theta3 = np.array(theta3).T
    # theta4 = np.array(theta4).T
    numTarget = features.shape[1]

    # init before[10], after[10], outcomes2[10] (MetaboD_estimated)
    model = sklearn.linear_model.LinearRegression(fit_intercept=True, normalize=False, copy_X=True, n_jobs=1)
    model.fit(np.hstack((before[:9, idxHCValid], after[:9, idxHCValid])).T, np.hstack((before[9, idxHCValid], after[9, idxHCValid])))
    before10 = np.full(len(idxFull), None)
    after10 = np.full(len(idxFull), None)
    outcomes10 = np.full(len(idxFull), None)
    before10[idxHCValid] = model.predict(before[:9, idxHCValid].T)
    after10[idxHCValid] = model.predict(after[:9, idxHCValid].T)
    outcomes10[idxHCValid] = after10[idxHCValid] - before10[idxHCValid]
    before = np.vstack((before, before10))
    after = np.vstack((after, after10))
    outcomes2 = np.vstack((outcomes2, outcomes10))

    # init idxAbnormals
    for i in range(len(HCLabel)):
        if(abnormalValues[i] is not None):
            if(abnormalFlags[i]):
                idxAbnormals.append(before[i, idxHCValid] >= abnormalValues[i])
            else:
                idxAbnormals.append(before[i, idxHCValid] < abnormalValues[i])
        else:
            idxAbnormals.append(None)

    # init outcome fixed
    outcomes_fixed = np.full_like(outcomes2, None)
    for i in range(len(HCLabel)):
        model = sklearn.linear_model.LinearRegression(fit_intercept=True, normalize=False, copy_X=True, n_jobs=1)
        model.fit(np.transpose([before[i][idxHCValid]]), outcomes2[i][idxHCValid])
        outcomes_fixed[i][idxHCValid] = outcomes2[i][idxHCValid] - model.coef_[0] * before[i][idxHCValid] - model.intercept_


def makeStatsSheet(wb):
    ws = wb.create_sheet("stats")
    ws.append(["THETA_KEY", THETA_KEY])
    ws.append(["MIN_DIRECTION_TIMES", MIN_DIRECTION_TIMES])
    ws.append(["numTarget", numTarget])
    ws.append(["numAdviser", len(advisors)])


def makeFeatureSheet(wb):
    ws = wb.create_sheet("feature")
    ws.append(featureLabels)
    numFormat = ["##.0", "#", "###", "###.0", "###.0", "###.0", "###", "#.0"]
    for i in range(numTarget):
        ws.append(features[:, i].tolist())
        for j in range(len(featureLabels)):
            ws.cell(row=i + 2, column=j + 1).number_format = numFormat[j]


def makeDirectionSheet(wb):
    ws = wb.create_sheet("direction")
    ws.append([])
    ws.append(thetaLabels * 2)
    ws.cell(row=1, column=len(thetaLabels) * 0 + 1, value="fkr_category")
    for n in range(MIN_DIRECTION_TIMES):
        ws.cell(row=1, column=len(thetaLabels) * (n + 1) + 1, value=f"theta{n+1}")
    for i in range(numTarget):
        for j in range(len(thetaLabels)):
            ws.cell(row=i + 3, column=len(thetaLabels) * 0 + j + 1, value=fkr[j, i])
            for n in range(MIN_DIRECTION_TIMES):
                ws.cell(row=i + 3, column=len(thetaLabels) * (n * 1) + j + 1, value=thetas[n][j, i])
                ws.cell(row=i + 3, column=len(thetaLabels) * (n * 1) + j + 1).number_format = "0.000"


def makeDirection2Sheet(wb):
    def func1(ws, a, b, i):
        plt.figure(figsize=(12, 4))
        plt.boxplot(a.T, labels=thetaLabels, showfliers=False, showmeans=True)
        plt.title(b)
        plt.xticks(rotation='vertical')
        plt.subplots_adjust(bottom=0.28)
        plt.savefig(str(temp.joinpath(b + '.png')))
        plt.close()
        img = openpyxl.drawing.image.Image(str(temp.joinpath(b + '.png')))
        img.height = 400
        img.width = 1200
        ws.add_image(img, "B" + str(i * 24 + 2))
    ws = wb.create_sheet("direction2")
    func1(ws, fkr, "fkr_category", 0)
    for n in range(MIN_DIRECTION_TIMES):
        func1(ws, thetas[n], f"theta{n+1}", n + 1)


def makePhi_simSheet(wb):
    ws = wb.create_sheet("phi_sim")
    lst = [LDA_MODEL.getTopicName(k) for k in range(LDA_MODEL.K)]
    ws.append(lst)
    for k1 in range(LDA_MODEL.K):
        lst = []
        for k2 in range(LDA_MODEL.K):
            lst.append(np.dot(LDA_MODEL.phi[k1], LDA_MODEL.phi[k2]) / np.linalg.norm(LDA_MODEL.phi[k1]) / np.linalg.norm(LDA_MODEL.phi[k2]))
        ws.append(lst)


def makeOutcomeSheet(wb):
    ws = wb.create_sheet("outcome")
    ws.append(outcomeLabels)
    numFormat = ["##0.0", "##0.0"]
    for i in range(numTarget):
        ws.append(outcomes[:, i].tolist())
        for j in range(len(outcomeLabels)):
            ws.cell(row=i + 2, column=j + 1).number_format = numFormat[j]


def makeOutcome2Sheet(wb):
    ws = wb.create_sheet("outcome2")
    ws.append(outcomeLabels2 + ["FirstGuidance", "LastGuidance", "BeforeCheck", "AfterCheck"])
    numFormat = ["##0.0", "##0.0", "##0.0", "##0.0", "##0.0", "##0.0", "##0.0", "##0.0", "##0.0", "#0.00", "#0.00"]
    for i in idxHCValid:
        ws.append(outcomes2[:, i].tolist() + [idxToSDATE[i].strftime("%y-%m-%d"), idxToEDATE[i].strftime("%y-%m-%d"),
                                              idxToHCDATE1[i].strftime("%y-%m-%d"), idxToHCDATE2[i].strftime("%y-%m-%d")])
        for j in range(len(outcomeLabels2)):
            ws.cell(row=i + 2, column=j + 1).number_format = numFormat[j]


def makeHistSheet(wb):
    def func2(ws, a, b, i):
        a = np.array(a).astype("float64")
        ws["S" + str(i * 24 + 2)] = b
        ws["S" + str(i * 24 + 2)].font = Font(color="dc143c")
        ws["S" + str(i * 24 + 3)] = "average"
        ws["S" + str(i * 24 + 3)].font = Font(color="006400")
        ws["T" + str(i * 24 + 3)] = np.mean(a)
        ws["S" + str(i * 24 + 4)] = "var"
        ws["S" + str(i * 24 + 4)].font = Font(color="006400")
        ws["T" + str(i * 24 + 4)] = np.var(a)
        plt.figure(figsize=(12, 4))
        plt.hist(a, bins=20)
        plt.title(b)
        plt.savefig(str(temp.joinpath(b + '.png')))
        plt.close()
        img = openpyxl.drawing.image.Image(str(temp.joinpath(b + '.png')))
        img.height = 400
        img.width = 1200
        ws.add_image(img, "B" + str(i * 24 + 2))
    ws = wb.create_sheet("hist")
    for i in range(len(featureLabels)):
        func2(ws, features[i], featureLabels[i], i)
    # for i in range(len(outcomeLabels)):
    #     func2(ws, outcomes[i], outcomeLabels[i], i + len(featureLabels))
    # for i in range(len(outcomeLabels2)):
    #     func2(ws, outcomes2[i][idxHCValid], outcomeLabels2[i], i + len(featureLabels) + len(outcomeLabels) + 1)
    # func2(ws, list(advisors.values()), "advisors", len(featureLabels) + len(outcomeLabels))
    for i in range(len(outcomeLabels2)):
        func2(ws, outcomes2[i, idxHCValid], outcomeLabels2[i], i + len(featureLabels) + 1)
    func2(ws, list(advisors.values()), "advisors", len(featureLabels))


def makePCA_phiSheet(wb):
    ws = wb.create_sheet("PCA_phi")
    lst = [LDA_MODEL.getTopicName(k) for k in range(LDA_MODEL.K)]
    pca = PCA(n_components=2)
    pca.fit(LDA_MODEL.phi.T)
    ws.cell(row=1, column=1, value="topik name")
    for k in range(pca.components_.shape[1]):
        ws.cell(row=k + 2, column=1, value=lst[k])
        for j in range(pca.components_.shape[0]):
            ws.cell(row=k + 2, column=j + 2, value=pca.components_[j, k])


def makePCA_thetaSheet(wb, folderToWriteResultIn):
    ws = wb.create_sheet("PCA_theta")
    n_clucsters = 6
    pca = PCA(n_components=n_clucsters)
    pca.fit(theta1)
    km = KMeans(n_clusters=n_clucsters, random_state=1)
    cluster = np.array(km.fit_predict(theta1[:len(LDA_MODEL.labelTypes)].T))
    lengths = km.transform(theta1[:len(LDA_MODEL.labelTypes)].T)

    ws.cell(row=1, column=1, value="m_id")
    ws.cell(row=1, column=pca.components_.shape[0] + 2, value="cluster id")
    ws.cell(row=1, column=pca.components_.shape[0] + 3, value="text")
    for d in range(pca.components_.shape[1]):
        ws.cell(row=d + 2, column=1, value=idxToMID[d])
        ws.cell(row=d + 2, column=pca.components_.shape[0] + 2, value=cluster[d])
        ws.cell(row=d + 2, column=pca.components_.shape[0] + 3, value=idxToTarget[d][0]["p_r_tgtset_explan"])
        for j in range(pca.components_.shape[0]):
            ws.cell(row=d + 2, column=j + 2, value=pca.components_[j, d])

    ws = wb.create_sheet("clustering_theta")
    ws.cell(row=1, column=1, value="cluster id")
    ws.cell(row=1, column=2, value="size")
    for k in range(len(LDA_MODEL.labelTypes)):
        ws.cell(row=1, column=k + 3, value=LDA_MODEL.getTopicName(k))
    for c in range(n_clucsters):
        ave = np.average(theta1[:, cluster == c], axis=1)
        ws.cell(row=c + 2, column=1, value=c)
        ws.cell(row=c + 2, column=2, value=len(theta1[0, cluster == c]))
        for k in range(len(LDA_MODEL.labelTypes)):
            ws.cell(row=c + 2, column=k + 3, value=ave[k])

    plt.figure(figsize=(12, 12))
    for i in range(n_clucsters):
        plt.scatter(pca.components_[0][cluster == i], pca.components_[1][cluster == i], label=str(i))
    plt.title("scatter_PCAed_theta")
    plt.legend()
    plt.savefig(str(temp.joinpath('scatter_PCAed_theta.png')))
    plt.close()

    def func(dct, key, default=None):
        if(key in dct):
            return dct[key]
        else:
            return default

    folderToWriteResultIn.joinpath("text").mkdir(exist_ok=True)
    folder = folderToWriteResultIn.joinpath("text", "cluster")
    if(folder.exists()):
        shutil.rmtree(folder)
    folder.mkdir(exist_ok=True)
    for c in range(n_clucsters):
        with open(str(folderToWriteResultIn.joinpath("text", "cluster", 'cluster{}.json'.format(c))), "w", encoding="utf_8_sig") as output:
            idx = np.argsort(lengths[:, c])
            a0 = []
            for n in range(numTarget):
                iii = idx[n]
                if(iii in idxHCValid):
                    a1 = idxToTarget[iii][0]
                    d0 = {}
                    d0["__p_r_explan_iyoku"] = func(a1, "p_r_explan_iyoku")
                    d0["__p_r_explained"] = func(a1, "p_r_explained")
                    d0["__snack"] = func(a1, "p_r_snack_text")
                    d0["__drink"] = func(a1, "p_r_drink_text")
                    d0["__sake"] = func(a1, "p_r_sake_text")
                    d0["__exercise"] = func(a1, "p_r_exer")
                    d0["__smoke"] = func(a1, "p_r_other")
                    d0["__syokuji_time"] = func(a1, "p_r_other_text")
                    d0["__sleep"] = func(a1, "p_r_sleep_text")
                    d0["__stress"] = func(a1, "p_r_stress_rel")
                    for nnn in range(MIN_DIRECTION_TIMES):
                        d0[f"__p_r_tgtset_explan{nnn+1}"] = func(idxToTarget[iii][nnn], "p_r_tgtset_explan")
                    d0["__feature"] = str(features[:, iii])
                    d0["__before"] = str(before[:, iii])
                    d0["__after"] = str(after[:, iii])
                    d0["__outcome"] = str(outcomes2[:, iii])
                    d0["__fkr_label"] = str(fkr[:, iii])
                    for nnn in range(MIN_DIRECTION_TIMES):
                        d0[f"__fkr_text{nnn+1}"] = func(idxToTarget[iii][nnn], "p_r_tgtset_explan_seqs_text")
                    d0["__p_shienkeikaku"] = func(a1, "__p_shienkeikaku")
                    d0["date"] = "HC1={}, start={} end={}, HC2={}".format(idxToHCDATE1[iii], idxToSDATE[iii],
                                                                          idxToEDATE[iii], idxToHCDATE2[iii])
                    a0.append(d0)
                    if(len(a0) >= 10):
                        break
            json.dump(a0, output, ensure_ascii=False, indent=2)


def makeAverageSheet(wb):
    def func3(ws, a, b, name, compare=None):
        ave = np.average(a, axis=1)
        ws.cell(row=nowRow[0], column=1, value=name).font = Font(color="0000cd")
        for i in range(len(b)):
            ws.cell(row=nowRow[0], column=i + 2, value=b[i]).font = Font(color="006400")
            ws.cell(row=nowRow[0] + 1, column=i + 2, value=ave[i])
            if(compare is not None):
                aaa = a[i].astype("float64")
                bbb = compare[i].astype("float64")
                ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                ws.cell(row=nowRow[0] + 1, column=i + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                if(ttest[1] <= 0.01):
                    ws.cell(row=nowRow[0] + 1, column=i + 2).fill = fill1 if ttest[0] > 0 else fill3
                elif(ttest[1] <= 0.05):
                    ws.cell(row=nowRow[0] + 1, column=i + 2).fill = fill2 if ttest[0] > 0 else fill4
        nowRow[0] += 2
    ws = wb.create_sheet("average")
    nowRow[0] = 4
    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 2, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Full Data").font = Font(color="dc143c")
    func3(ws, features[:, idxFull], featureLabels, "features")
    func3(ws, before[:, idxHCValid], HCLabel, "before")
    func3(ws, outcomes[:, idxFull], outcomeLabels, "outcomes")
    func3(ws, outcomes2[:, listAnd(idxHCValid, idxFull)], outcomeLabels2, "outcomes2")
    func3(ws, theta1[:, idxFull], thetaLabels, "theta1")
    func3(ws, fkr[:, idxFull], thetaLabels, "fkr_category")
    # func3(ws, thetas[1][:, idxFull], thetaLabels, "theta2")
    # func3(ws, thetas[2][:, idxFull], thetaLabels, "theta3")
    nowRow[0] += 4
    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Data recorded by skilled advisors(c_num >= {:<.1f})".format(aveCNum)).font = Font(color="dc143c")
    func3(ws, features[:, idxCNum0], featureLabels, "features", features[:, idxCNum1])
    func3(ws, before[:, listAnd(idxHCValid, idxCNum0)], HCLabel, "before", before[:, listAnd(idxHCValid, idxCNum1)])
    func3(ws, outcomes[:, idxCNum0], outcomeLabels, "outcomes", outcomes[:, idxCNum1])
    func3(ws, outcomes2[:, listAnd(idxHCValid, idxCNum0)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idxCNum1)])
    func3(ws, theta1[:, idxCNum0], thetaLabels, "theta1", theta1[:, idxCNum1])
    func3(ws, fkr[:, idxCNum0], thetaLabels, "fkr_category", fkr[:, idxCNum1])
    # func3(ws, thetas[1][:, idxCNum0], thetaLabels, "theta2", thetas[1][:, idxCNum1])
    # func3(ws, thetas[2][:, idxCNum0], thetaLabels, "theta3", thetas[2][:, idxCNum1])
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Data recorded by non-skilled advisors").font = Font(color="dc143c")
    func3(ws, features[:, idxCNum1], featureLabels, "features", features[:, idxCNum0])
    func3(ws, before[:, listAnd(idxHCValid, idxCNum1)], HCLabel, "before", before[:, listAnd(idxHCValid, idxCNum0)])
    func3(ws, outcomes[:, idxCNum1], outcomeLabels, "outcomes", outcomes[:, idxCNum0])
    func3(ws, outcomes2[:, listAnd(idxHCValid, idxCNum1)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idxCNum0)])
    func3(ws, theta1[:, idxCNum1], thetaLabels, "theta1", theta1[:, idxCNum0])
    func3(ws, fkr[:, idxCNum1], thetaLabels, "fkr_category", fkr[:, idxCNum0])
    # func3(ws, thetas[1][:, idxCNum1], thetaLabels, "theta2", thetas[1][:, idxCNum0])
    # func3(ws, thetas[2][:, idxCNum1], thetaLabels, "theta3", thetas[2][:, idxCNum0])
    nowRow[0] += 4
    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with good outcomes(dWaist >= {:<.2f})".format(aveWaistD)).font = Font(color="dc143c")
    func3(ws, features[:, idxD0], featureLabels, "features", features[:, idxD1])
    func3(ws, before[:, listAnd(idxHCValid, idxD0)], HCLabel, "before", before[:, listAnd(idxHCValid, idxD1)])
    func3(ws, outcomes[:, idxD0], outcomeLabels, "outcomes", outcomes[:, idxD1])
    func3(ws, outcomes2[:, listAnd(idxHCValid, idxD0)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idxD1)])
    func3(ws, theta1[:, idxD0], thetaLabels, "theta1", theta1[:, idxD1])
    func3(ws, fkr[:, idxD0], thetaLabels, "fkr_category", fkr[:, idxD1])
    # func3(ws, thetas[1][:, idxD0], thetaLabels, "theta2", thetas[1][:, idxD1])
    # func3(ws, thetas[2][:, idxD0], thetaLabels, "theta3", thetas[2][:, idxD1])
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with bad outcomes").font = Font(color="dc143c")
    func3(ws, features[:, idxD1], featureLabels, "features", features[:, idxD0])
    func3(ws, before[:, listAnd(idxHCValid, idxD1)], HCLabel, "before", before[:, listAnd(idxHCValid, idxD0)])
    func3(ws, outcomes[:, idxD1], outcomeLabels, "outcomes", outcomes[:, idxD0])
    func3(ws, outcomes2[:, listAnd(idxHCValid, idxD1)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idxD0)])
    func3(ws, theta1[:, idxD1], thetaLabels, "theta1", theta1[:, idxD0])
    func3(ws, fkr[:, idxD1], thetaLabels, "fkr_category", fkr[:, idxD0])
    # func3(ws, thetas[1][:, idxD1], thetaLabels, "theta2", thetas[1][:, idxD0])
    # func3(ws, thetas[2][:, idxD1], thetaLabels, "theta3", thetas[2][:, idxD0])

    nowRow[0] += 4
    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Metabo dicision 1 to 0 (n = {})".format(len(idxMetaboD[1][0]))).font = Font(color="dc143c")
    func3(ws, features[:, idxMetaboD[1][0]], featureLabels, "features", features[:, idxMetaboD[1][2]])
    func3(ws, before[:, idxMetaboD[1][0]], HCLabel, "before", before[:, idxMetaboD[1][2]])
    func3(ws, outcomes[:, idxMetaboD[1][0]], outcomeLabels, "outcomes", outcomes[:, idxMetaboD[1][2]])
    func3(ws, outcomes2[:, idxMetaboD[1][0]], outcomeLabels2, "outcomes2", outcomes2[:, idxMetaboD[1][2]])
    func3(ws, theta1[:, idxMetaboD[1][0]], thetaLabels, "theta1", theta1[:, idxMetaboD[1][2]])
    # func3(ws, thetas[1][:, idxMetaboD[1][0]], thetaLabels, "theta2", thetas[1][:, idxMetaboD[1][2]])
    # func3(ws, thetas[2][:, idxMetaboD[1][0]], thetaLabels, "theta3", thetas[2][:, idxMetaboD[1][2]])
    # func3(ws, thetas[3][:, idxMetaboD[1][0]], thetaLabels, "theta4", thetas[3][:, idxMetaboD[1][2]])
    func3(ws, fkr[:, idxMetaboD[1][0]], thetaLabels, "fkr_category", fkr[:, idxMetaboD[1][2]])
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Metabo dicision 1 to 1 (n = {})".format(len(idxMetaboD[1][1]))).font = Font(color="dc143c")
    func3(ws, features[:, idxMetaboD[1][1]], featureLabels, "features", None)
    func3(ws, before[:, idxMetaboD[1][1]], HCLabel, "before", None)
    func3(ws, outcomes[:, idxMetaboD[1][1]], outcomeLabels, "outcomes", None)
    func3(ws, outcomes2[:, idxMetaboD[1][1]], outcomeLabels2, "outcomes2", None)
    func3(ws, theta1[:, idxMetaboD[1][1]], thetaLabels, "theta1", None)
    func3(ws, fkr[:, idxMetaboD[1][1]], thetaLabels, "fkr_category", None)
    # func3(ws, thetas[1][:, idxMetaboD[1][1]], thetaLabels, "theta2", None)
    # func3(ws, thetas[2][:, idxMetaboD[1][1]], thetaLabels, "theta3", None)
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Metabo dicision 1 to 2 (n = {})".format(len(idxMetaboD[1][2]))).font = Font(color="dc143c")
    func3(ws, features[:, idxMetaboD[1][2]], featureLabels, "features", features[:, idxMetaboD[1][0]])
    func3(ws, before[:, idxMetaboD[1][2]], HCLabel, "features", before[:, idxMetaboD[1][0]])
    func3(ws, outcomes[:, idxMetaboD[1][2]], outcomeLabels, "outcomes", outcomes[:, idxMetaboD[1][0]])
    func3(ws, outcomes2[:, idxMetaboD[1][2]], outcomeLabels2, "outcomes2", outcomes2[:, idxMetaboD[1][0]])
    func3(ws, theta1[:, idxMetaboD[1][2]], thetaLabels, "theta1", theta1[:, idxMetaboD[1][0]])
    # func3(ws, thetas[1][:, idxMetaboD[1][2]], thetaLabels, "theta2", thetas[1][:, idxMetaboD[1][0]])
    # func3(ws, thetas[2][:, idxMetaboD[1][2]], thetaLabels, "theta3", thetas[2][:, idxMetaboD[1][0]])
    # func3(ws, thetas[3][:, idxMetaboD[1][2]], thetaLabels, "theta4", thetas[3][:, idxMetaboD[1][0]])
    func3(ws, fkr[:, idxMetaboD[1][2]], thetaLabels, "fkr_category", fkr[:, idxMetaboD[1][0]])
    ws.cell(row=nowRow[0] + 1, column=1, value="Tukey HSD test").font = Font(color="dc143c")
    ws.cell(row=nowRow[0] + 2, column=1, value="1to0 and 1to1").font = Font(color="0000cd")
    ws.cell(row=nowRow[0] + 3, column=1, value="1to0 and 1to2").font = Font(color="0000cd")
    ws.cell(row=nowRow[0] + 4, column=1, value="1to1 and 1to2").font = Font(color="0000cd")
    for k in range(len(thetaLabels)):
        tresult = tukey_hsd(list('ABC'), theta1[k, idxMetaboD[1][0]], theta1[k, idxMetaboD[1][1]], theta1[k, idxMetaboD[1][2]])
        for i in range(3):
            ws.cell(row=nowRow[0] + i + 2, column=k + 2, value=tresult.reject[i])
    nowRow[0] += 5

    # nowRow[0] += 4
    # idx2to0or1 = listOr(idxMetaboD[2][0], idxMetaboD[2][1])
    # for i in range(1, 21):
    #     ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    # ws.cell(row=nowRow[0] - 1, column=1, value="Metabo dicision 2 to 0 or 1 (n = {})".format(len(idx2to0or1))).font = Font(color="dc143c")
    # func3(ws, features[:, idx2to0or1], featureLabels, "features", features[:, idxMetaboD[2][2]])
    # func3(ws, before[:, idx2to0or1], HCLabel, "before", before[:, idxMetaboD[2][2]])
    # func3(ws, outcomes[:, idx2to0or1], outcomeLabels, "outcomes", outcomes[:, idxMetaboD[2][2]])
    # func3(ws, outcomes2[:, idx2to0or1], outcomeLabels2, "outcomes2", outcomes2[:, idxMetaboD[2][2]])
    # func3(ws, theta1[:, idx2to0or1], thetaLabels, "theta1", theta1[:, idxMetaboD[2][2]])
    # func3(ws, fkr[:, idx2to0or1], thetaLabels, "fkr_category", fkr[:, idxMetaboD[2][2]])
    # # func3(ws, thetas[1][:, idx2to0or1], thetaLabels, "theta2", thetas[1][:, idxMetaboD[2][2]])
    # # func3(ws, thetas[2][:, idx2to0or1], thetaLabels, "theta3", thetas[2][:, idxMetaboD[2][2]])
    # nowRow[0] += 2
    # ws.cell(row=nowRow[0] - 1, column=1, value="Metabo dicision 2 to 2 (n = {})".format(len(idxMetaboD[2][2]))).font = Font(color="dc143c")
    # func3(ws, features[:, idxMetaboD[2][2]], featureLabels, "features", features[:, idx2to0or1])
    # func3(ws, before[:, idxMetaboD[2][2]], HCLabel, "before", before[:, idx2to0or1])
    # func3(ws, outcomes[:, idxMetaboD[2][2]], outcomeLabels, "outcomes", outcomes[:, idx2to0or1])
    # func3(ws, outcomes2[:, idxMetaboD[2][2]], outcomeLabels2, "outcomes2", outcomes2[:, idx2to0or1])
    # func3(ws, theta1[:, idxMetaboD[2][2]], thetaLabels, "theta1", theta1[:, idx2to0or1])
    # func3(ws, fkr[:, idxMetaboD[2][2]], thetaLabels, "fkr_category", fkr[:, idx2to0or1])
    # # func3(ws, thetas[1][:, idxMetaboD[2][2]], thetaLabels, "theta2", thetas[1][:, idx2to0or1])
    # # func3(ws, thetas[2][:, idxMetaboD[2][2]], thetaLabels, "theta3", thetas[2][:, idx2to0or1])
    # nowRow[0] += 4
    # idx0to1or2 = listOr(idxMetaboD[0][1], idxMetaboD[0][2])
    # for i in range(1, 21):
    #     ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    # ws.cell(row=nowRow[0] - 1, column=1, value="Metabo dicision 0 to 0 (n = {})".format(len(idxMetaboD[0][0]))).font = Font(color="dc143c")
    # func3(ws, features[:, idxMetaboD[0][0]], featureLabels, "features", features[:, idx0to1or2])
    # func3(ws, before[:, idxMetaboD[0][0]], HCLabel, "before", before[:, idx0to1or2])
    # func3(ws, outcomes[:, idxMetaboD[0][0]], outcomeLabels, "outcomes", outcomes[:, idx0to1or2])
    # func3(ws, outcomes2[:, idxMetaboD[0][0]], outcomeLabels2, "outcomes2", outcomes2[:, idx0to1or2])
    # func3(ws, theta1[:, idxMetaboD[0][0]], thetaLabels, "theta1", theta1[:, idx0to1or2])
    # func3(ws, fkr[:, idxMetaboD[0][0]], thetaLabels, "fkr_category", fkr[:, idx0to1or2])
    # # func3(ws, thetas[1][:, idxMetaboD[0][0]], thetaLabels, "theta2", thetas[1][:, idx0to1or2])
    # # func3(ws, thetas[2][:, idxMetaboD[0][0]], thetaLabels, "theta3", thetas[2][:, idx0to1or2])
    # nowRow[0] += 2
    # ws.cell(row=nowRow[0] - 1, column=1, value="Metabo dicision 0 to 1 or 2 (n = {})".format(len(idx0to1or2))).font = Font(color="dc143c")
    # func3(ws, features[:, idx0to1or2], featureLabels, "features", features[:, idxMetaboD[0][0]])
    # func3(ws, before[:, idx0to1or2], HCLabel, "before", before[:, idxMetaboD[0][0]])
    # func3(ws, outcomes[:, idx0to1or2], outcomeLabels, "outcomes", outcomes[:, idxMetaboD[0][0]])
    # func3(ws, outcomes2[:, idx0to1or2], outcomeLabels2, "outcomes2", outcomes2[:, idxMetaboD[0][0]])
    # func3(ws, theta1[:, idx0to1or2], thetaLabels, "theta1", theta1[:, idxMetaboD[0][0]])
    # func3(ws, fkr[:, idx0to1or2], thetaLabels, "fkr_category", fkr[:, idxMetaboD[0][0]])
    # # func3(ws, thetas[1][:, idx0to1or2], thetaLabels, "theta2", thetas[1][:, idxMetaboD[0][0]])
    # # func3(ws, thetas[2][:, idx0to1or2], thetaLabels, "theta3", thetas[2][:, idxMetaboD[0][0]])

    aveDGammaGT = np.average(outcomes2[8, idxGammaGTAbnorm].astype("float64"))
    idxGood = outcomes2[8, idxGammaGTAbnorm] <= aveDGammaGT
    idxBad = outcomes2[8, idxGammaGTAbnorm] > aveDGammaGT
    nowRow[0] += 4
    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Good outcome(gammaGT) in abnormal gammagt targets (ave dGammaGT={:<.2f}) (n = {})".format(aveDGammaGT, idxGood.sum())).font = Font(color="dc143c")
    func3(ws, features[:, idxGammaGTAbnorm][:, idxGood], featureLabels, "features", features[:, idxGammaGTAbnorm][:, idxBad])
    func3(ws, before[:, idxGammaGTAbnorm][:, idxGood], HCLabel, "before", before[:, idxGammaGTAbnorm][:, idxBad])
    func3(ws, outcomes[:, idxGammaGTAbnorm][:, idxGood], outcomeLabels, "outcomes", outcomes[:, idxGammaGTAbnorm][:, idxBad])
    func3(ws, outcomes2[:, idxGammaGTAbnorm][:, idxGood], outcomeLabels2, "outcomes2", outcomes2[:, idxGammaGTAbnorm][:, idxBad])
    func3(ws, theta1[:, idxGammaGTAbnorm][:, idxGood], thetaLabels, "theta1", theta1[:, idxGammaGTAbnorm][:, idxBad])
    func3(ws, fkr[:, idxGammaGTAbnorm][:, idxGood], thetaLabels, "fkr_category", fkr[:, idxGammaGTAbnorm][:, idxBad])
    # func3(ws, thetas[1][:, idxGood], thetaLabels, "theta2", thetas[1][:, idxBad])
    # func3(ws, thetas[2][:, idxGood], thetaLabels, "theta3", thetas[2][:, idxBad])
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Bad outcome(gammaGT) in abnormal gammagt targets (n = {})".format(idxBad.sum())).font = Font(color="dc143c")
    func3(ws, features[:, idxGammaGTAbnorm][:, idxBad], featureLabels, "features", features[:, idxGammaGTAbnorm][:, idxGood])
    func3(ws, before[:, idxGammaGTAbnorm][:, idxBad], HCLabel, "before", before[:, idxGammaGTAbnorm][:, idxGood])
    func3(ws, outcomes[:, idxGammaGTAbnorm][:, idxBad], outcomeLabels, "outcomes", outcomes[:, idxGammaGTAbnorm][:, idxGood])
    func3(ws, outcomes2[:, idxGammaGTAbnorm][:, idxBad], outcomeLabels2, "outcomes2", outcomes2[:, idxGammaGTAbnorm][:, idxGood])
    func3(ws, theta1[:, idxGammaGTAbnorm][:, idxBad], thetaLabels, "theta1", theta1[:, idxGammaGTAbnorm][:, idxGood])
    func3(ws, fkr[:, idxGammaGTAbnorm][:, idxBad], thetaLabels, "fkr_category", fkr[:, idxGammaGTAbnorm][:, idxGood])
    # func3(ws, thetas[1][:, idxBad], thetaLabels, "theta2", thetas[1][:, idxGood])
    # func3(ws, thetas[2][:, idxBad], thetaLabels, "theta3", thetas[2][:, idxGood])

    aveBDBP = np.average(before[4, idxHCValid].astype("float64"))
    idxGood = before[4, idxHCValid] <= aveBDBP
    idxBad = before[4, idxHCValid] > aveBDBP
    nowRow[0] += 4
    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with good before DBP (ave DBP={:<.2f}) (n = {})".format(aveBDBP, idxGood.sum())).font = Font(color="dc143c")
    func3(ws, features[:, idxHCValid][:, idxGood], featureLabels, "features", features[:, idxHCValid][:, idxBad])
    func3(ws, before[:, idxHCValid][:, idxGood], HCLabel, "before", before[:, idxHCValid][:, idxBad])
    func3(ws, outcomes[:, idxHCValid][:, idxGood], outcomeLabels, "outcomes", outcomes[:, idxHCValid][:, idxBad])
    func3(ws, outcomes2[:, idxHCValid][:, idxGood], outcomeLabels2, "outcomes2", outcomes2[:, idxHCValid][:, idxBad])
    func3(ws, theta1[:, idxHCValid][:, idxGood], thetaLabels, "theta1", theta1[:, idxHCValid][:, idxBad])
    func3(ws, fkr[:, idxHCValid][:, idxGood], thetaLabels, "fkr_category", fkr[:, idxHCValid][:, idxBad])
    # func3(ws, thetas[1][:, idxGood], thetaLabels, "theta2", thetas[1][:, idxBad])
    # func3(ws, thetas[2][:, idxGood], thetaLabels, "theta3", thetas[2][:, idxBad])
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with bad before DBP (n = {})".format(idxBad.sum())).font = Font(color="dc143c")
    func3(ws, features[:, idxHCValid][:, idxBad], featureLabels, "features", features[:, idxHCValid][:, idxGood])
    func3(ws, before[:, idxHCValid][:, idxBad], HCLabel, "before", before[:, idxHCValid][:, idxGood])
    func3(ws, outcomes[:, idxHCValid][:, idxBad], outcomeLabels, "outcomes", outcomes[:, idxHCValid][:, idxGood])
    func3(ws, outcomes2[:, idxHCValid][:, idxBad], outcomeLabels2, "outcomes2", outcomes2[:, idxHCValid][:, idxGood])
    func3(ws, theta1[:, idxHCValid][:, idxBad], thetaLabels, "theta1", theta1[:, idxHCValid][:, idxGood])
    func3(ws, fkr[:, idxHCValid][:, idxBad], thetaLabels, "fkr_category", fkr[:, idxHCValid][:, idxGood])
    # func3(ws, thetas[1][:, idxBad], thetaLabels, "theta2", thetas[1][:, idxGood])
    # func3(ws, thetas[2][:, idxBad], thetaLabels, "theta3", thetas[2][:, idxGood])

    nowRow[0] += 4
    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with each labels.").font = Font(color="dc143c")
    for k in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=k + 2, value=thetaLabels[k]).font = Font(color="006400")
    for i in range(len(outcomeLabels2)):
        ws.cell(row=nowRow[0] + i + 1, column=1, value=outcomeLabels2[i]).font = Font(color="006400")
    for k in range(len(thetaLabels)):
        for i in range(len(outcomeLabels2)):
            aaa = outcomes2[i, idxHCValid][theta1[k, idxHCValid] != 0]
            bbb = outcomes2[i, idxHCValid][theta1[k, idxHCValid] == 0]
            if(len(aaa) > 1 and len(bbb) > 1):
                ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                ws.cell(row=nowRow[0] + i + 1, column=k + 2, value=ttest[0])
                ws.cell(row=nowRow[0] + i + 1, column=k + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                if(ttest[1] <= 0.01):
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill1 if ttest[0] > 0 else fill3
                elif(ttest[1] <= 0.05):
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill2 if ttest[0] > 0 else fill4
    nowRow[0] += len(outcomeLabels2) + 5

    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with each labels (in abnormal before values).").font = Font(color="dc143c")
    for k in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=k + 2, value=thetaLabels[k]).font = Font(color="006400")
    for i in range(len(outcomeLabels2)):
        ws.cell(row=nowRow[0] + i + 1, column=1, value=outcomeLabels2[i]).font = Font(color="006400")
    for k in range(len(thetaLabels)):
        for i in range(len(outcomeLabels2)):
            aaa = outcomes2[i, idxHCValid][idxAbnormals[i]][theta1[k, idxHCValid][idxAbnormals[i]] != 0].astype("float")
            bbb = outcomes2[i, idxHCValid][idxAbnormals[i]][theta1[k, idxHCValid][idxAbnormals[i]] == 0].astype("float")
            if(len(aaa) != 0 and len(bbb) != 0):
                ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                ws.cell(row=nowRow[0] + i + 1, column=k + 2, value=ttest[0])
                ws.cell(row=nowRow[0] + i + 1, column=k + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                if(ttest[1] <= 0.01):
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill1 if ttest[0] > 0 else fill3
                elif(ttest[1] <= 0.05):
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill2 if ttest[0] > 0 else fill4
    nowRow[0] += len(outcomeLabels2) + 5

    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Average direction vector (in abnormal before values).").font = Font(color="dc143c")
    for k in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=k + 2, value=thetaLabels[k]).font = Font(color="006400")
    for i in range(len(outcomeLabels2)):
        ws.cell(row=nowRow[0] + i + 1, column=1, value=outcomeLabels2[i]).font = Font(color="006400")
    for k in range(len(thetaLabels)):
        for i in range(len(outcomeLabels2)):
            if(idxAbnormals[i] is not None):
                ave = np.average(outcomes2[i, idxHCValid][idxAbnormals[i]].astype("float"))
                aaa = theta1[k, idxHCValid][idxAbnormals[i]][outcomes2[i, idxHCValid][idxAbnormals[i]] > ave].astype("float")
                bbb = theta1[k, idxHCValid][idxAbnormals[i]][outcomes2[i, idxHCValid][idxAbnormals[i]] <= ave].astype("float")
                if(len(aaa) != 0 and len(bbb) != 0):
                    ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2, value=np.average(aaa) - np.average(bbb))
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                    if(ttest[1] <= 0.05):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill1 if ttest[0] > 0 else fill3
                    elif(ttest[1] <= 0.1):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill2 if ttest[0] > 0 else fill4
    nowRow[0] += len(outcomeLabels2) + 5

    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Average direction vector (fixed)(in abnormal before values).").font = Font(color="dc143c")
    for k in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=k + 2, value=thetaLabels[k]).font = Font(color="006400")
    for i in range(len(outcomeLabels2)):
        ws.cell(row=nowRow[0] + i + 1, column=1, value=outcomeLabels2[i]).font = Font(color="006400")
    for k in range(len(thetaLabels)):
        for i in range(len(outcomeLabels2)):
            if(idxAbnormals[i] is not None):
                ave = np.average(outcomes_fixed[i, idxHCValid][idxAbnormals[i]].astype("float"))
                aaa = theta1[k, idxHCValid][idxAbnormals[i]][outcomes_fixed[i, idxHCValid][idxAbnormals[i]] > ave].astype("float")
                bbb = theta1[k, idxHCValid][idxAbnormals[i]][outcomes_fixed[i, idxHCValid][idxAbnormals[i]] <= ave].astype("float")
                if(len(aaa) != 0 and len(bbb) != 0):
                    ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2, value=np.average(aaa) - np.average(bbb))
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                    if(ttest[1] <= 0.05):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill1 if ttest[0] > 0 else fill3
                    elif(ttest[1] <= 0.1):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill2 if ttest[0] > 0 else fill4
    nowRow[0] += len(outcomeLabels2) + 5

    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Average label frequency (in abnormal before values).").font = Font(color="dc143c")
    for k in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=k + 2, value=thetaLabels[k]).font = Font(color="006400")
    for i in range(len(outcomeLabels2)):
        ws.cell(row=nowRow[0] + i + 1, column=1, value=outcomeLabels2[i]).font = Font(color="006400")
    for k in range(len(thetaLabels)):
        for i in range(len(outcomeLabels2)):
            if(idxAbnormals[i] is not None):
                ave = np.average(outcomes2[i, idxHCValid][idxAbnormals[i]].astype("float"))
                aaa = fkr[k, idxHCValid][idxAbnormals[i]][outcomes2[i, idxHCValid][idxAbnormals[i]] > ave].astype("float")
                bbb = fkr[k, idxHCValid][idxAbnormals[i]][outcomes2[i, idxHCValid][idxAbnormals[i]] <= ave].astype("float")
                if(len(aaa) != 0 and len(bbb) != 0):
                    ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2, value=np.average(aaa) - np.average(bbb))
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                    if(ttest[1] <= 0.05):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill1 if ttest[0] > 0 else fill3
                    elif(ttest[1] <= 0.1):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill2 if ttest[0] > 0 else fill4
    nowRow[0] += len(outcomeLabels2) + 5

    for i in range(1, 21):
        ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
    ws.cell(row=nowRow[0] - 1, column=1, value="Average label frequency (fixed)(in abnormal before values).").font = Font(color="dc143c")
    for k in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=k + 2, value=thetaLabels[k]).font = Font(color="006400")
    for i in range(len(outcomeLabels2)):
        ws.cell(row=nowRow[0] + i + 1, column=1, value=outcomeLabels2[i]).font = Font(color="006400")
    for k in range(len(thetaLabels)):
        for i in range(len(outcomeLabels2)):
            if(idxAbnormals[i] is not None):
                ave = np.average(outcomes_fixed[i, idxHCValid][idxAbnormals[i]].astype("float"))
                aaa = fkr[k, idxHCValid][idxAbnormals[i]][outcomes_fixed[i, idxHCValid][idxAbnormals[i]] > ave].astype("float")
                bbb = fkr[k, idxHCValid][idxAbnormals[i]][outcomes_fixed[i, idxHCValid][idxAbnormals[i]] <= ave].astype("float")
                if(len(aaa) != 0 and len(bbb) != 0):
                    ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2, value=np.average(aaa) - np.average(bbb))
                    ws.cell(row=nowRow[0] + i + 1, column=k + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                    if(ttest[1] <= 0.05):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill1 if ttest[0] > 0 else fill3
                    elif(ttest[1] <= 0.1):
                        ws.cell(row=nowRow[0] + i + 1, column=k + 2).fill = fill2 if ttest[0] > 0 else fill4
    nowRow[0] += len(outcomeLabels2) + 5


def makeCoefSheet(wb, flag7_2):
    def func4(ws1, ws2, name, a, b, aLabel, bLabel, idxs=None):
        ws1.cell(row=nowRow[0], column=1, value=name).font = Font(color="0000cd")
        for j in range(len(aLabel)):
            ws1.cell(row=nowRow[0], column=j + 2, value=aLabel[j]).font = Font(color="006400")
        for i in range(len(bLabel)):
            ws1.cell(row=nowRow[0] + i + 1, column=1, value=bLabel[i]).font = Font(color="006400")
        for i in range(len(bLabel)):
            for j in range(len(aLabel)):
                if(idxs is None or idxs[i] is None):
                    r, p = stats.pearsonr(a[j], b[i])
                else:
                    r, p = stats.pearsonr(a[j][idxs[i]], b[i][idxs[i]])
                ws1.cell(row=nowRow[0] + i + 1, column=j + 2, value=r)
                ws1.cell(row=nowRow[0] + i + 1, column=j + 2).comment = openpyxl.comments.Comment(p, None)
                if(p <= 0.05):
                    if(flag7_2):
                        # plt.figure(figsize=(6, 4))
                        plt.figure(figsize=(10, 10))
                        plt.scatter(a[j], b[i])
                        # plt.xlabel(aLabel[j])
                        # plt.ylabel(bLabel[i])
                        plt.savefig(str(temp.joinpath('coef{}.png'.format(imageID[0]))))
                        plt.close()
                        img = openpyxl.drawing.image.Image(str(temp.joinpath('coef{}.png'.format(imageID[0]))))
                        # img.height = 400
                        # img.width = 600
                        img.height = 180
                        img.width = 180
                        # ws1.add_image(img, "T12")
                        ws1.add_image(img, "K2")
                        ws2.cell(row=nowRow[0] + i + 1, column=j + 2, value=imageID[0])
                        imageID[0] += 1
        setColorScale(ws1, nowRow[0] + 1, 2, nowRow[0] + len(bLabel) + 1, len(aLabel) + 2)
        nowRow[0] += len(bLabel) + 1
    ws1 = wb.get_sheet_by_name("coef")
    ws2 = wb.get_sheet_by_name("coef_meta")
    imageID = [1]
    nowRow[0] = 2
    # TODO move sheet
    # ws1.cell(row=1, column=1, value="Full Data").font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxFull], outcomes[:, idxFull], featureLabels, outcomeLabels)
    # func4(ws1, ws2, "theta1", theta1[:, idxFull], outcomes[:, idxFull], thetaLabels, outcomeLabels)
    # func4(ws1, ws2, "fkr_category", fkr[:, idxFull], outcomes[:, idxFull], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta2", thetas[1][:, idxFull], outcomes[:, idxFull], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta3", thetas[2][:, idxFull], outcomes[:, idxFull], thetaLabels, outcomeLabels)
    nowRow[0] += 2
    ws1.cell(row=nowRow[0] - 1, column=1, value="All targets (n={})".format(len(idxHCValid))).font = Font(color="dc143c")
    func4(ws1, ws2, "features", features[:, idxHCValid], outcomes2[:, idxHCValid], featureLabels, outcomeLabels2)
    func4(ws1, ws2, "theta1", theta1[:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2)
    func4(ws1, ws2, "fkr_category", fkr[:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta2", thetas[1][:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta3", thetas[2][:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2)
    nowRow[0] += 2
    ws1.cell(row=nowRow[0] - 1, column=1, value="Targets with abnormal GammaGT (>100) (n={})".format(len(idxGammaGTAbnorm))).font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxGammaGTAbnorm], outcomes2[:, idxGammaGTAbnorm], featureLabels, outcomeLabels2)
    func4(ws1, ws2, "theta1", theta1[:, idxGammaGTAbnorm], outcomes2[:, idxGammaGTAbnorm], thetaLabels[:], outcomeLabels2)
    func4(ws1, ws2, "fkr_category", fkr[:, idxGammaGTAbnorm], outcomes2[:, idxGammaGTAbnorm], thetaLabels[:], outcomeLabels2)
    # func4(ws1, ws2, "theta2", thetas[1][:, idxGammaGTAbnorm], outcomes2[:, idxGammaGTAbnorm], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta3", thetas[2][:, idxGammaGTAbnorm], outcomes2[:, idxGammaGTAbnorm], thetaLabels, outcomeLabels2)
    nowRow[0] += 2
    idxBad = before[4, idxHCValid] >= 85
    ws1.cell(row=nowRow[0] - 1, column=1, value="Targets with abnormal DBP (>=85) (n={})".format(idxBad.sum())).font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxHCValid][:, idxBad], outcomes2[:, idxHCValid][:, idxBad], featureLabels, outcomeLabels2)
    func4(ws1, ws2, "theta1", theta1[:, idxHCValid][:, idxBad], outcomes2[:, idxHCValid][:, idxBad], thetaLabels, outcomeLabels2)
    func4(ws1, ws2, "fkr_category", fkr[:, idxHCValid][:, idxBad], outcomes2[:, idxHCValid][:, idxBad], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta2", thetas[1][:, idxHCValid][:, idxBad], outcomes2[:, idxHCValid][:, idxBad], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta3", thetas[2][:, idxHCValid][:, idxBad], outcomes2[:, idxHCValid][:, idxBad], thetaLabels, outcomeLabels2)
    nowRow[0] += 2
    ws1.cell(row=nowRow[0] - 1, column=1, value="Targets with abnormal values").font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxHCValid], outcomes2[:, idxHCValid], featureLabels, outcomeLabels2, idxs=idxAbnormals)
    func4(ws1, ws2, "theta1", theta1[:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2, idxs=idxAbnormals)
    func4(ws1, ws2, "fkr_category", fkr[:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2, idxs=idxAbnormals)
    # func4(ws1, ws2, "theta2", thetas[1][:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2, idxs=idxAbnormals)
    # func4(ws1, ws2, "theta3", thetas[2][:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2, idxs=idxAbnormals)
    # nowRow[0] += 2
    # ws1.cell(row=nowRow[0] - 1, column=1, value="Data recorded by skilled advisors(c_num >= {:<.1f})".format(aveCNum)).font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxCNum0], outcomes[:, idxCNum0], featureLabels, outcomeLabels)
    # func4(ws1, ws2, "theta1", theta1[:, idxCNum0], outcomes[:, idxCNum0], thetaLabels, outcomeLabels)
    # func4(ws1, ws2, "fkr_category", fkr[:, idxCNum0], outcomes[:, idxCNum0], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta2", thetas[1][:, idxCNum0], outcomes[:, idxCNum0], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta3", thetas[2][:, idxCNum0], outcomes[:, idxCNum0], thetaLabels, outcomeLabels)
    # nowRow[0] += 2
    # ws1.cell(row=nowRow[0] - 1, column=1, value="Data recorded by non-skilled advisors").font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxCNum1], outcomes[:, idxCNum1], featureLabels, outcomeLabels)
    # func4(ws1, ws2, "theta1", theta1[:, idxCNum1], outcomes[:, idxCNum1], thetaLabels, outcomeLabels)
    # func4(ws1, ws2, "fkr_category", fkr[:, idxCNum1], outcomes[:, idxCNum1], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta2", thetas[1][:, idxCNum1], outcomes[:, idxCNum1], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta3", thetas[2][:, idxCNum1], outcomes[:, idxCNum1], thetaLabels, outcomeLabels)
    # nowRow[0] += 2
    # ws1.cell(row=nowRow[0] - 1, column=1, value="Data with good outcomes(dWaist >= {:<.2f})".format(aveWaistD)).font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxD0], outcomes[:, idxD0], featureLabels, outcomeLabels)
    # func4(ws1, ws2, "theta1", theta1[:, idxD0], outcomes[:, idxD0], thetaLabels, outcomeLabels)
    # func4(ws1, ws2, "fkr_category", fkr[:, idxD0], outcomes[:, idxD0], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta2", thetas[1][:, idxD0], outcomes[:, idxD0], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta3", thetas[2][:, idxD0], outcomes[:, idxD0], thetaLabels, outcomeLabels)
    # nowRow[0] += 2
    # ws1.cell(row=nowRow[0] - 1, column=1, value="Data with bad outcomes").font = Font(color="dc143c")
    # func4(ws1, ws2, "features", features[:, idxD1], outcomes[:, idxD1], featureLabels, outcomeLabels)
    # func4(ws1, ws2, "theta1", theta1[:, idxD1], outcomes[:, idxD1], thetaLabels, outcomeLabels)
    # func4(ws1, ws2, "fkr_category", fkr[:, idxD1], outcomes[:, idxD1], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta2", thetas[1][:, idxD1], outcomes[:, idxD1], thetaLabels, outcomeLabels)
    # # func4(ws1, ws2, "theta3", thetas[2][:, idxD1], outcomes[:, idxD1], thetaLabels, outcomeLabels)
    nowRow[0] += 2
    ws1.cell(row=nowRow[0] - 1, column=1, value="Before and Direction").font = Font(color="dc143c")
    func4(ws1, ws2, "", theta1[:, idxHCValid], before[:, idxHCValid], thetaLabels, HCLabel)
    nowRow[0] += 2
    ws1.cell(row=nowRow[0] - 1, column=1, value="Before and Outcome").font = Font(color="dc143c")
    func4(ws1, ws2, "", outcomes2[:, idxHCValid], before[:, idxHCValid], outcomeLabels2, HCLabel)
    nowRow[0] += 2
    ws1.cell(row=nowRow[0] - 1, column=1, value="Correlations with directions and outcomes_fixed").font = Font(color="dc143c")
    func4(ws1, ws2, "features", features[:, idxHCValid], outcomes_fixed[:, idxHCValid], featureLabels, outcomeLabels2)
    func4(ws1, ws2, "theta1", theta1[:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta2", thetas[1][:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta3", thetas[2][:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta4", thetas[3][:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    func4(ws1, ws2, "fkr_category", fkr[:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta2", thetas[1][:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    # func4(ws1, ws2, "theta3", thetas[2][:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    nowRow[0] += 2
    ws1.cell(row=nowRow[0] - 1, column=1, value="Before and Outcome_fixed").font = Font(color="dc143c")
    func4(ws1, ws2, "", outcomes_fixed[:, idxHCValid], before[:, idxHCValid], outcomeLabels2, HCLabel)


def makeCoef2Sheet(wb):
    def func5(ws1, name, a, b, aLabel, bLabel, idxs=None):
        ws1.cell(row=nowRow[0], column=1, value=name).font = Font(color="0000cd")
        for j in range(len(aLabel)):
            ws1.cell(row=nowRow[0], column=j + 2, value=aLabel[j]).font = Font(color="006400")
        for i in range(len(bLabel)):
            ws1.cell(row=nowRow[0] + i + 1, column=1, value=bLabel[i]).font = Font(color="006400")
        for i in range(len(bLabel)):
            for j in range(len(aLabel)):
                indexes = a[j] != 0
                a2 = a[j, indexes]
                b2 = b[i, indexes]
                if(len(a2) > 1 and len(b2) > 1):
                    r, p = stats.pearsonr(a2, b2)
                    if(idxs is None or idxs[i] is None):
                        a2 = a[j, indexes]
                        b2 = b[i, indexes]
                        r, p = stats.pearsonr(a2, b2)
                    else:
                        idx = np.logical_and(indexes, idxs[i])
                        if(np.sum(idx) < 2):
                            r = None
                            p = 0
                        else:
                            a2 = a[j, idx]
                            b2 = b[i, idx]
                            r, p = stats.pearsonr(a2, b2)
                        p = "{}, (n={})".format(p, idx.sum())
                    ws1.cell(row=nowRow[0] + i + 1, column=j + 2, value=r)
                    ws1.cell(row=nowRow[0] + i + 1, column=j + 2).comment = openpyxl.comments.Comment(p, None)
        setColorScale(ws1, nowRow[0] + 1, 2, nowRow[0] + len(bLabel) + 1, len(aLabel) + 2)
        nowRow[0] += len(bLabel) + 1
    ws = wb.create_sheet("coef2")
    nowRow[0] = 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Correlations with directions and outcomes (removed unrelated targets)").font = Font(color="dc143c")
    func5(ws, "theta1", theta1[:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2)
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Correlations with directions and outcomes_fixed (removed unrelated targets)").font = Font(color="dc143c")
    func5(ws, "theta1", theta1[:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2)
    # nowRow[0] += 2
    # idxBad = before[4, idxHCValid] > 85
    # ws.cell(row=nowRow[0] - 1, column=1, value="Data with abnormal DBP (removed unrelated targets)").font = Font(color="dc143c")
    # func5(ws, "theta1", theta1[:, idxHCValid][:, idxBad], outcomes_fixed[:, idxHCValid][:, idxBad], thetaLabels, outcomeLabels2)
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with abnormal values (removed unrelated targets)").font = Font(color="dc143c")
    func5(ws, "theta1", theta1[:, idxHCValid], outcomes2[:, idxHCValid], thetaLabels, outcomeLabels2, idxs=idxAbnormals)
    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Data with abnormal values (fixed)(removed unrelated targets)").font = Font(color="dc143c")
    func5(ws, "theta1", theta1[:, idxHCValid], outcomes_fixed[:, idxHCValid], thetaLabels, outcomeLabels2, idxs=idxAbnormals)
    nowRow[0] += 2
    ws.cell(row=nowRow[0], column=2, value="total").font = Font(color="dc143c")
    ws.cell(row=nowRow[0] + 1, column=1, value="total").font = Font(color="dc143c")
    ws.cell(row=nowRow[0] + 1, column=2, value=len(idxHCValid))
    for j in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=j + 3, value=thetaLabels[j]).font = Font(color="006400")
        ws.cell(row=nowRow[0] + 1, column=j + 3, value=np.sum((theta1[j] != 0)[idxHCValid]))
    for i in range(len(outcomeLabels2)):
        ws.cell(row=nowRow[0] + i + 2, column=1, value=outcomeLabels2[i]).font = Font(color="006400")
        if(idxAbnormals[i] is not None):
            ws.cell(row=nowRow[0] + i + 2, column=2, value=np.sum(idxAbnormals[i]))
    for i in range(len(outcomeLabels2)):
        for j in range(len(thetaLabels)):
            if(idxAbnormals[i] is not None):
                idx = np.logical_and((theta1[j] != 0)[idxHCValid], idxAbnormals[i])
                ws.cell(row=nowRow[0] + i + 2, column=j + 3, value=np.sum(idx))
    nowRow[0] += len(outcomeLabels2) + 2

    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Before with abnormal values (removed unrelated targets)").font = Font(color="dc143c")
    func5(ws, "theta1", theta1[:, idxHCValid], before[:, idxHCValid], thetaLabels, HCLabel, idxs=idxAbnormals)

    nowRow[0] += 2
    ws.cell(row=nowRow[0] - 1, column=1, value="Features before direction (removed unrelated targets)").font = Font(color="dc143c")
    for j in range(len(thetaLabels)):
        ws.cell(row=nowRow[0], column=j + 2, value=thetaLabels[j]).font = Font(color="006400")
    for i in range(len(HCLabel)):
        ws.cell(row=nowRow[0] + i + 1, column=1, value=HCLabel[i]).font = Font(color="006400")
        for j in range(len(thetaLabels)):
            indexes = theta1[j, idxHCValid] != 0
            b = before[i, idxHCValid][indexes].astype("float64")
            ttest = stats.ttest_ind(b, before[i, idxHCValid], equal_var=False)
            ws.cell(row=nowRow[0] + i + 1, column=j + 2, value=np.average(b)).comment = openpyxl.comments.Comment(ttest[1], None)
            if(ttest[1] <= 0.01):
                ws.cell(row=nowRow[0] + i + 1, column=j + 2).fill = fill1 if ttest[0] > 0 else fill3
            elif(ttest[1] <= 0.05):
                ws.cell(row=nowRow[0] + i + 1, column=j + 2).fill = fill2 if ttest[0] > 0 else fill4
    # ws.cell(row=nowRow[0] - 1, column=1, value="Features outcome direction (removed unrelated targets)").font = Font(color="dc143c")
    # for j in range(len(thetaLabels)):
    #     ws.cell(row=nowRow[0], column=j + 2, value=thetaLabels[j]).font = Font(color="006400")
    # for i in range(len(HCLabel)):
    #     ws.cell(row=nowRow[0] + i + 1, column=1, value=HCLabel[i]).font = Font(color="006400")
    #     for j in range(len(thetaLabels)):
    #         indexes = theta1[j, idxHCValid] != 0
    #         b = outcomes2[i, idxHCValid][indexes].astype("float64")
    #         ttest = stats.ttest_ind(b, outcomes2[i, idxHCValid], equal_var=False)
    #         ws.cell(row=nowRow[0] + i + 1, column=j + 2, value=np.average(b)).comment = openpyxl.comments.Comment(ttest[1], None)
    #         if(ttest[1] <= 0.01):
    #             ws.cell(row=nowRow[0] + i + 1, column=j + 2).fill = fill1 if ttest[0] > 0 else fill3
    #         elif(ttest[1] <= 0.05):
    #             ws.cell(row=nowRow[0] + i + 1, column=j + 2).fill = fill2 if ttest[0] > 0 else fill4


def make2dSheet(wb):
    before_s = stats.zscore(before[:, idxHCValid].astype("float64"), axis=1)
    theta_s = stats.zscore(theta1[:, idxHCValid].astype("float64"), axis=1)
    a3d = np.zeros((len(idxHCValid), len(HCLabel), len(thetaLabels)))
    for i, idx in enumerate(idxHCValid):
        a3d[i] = np.tensordot(before_s[:, i], theta_s[:, i], axes=0)

    ws = wb.create_sheet("2d_coef")
    nowRow[0] = 3
    for nOutcome in range(len(outcomeLabels2)):
        name = outcomeLabels2[nOutcome]
        y = outcomes2[nOutcome, idxHCValid].astype("float64")
        for i in range(1, 21):
            ws.cell(row=nowRow[0] - 2, column=i).fill = fill5
        ws.cell(row=nowRow[0] - 1, column=1, value="correlation between outcome value ({}) and each feature-direction pair".format(name)).font = Font(color="dc143c")
        for j in range(len(thetaLabels)):
            ws.cell(row=nowRow[0], column=j + 2, value=thetaLabels[j]).font = Font(color="006400")
        for i in range(len(HCLabel)):
            ws.cell(row=i + nowRow[0] + 1, column=1, value=HCLabel[i]).font = Font(color="006400")
            for j in range(len(thetaLabels)):
                r, p = stats.pearsonr(a3d[:, i, j], y)
                ws.cell(row=i + nowRow[0] + 1, column=j + 2, value=r).comment = openpyxl.comments.Comment(p, None)
                # if(p < 0.0001):
                #     plt.figure(figsize=(10, 10))
                #     plt.scatter(a3d[:, i, j], outcomes2[nOutcome, idxHCValid])
                #     plt.xlabel(HCLabel[i] + " - " + thetaLabels[j])
                #     plt.ylabel(outcomeLabels2[nOutcome])
                #     plt.savefig(str(pathLDAResultFolder.joinpath("scatter", "scatter_({}-{})-{}.png".format(HCLabel[i], thetaLabels[j], outcomeLabels2[nOutcome]))))
                #     plt.close()
        ws.cell(row=len(HCLabel) + nowRow[0] + 1, column=1, value="correlation between outcome value ({}) and each direction".format(name)).font = Font(color="dc143c")
        for j in range(len(thetaLabels)):
            ws.cell(row=len(HCLabel) + nowRow[0] + 2, column=j + 2, value=stats.pearsonr(theta1[j, idxHCValid], y)[0])
        setColorScale(ws, nowRow[0], 1, len(HCLabel) + nowRow[0] + 2, len(thetaLabels) + 1)
        nowRow[0] += len(HCLabel) + 5

    # # save scatter
    # i = 8
    # j = 2
    # k = 8
    # plt.figure(figsize=(10, 10))
    # # plt.scatter(a3d[:, i, j], outcomes2[k, idxHCValid])
    # indexes = np.full((len(idxFull)), False)
    # indexes[idxGammaGTAbnorm] = True
    # indexes = indexes[idxHCValid]
    # plt.scatter(a3d[indexes, i, j], outcomes2[k, idxGammaGTAbnorm])
    # plt.xlabel(HCLabel[i] + " - " + thetaLabels[j])
    # plt.ylabel(outcomeLabels2[k])
    # plt.savefig(str(pathLDAResultFolder.joinpath("scatter_({}-{})-{}.png".format(HCLabel[i], thetaLabels[j], outcomeLabels2[k]))))
    # plt.close()

    # model = sklearn.linear_model.LinearRegression(fit_intercept=True, normalize=False, copy_X=True, n_jobs=1)
    # converter = converter32(len(HCLabel), len(thetaLabels))
    # X = converter.con(a3d)
    # y = -outcomes2[10, idxHCValid].astype("float64")
    # model.fit(X, y)
    # ws = wb.create_sheet("2d_regression")
    # coef = model.coef_.reshape(len(HCLabel), len(thetaLabels))
    # for j in range(len(thetaLabels)):
    #     ws.cell(row=1, column=j + 2, value=thetaLabels[j]).font = Font(color="006400")
    # for i in range(len(HCLabel)):
    #     ws.cell(row=i + 2, column=1, value=HCLabel[i]).font = Font(color="006400")
    #     for j in range(len(thetaLabels)):
    #         ws.cell(row=i + 2, column=j + 2, value=coef[i, j])
    # p1 = openpyxl.utils.cell.get_column_letter(1) + "1:" + openpyxl.utils.cell.get_column_letter(1 + len(thetaLabels)) + str(len(HCLabel) + 1)
    # cs = ColorScale(cfvo=[FormatObject(type='min'), FormatObject(type='num', val=0), FormatObject(type='max')],
    #                 color=[Color('80FF00'), Color('FFFFFF'), Color('FF8000')])
    # ws.conditional_formatting.add(p1, Rule(type='colorScale', colorScale=cs))


def makeAnovaSheet(wb):
    ws = wb.create_sheet("anova")
    nowRow[0] = 2
    sig = DF_ANOVA < 1e-10
    for i, label in enumerate(HCLabel):
        ws.cell(row=nowRow[0] + 2 * i, column=1, value=label)
        if(label in sig.columns):
            eff_habit = sig[sig[label]].index.tolist()
            eff_direction = list(set([relation_habit_direction[habit] for habit in eff_habit]))
            setValuesFromArray(eff_habit, ws, nowRow[0] + 2 * i, 2)
            setValuesFromArray(eff_direction, ws, nowRow[0] + 2 * i + 1, 2)
    nowRow[0] += 2 * len(HCLabel) + 3

    def func(outcomes):
        for n, label in enumerate(HCLabel):
            ws.cell(row=nowRow[0], column=n + 1, value=outcomeLabels2[n])
            if(abnormalValues[n] is not None):
                eff_habit = sig[sig[label]].index.tolist()
                eff_direction = list(set([relation_habit_direction[habit] for habit in eff_habit]))
                dr = np.zeros(np.sum(idxAbnormals[n]))
                for ed in eff_direction:
                    dr += theta1[thetaLabels.index(ed), idxHCValid][idxAbnormals[n]]
                oc = outcomes[n, idxHCValid][idxAbnormals[n]]
                r, p = stats.pearsonr(dr, oc)
                ws.cell(row=nowRow[0] + 1, column=n + 1, value=r).comment = openpyxl.comments.Comment(p, None)
        nowRow[0] += 4

    ws.cell(row=nowRow[0], column=1, value="coefs between outcome and effective direction point(decided with anova)")
    nowRow[0] += 1
    func(outcomes2)
    ws.cell(row=nowRow[0], column=1, value="coefs between outcome and effective direction point(decided with anova)(fixed)")
    nowRow[0] += 1
    func(outcomes_fixed)


def makeLeeSheet(wb):
    # 対象文書がano_p_messageのときのみ
    # MIN_DIRECTION_TIMES=2とすること

    ws = wb.create_sheet("lee")

    def func3(ws, a, b, name, compare=None):
        ave = np.average(a, axis=1)
        ws.cell(row=nowRow[0], column=1, value=name).font = Font(color="0000cd")
        for i in range(len(b)):
            ws.cell(row=nowRow[0], column=i + 2, value=b[i]).font = Font(color="006400")
            ws.cell(row=nowRow[0] + 1, column=i + 2, value=ave[i])
            if(compare is not None):
                aaa = a[i].astype("float64")
                bbb = compare[i].astype("float64")
                ttest = stats.ttest_ind(aaa, bbb, equal_var=False)
                ws.cell(row=nowRow[0] + 1, column=i + 2).comment = openpyxl.comments.Comment(ttest[1], None)
                if(ttest[1] <= 0.01):
                    ws.cell(row=nowRow[0] + 1, column=i + 2).fill = fill1 if ttest[0] > 0 else fill3
                elif(ttest[1] <= 0.05):
                    ws.cell(row=nowRow[0] + 1, column=i + 2).fill = fill2 if ttest[0] > 0 else fill4
        nowRow[0] += 2

    if(MIN_DIRECTION_TIMES < 4):
        idx0 = np.argsort(outcomes2[2, :])[:300]
        idx1 = np.argsort(outcomes2[2, :])[-300:]
        nowRow[0] = 4
        for i in range(1, 21):
            ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
        ws.cell(row=nowRow[0] - 1, column=1, value=f"Data with good outcomes(dBMI)(n={len(idx0)})(pnum>=2)").font = Font(color="dc143c")
        func3(ws, features[:, idx0], featureLabels, "features", features[:, idx1])
        func3(ws, before[:, listAnd(idxHCValid, idx0)], HCLabel, "before", before[:, listAnd(idxHCValid, idx1)])
        func3(ws, after[:, listAnd(idxHCValid, idx0)], HCLabel, "after", after[:, listAnd(idxHCValid, idx1)])
        func3(ws, outcomes[:, idx0], outcomeLabels, "outcomes", outcomes[:, idx1])
        func3(ws, outcomes2[:, listAnd(idxHCValid, idx0)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idx1)])
        func3(ws, theta1[:, idx0], thetaLabels, "theta1", theta1[:, idx1])
        func3(ws, fkr[:, idx0], thetaLabels, "fkr_category", fkr[:, idx1])
        nowRow[0] += 2
        ws.cell(row=nowRow[0] - 1, column=1, value=f"Data with bad outcomes(dBMI)(n={len(idx1)})(pnum>=2)").font = Font(color="dc143c")
        func3(ws, features[:, idx1], featureLabels, "features", features[:, idx0])
        func3(ws, before[:, listAnd(idxHCValid, idx1)], HCLabel, "before", before[:, listAnd(idxHCValid, idx0)])
        func3(ws, after[:, listAnd(idxHCValid, idx1)], HCLabel, "after", after[:, listAnd(idxHCValid, idx0)])
        func3(ws, outcomes[:, idx1], outcomeLabels, "outcomes", outcomes[:, idx0])
        func3(ws, outcomes2[:, listAnd(idxHCValid, idx1)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idx0)])
        func3(ws, theta1[:, idx1], thetaLabels, "theta1", theta1[:, idx0])
        func3(ws, fkr[:, idx1], thetaLabels, "fkr_category", fkr[:, idx0])

    else:
        if(numTarget >= 200):
            idx0 = np.argsort(outcomes2[2, :])[:100]
            idx1 = np.argsort(outcomes2[2, :])[-100:]
        else:
            idx0 = np.argsort(outcomes2[2, :])[:(numTarget // 2)]
            idx1 = np.argsort(outcomes2[2, :])[-(numTarget // 2):]

        nowRow[0] = 4
        for i in range(1, 21):
            ws.cell(row=nowRow[0] - 3, column=i).fill = fill5
        ws.cell(row=nowRow[0] - 1, column=1, value=f"Data with good outcomes(dBMI)(n={len(idx0)})(pnum>=4)").font = Font(color="dc143c")
        func3(ws, features[:, idx0], featureLabels, "features", features[:, idx1])
        func3(ws, before[:, listAnd(idxHCValid, idx0)], HCLabel, "before", before[:, listAnd(idxHCValid, idx1)])
        func3(ws, after[:, listAnd(idxHCValid, idx0)], HCLabel, "after", after[:, listAnd(idxHCValid, idx1)])
        func3(ws, outcomes[:, idx0], outcomeLabels, "outcomes", outcomes[:, idx1])
        func3(ws, outcomes2[:, listAnd(idxHCValid, idx0)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idx1)])
        func3(ws, thetas[0][:, idx0], thetaLabels, "theta1", thetas[0][:, idx1])
        func3(ws, thetas[1][:, idx0], thetaLabels, "theta2", thetas[1][:, idx1])
        func3(ws, thetas[2][:, idx0], thetaLabels, "theta3", thetas[2][:, idx1])
        func3(ws, thetas[3][:, idx0], thetaLabels, "theta4", thetas[3][:, idx1])
        func3(ws, fkr[:, idx0], thetaLabels, "fkr_category", fkr[:, idx1])
        nowRow[0] += 2
        ws.cell(row=nowRow[0] - 1, column=1, value=f"Data with bad outcomes(dBMI)(n={len(idx1)})(pnum>=4)").font = Font(color="dc143c")
        func3(ws, features[:, idx1], featureLabels, "features", features[:, idx0])
        func3(ws, before[:, listAnd(idxHCValid, idx1)], HCLabel, "before", before[:, listAnd(idxHCValid, idx0)])
        func3(ws, after[:, listAnd(idxHCValid, idx1)], HCLabel, "after", after[:, listAnd(idxHCValid, idx0)])
        func3(ws, outcomes[:, idx1], outcomeLabels, "outcomes", outcomes[:, idx0])
        func3(ws, outcomes2[:, listAnd(idxHCValid, idx1)], outcomeLabels2, "outcomes2", outcomes2[:, listAnd(idxHCValid, idx0)])
        func3(ws, thetas[0][:, idx1], thetaLabels, "theta1", thetas[0][:, idx0])
        func3(ws, thetas[1][:, idx1], thetaLabels, "theta2", thetas[1][:, idx0])
        func3(ws, thetas[2][:, idx1], thetaLabels, "theta3", thetas[2][:, idx0])
        func3(ws, thetas[3][:, idx1], thetaLabels, "theta4", thetas[3][:, idx0])
        func3(ws, fkr[:, idx1], thetaLabels, "fkr_category", fkr[:, idx0])


def tukey_hsd(ind, *args):
    data_arr = np.hstack(args)
    ind_arr = np.array([])
    for x in range(len(args)):
        ind_arr = np.append(ind_arr, np.repeat(ind[x], len(args[x])))
    return pairwise_tukeyhsd(data_arr, ind_arr)


def getCellString(row, column):
    return openpyxl.utils.cell.get_column_letter(column) + str(row)


def getRangeString(startRow, startColumn, endRow, endColumn):
    return getCellString(startRow, startColumn) + ":" + getCellString(endRow, endColumn)


def setColorScale(ws, startRow, startColumn, endRow, endColumn):
    area = getRangeString(startRow, startColumn, endRow, endColumn)
    cs = ColorScale(cfvo=[FormatObject(type='min'), FormatObject(type='num', val=0), FormatObject(type='max')],
                    color=[Color('80FF00'), Color('FFFFFF'), Color('FF8000')])
    ws.conditional_formatting.add(area, Rule(type='colorScale', colorScale=cs))
    for row in range(startRow, endRow + 1):
        for column in range(startColumn, endColumn + 1):
            ws.cell(row=row, column=column).number_format = "#0.000"


def setValuesFromArray(array, ws, startRow, startColumn, axis="column"):
    if(axis == "row"):
        for i in range(len(array)):
            ws.cell(row=startRow + i, column=startColumn, value=array[i])
    elif(axis == "column"):
        for j in range(len(array)):
            ws.cell(row=startRow, column=startColumn + j, value=array[j])
    else:
        print("invalid parameter--------------------------------------------")


def listAnd(l0, l1):
    return list(set(l0) & set(l1))


def listOr(l0, l1):
    return list(set(l0) | set(l1))


class converter32:

    def __init__(self, a, b):
        self.a = a
        self.b = b

    def con(self, a3d):
        return a3d.reshape(a3d.shape[0], self.a * self.b)

    def dcon(self, a2d):
        return a2d.reshape(a2d.shape[0], self.a, self.b)
