import pandas as pd
import numpy as np
import statsmodels.api as sm
from statsmodels.formula.api import ols
import csv
import openpyxl


# 年齢, メタボリックシンドローム判定
dataLabels = ["体重", "腹囲", "ＢＭＩ",
              "空腹時血糖", "ＨｂＡ１ｃ（ＮＧＳＰ）",
              "拡張期血圧", "収縮期血圧",
              "中性脂肪", "ＨＤＬコレステロール", "ＬＤＬコレステロール",
              "γ－ＧＴ（γ－ＧＴＰ）", "ＧＯＴ（ＡＳＴ）", "ＧＰＴ（ＡＬＴ）"]
habitLabels = ["３０分以上の運動習慣", "歩行又は身体活動", "歩行速度",
               "食べ方２（就寝前）", "食べ方３（夜食／間食）", "食習慣", "食べ方１（早食い等）",
               "飲酒", "飲酒量",
               "喫煙", "睡眠"]
habitLevels = [["いいえ", "はい"], ["いいえ", "はい"], ["いいえ", "はい"],
               ["いいえ", "はい"], ["いいえ", "はい"], ["いいえ", "はい"], ["遅い", "ふつう", "速い"],
               ["ほとんど飲まない", "時々", "毎日"], ["", "１合未満", "１～２合未満", "２～３合未満", "３合以上"],
               ["いいえ", "はい"], ["いいえ", "はい"]]
# habitLabels = ["３０分以上の運動習慣", "食べ方３（夜食／間食）", "飲酒", "喫煙"]
# habitLevels = [["いいえ", "はい"], ["いいえ", "はい"], ["ほとんど飲まない", "時々", "毎日"], ["いいえ", "はい"]]
medicineLabels = ["服薬１（血圧）", "服薬２（血糖）", "服薬３（脂質）"]
data_medicine_relation = [None, None, None, 1, 1, 0, 0, 2, 2, 2, None, None, None]
medicineLevels = [["いいえ", "2.0"], ["いいえ", "2.0"], ["いいえ", "2.0"]]

habitLabels_jp_replaced = [f"h{i}" for i in range(len(habitLabels))]
dataLabels_jp_replaced = [f"d{i}" for i in range(len(dataLabels))]
medicineLabels_jp_replaced = [f"m{i}" for i in range(len(medicineLabels))]


def analyze(pathHealthCheck, pathAnalysisHealthCheck):
    df = pd.read_csv(pathHealthCheck)
    df = df[dataLabels + habitLabels + medicineLabels]
    df["飲酒量"] = df["飲酒量"].fillna("")
    df = df.dropna()

    # check measured data valid
    for label in dataLabels:
        df[label] = pd.to_numeric(df[label], errors='coerce')
    df = df.dropna()
    df_jp_replaced = df.copy()

    # chech habit valid
    habit_valid = np.full(len(df), True)
    for n, label in enumerate(habitLabels):
        v = np.full(len(df), False)
        for m, level in enumerate(habitLevels[n]):
            v = v | (df[label] == level)
            df_jp_replaced[label][df_jp_replaced[label] == level] = m
        habit_valid = habit_valid & v
    df = df[habit_valid]
    df_jp_replaced = df_jp_replaced[habit_valid]

    # check medicine valid
    medicine_valid = np.full(len(df), True)
    for n, label in enumerate(medicineLabels):
        v = np.full(len(df), False)
        for m, level in enumerate(medicineLevels[n]):
            v = v | (df[label] == level)
            df_jp_replaced[label][df_jp_replaced[label] == level] = m
        medicine_valid = medicine_valid & v
    df = df[medicine_valid]
    df_jp_replaced = df_jp_replaced[medicine_valid]

    df_jp_replaced = df_jp_replaced.rename(columns={dataLabels[i]: dataLabels_jp_replaced[i] for i in range(len(dataLabels))})
    df_jp_replaced = df_jp_replaced.rename(columns={habitLabels[i]: habitLabels_jp_replaced[i] for i in range(len(habitLabels))})
    df_jp_replaced = df_jp_replaced.rename(columns={medicineLabels[i]: medicineLabels_jp_replaced[i] for i in range(len(medicineLabels))})
    df_jp_replaced = df_jp_replaced.astype(float)

    wb = openpyxl.Workbook()

    histHabit(df, wb)

    histMedicine(df, wb)

    anovaAll(df_jp_replaced, wb, pathAnalysisHealthCheck.joinpath("anova_p.csv"), medi_check=False)

    wb.save(str(pathAnalysisHealthCheck.joinpath("hc_analysi.xlsx")))


def histHabit(df, wb):
    ws = wb.create_sheet("hist habit")
    ws.cell(row=1, column=1, value="habit")
    ws.cell(row=1, column=2, value="level")
    ws.cell(row=1, column=3, value="count")
    row = 2
    for n, label in enumerate(habitLabels):
        series = df[label]
        series = series.fillna("")
        ws.cell(row=row, column=1, value=label)

        for level in habitLevels[n]:
            c = np.sum(series == level)
            ws.cell(row=row, column=2, value=level)
            ws.cell(row=row, column=3, value=c)
            row += 1


def histMedicine(df, wb):
    ws = wb.create_sheet("medicine habit")
    ws.cell(row=1, column=1, value="medicine")
    ws.cell(row=1, column=2, value="level")
    ws.cell(row=1, column=3, value="count")
    row = 2
    for n, label in enumerate(medicineLabels):
        series = df[label]
        ws.cell(row=row, column=1, value=label)

        for level in medicineLevels[n]:
            c = np.sum(series == level)
            ws.cell(row=row, column=2, value=level)
            ws.cell(row=row, column=3, value=c)
            row += 1

    ws.cell(row=row, column=1, value="全ていいえ")
    ws.cell(row=row, column=3, value=np.sum(np.any(df[medicineLabels] == "いいえ", axis=1)))
    row += 1
    ws.cell(row=row, column=1, value="全てはい")
    ws.cell(row=row, column=3, value=np.sum(np.any(df[medicineLabels] == "2.0", axis=1)))
    row += 1


def anovaAll(df_jp_replaced, wb, path, medi_check=True):
    ws_d = wb.create_sheet("anova data-habit detail")
    p = np.zeros((len(habitLabels), len(dataLabels)))
    row = 2
    for n, label in enumerate(dataLabels_jp_replaced):
        if(data_medicine_relation[n] is None or not medi_check):
            df = df_jp_replaced
        else:
            df = df_jp_replaced[df_jp_replaced[medicineLabels_jp_replaced[data_medicine_relation[n]]] == 0]
        result = anova(df, label)
        result = result.rename(index={habitLabels_jp_replaced[i]: habitLabels[i] for i in range(len(habitLabels))})
        p[:, n] = result["PR(>F)"][0:len(habitLabels)]

        ws_d.cell(row=row, column=1, value=dataLabels[n] + f" n={len(df)}-------------------------------------------------------------------")
        for j in range(len(result.columns)):
            ws_d.cell(row=row + 1, column=j + 2, value=result.columns[j])
        for i in range(len(result.index)):
            ws_d.cell(row=row + i + 2, column=1, value=result.index[i])
            for j in range(len(result.columns)):
                ws_d.cell(row=row + i + 2, column=j + 2, value=result.iloc[i, j])
        row += len(result.index) + 4

    ws_s = wb.create_sheet("anova data-habit summary")
    for j in range(len(dataLabels)):
        ws_s.cell(row=1, column=j + 2, value=dataLabels[j])
    for i in range(len(habitLabels)):
        ws_s.cell(row=i + 2, column=1, value=habitLabels[i])
        for j in range(len(dataLabels)):
            ws_s.cell(row=i + 2, column=j + 2, value=p[i, j])

    with open(str(path), "w", encoding="utf_8_sig") as f:
        writer = csv.writer(f)
        writer.writerow([""] + dataLabels)
        for i in range(len(habitLabels)):
            writer.writerow([habitLabels[i]] + p[i, :].tolist())


def anova(df_jp_replaced, outcome_key):
    # formula = outcome_key + "~" + "*".join(habitLabels_jp_replaced)
    formula = outcome_key + "~" + "+".join(habitLabels_jp_replaced)
    npk_lm = ols(formula, data=df_jp_replaced).fit()
    result = sm.stats.anova_lm(npk_lm, typ=2)
    # result = result[result["PR(>F)"] < 1e-10]
    return result
