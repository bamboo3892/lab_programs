import pickle
import json
import numpy as np
import matplotlib.pyplot as plt
import torch
import openpyxl
import cv2
import multiprocessing
import copy

from analysis.analyzeHealthCheck import habitLevels2, habitWorstLevels2, medicineLevels, medicineWorstLevels
import LDA_pytorch.LDA as LDA
import LDA_pytorch.MCLDA as MCLDA
from utils.general_util import Args, min_max_normalize, simple_moving_average
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix
from utils.graphic_util import makeColorMap, addImage, DEFAULT_COLOR_FUNC, DEFAULT_COLOR_FUNC2


seed = 1
np.random.seed(seed)
torch.manual_seed(seed)
DEVICE = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")


def excuteLDA(pathDocs, pathResult, *,
              args=None, summary_args=None):

    args = args if args is not None else Args()
    summary_args = summary_args if summary_args is not None else Args()

    with open(str(pathDocs), "r", encoding="utf_8_sig") as f:
        docs = json.load(f)

    model_class = LDA.LDA
    args.modelType = "LDA"
    args.random_seed = seed
    args.num_steps = 500
    args.step_subsample = 100000
    args.K = 7
    args.auto_beta = False
    args.auto_alpha = False
    args.coef_beta = 1
    args.coef_alpha = 1

    summary_args.morphome_key = "morphomes"
    summary_args.full_docs = docs

    data = [doc[summary_args.morphome_key] for doc in docs]
    print(f"coef_beta:   {args.coef_beta} (auto: {args.auto_beta})")
    print(f"coef_alpha:  {args.coef_alpha} (auto: {args.auto_alpha})")

    _excute(model_class, args, data, pathResult, summary_args)


def excuteMCLDA(pathDocs, pathTensors, pathResult, *,
                args=None, summary_args=None,
                pathTestdocs=None, pathTesttensors=None):

    args = args if args is not None else Args()
    summary_args = summary_args if summary_args is not None else Args()

    include_medicine = False
    documents, tensors, data = _make_data_1(pathDocs, pathTensors, include_medicine=include_medicine)
    testset = None
    if(pathTestdocs is not None and pathTesttensors is not None):
        _, __, testset = _make_data_1(pathTestdocs, pathTesttensors, include_medicine=include_medicine)

    model_class = MCLDA.MCLDA
    args.modelType = "MCLDA"
    args.random_seed = seed
    args.include_medicine = include_medicine
    args.num_steps = 1000
    args.step_subsample = 10
    args.K = 10
    args.D = len(data[0][0]) if len(data[0]) != 0 else (len(data[1][0]) if len(data[1]) != 0 else (len(data[2][0]) if len(data[2]) != 0 else 0))
    args.n_rh = [len(tensors["habit_levels"][rh]) for rh in range(len(tensors["habit_keys"]))]
    # args.n_rh = []
    args.auto_beta = False
    args.auto_alpha = False
    args.coef_beta = 1
    args.coef_alpha = 1
    args.nu_h = 1
    # args.deterministic_coefs = [None] * args.num_steps
    args.deterministic_coefs = np.linspace(0., 1, args.num_steps).tolist()

    summary_args.full_docs = documents
    summary_args.full_tensors = tensors
    summary_args.habitWorstLevels = habitWorstLevels2
    if(include_medicine):
        summary_args.habitWorstLevels += medicineWorstLevels
    print(f"D: {args.D}, K: {args.K}")
    print(f"coef_beta:   {args.coef_beta} (auto: {args.auto_beta})")
    print(f"coef_alpha:  {args.coef_alpha} (auto: {args.auto_alpha})")

    # _excute(model_class, args, data, pathResult, summary_args, testset=testset, from_pickle=False, do_hist_analysis=True)
    _excute(model_class, args, data, pathResult, summary_args, from_pickle=False, do_hist_analysis=True)
    # _excute(model_class, args, data, pathResult, summary_args, continue_from_pickle=True, do_hist_analysis=True)


def excuteMCLDA_K_range(pathDocs, pathTensors, pathResult, *,
                        pathTestdocs=None, pathTesttensors=None):

    args = Args()
    summary_args = Args()

    documents, tensors, data = _make_data_1(pathDocs, pathTensors)
    testset = None
    if(pathTestdocs is not None and pathTesttensors is not None):
        _, __, testset = _make_data_1(pathTestdocs, pathTesttensors)

    model_class = MCLDA.MCLDA
    args.modelType = "MCLDA"
    args.num_steps = 200
    args.step_subsample = 10
    args.D = len(data[0][0]) if len(data[0]) != 0 else (len(data[1][0]) if len(data[1]) != 0 else (len(data[2][0]) if len(data[2]) != 0 else 0))
    args.n_rh = [len(habitLevels2[rh]) for rh in range(len(tensors["habit_keys"]))]
    args.auto_beta = False
    args.auto_alpha = False
    args.coef_beta = 1
    args.coef_alpha = 1
    args.nu_h = 1
    args.deterministic_coefs = [None] * args.num_steps

    summary_args.full_docs = documents
    summary_args.full_tensors = tensors
    summary_args.habitWorstLevels = habitWorstLevels2

    # Ks = np.arange(1, 21, 1).tolist()
    # seeds = np.arange(0, 20).tolist()
    # fnames = [f"K{k}" for k in Ks]
    Ks = []
    seeds = []
    fnames = []
    # for seed in range(10):
    for seed in [0]:
        # for k in [1, 10, 20, 30, 40, 50]:
        for k in np.arange(1, 21):
            Ks.append(k)
            seeds.append(seed)
            fnames.append(f"K{k}_seed{seed}")

    # # simgle process
    # for i, K in enumerate(Ks):
    #     args.K = K
    #     args.random_seed = seeds[i]
    #     print(f"D: {args.D}, K: {args.K}")
    #     _excute(model_class, args, data, pathResult.joinpath(fnames[i]), summary_args,
    #             testset=testset, from_pickle=True, do_hist_analysis=False)
    #     # _excute(model_class, args, data, pathResult.joinpath(fnames[i]), summary_args,
    #     #         testset=None, from_pickle=True, do_hist_analysis=False)

    # multi process
    process_args = []
    for i in range(len(Ks)):
        args2 = copy.deepcopy(args)
        args2.K = Ks[i]
        args2.random_seed = seeds[i]
        summary_args2 = copy.deepcopy(summary_args)
        process_args.append([model_class, args2, data, pathResult.joinpath(fnames[i]), summary_args2,
                             testset, True, False, False])
    multiprocessing.set_start_method('spawn')
    processes = 3
    with multiprocessing.Pool(processes=processes, maxtasksperchild=1) as p:
        p.starmap(_excute, process_args, chunksize=(len(Ks) - 1) // processes + 1)

    accuracies_rt = []
    accuracies_rm = []
    accuracies_rh = []
    for i, K in enumerate(Ks):
        with open(str(pathResult.joinpath(fnames[i], "accuracy.json")), "r", encoding="utf_8_sig") as f:
            accuracy = json.load(f)
            accuracies_rt.append(accuracy["rt"])
            accuracies_rm.append(accuracy["rm"])
            accuracies_rh.append(accuracy["rh"])

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Rt")
    writeMatrix(ws, accuracies_rt, 1, 1,
                row_names=Ks,
                column_names=tensors["tensor_keys"],
                rule="databar")
    ws = wb.create_sheet("Rm")
    writeMatrix(ws, accuracies_rm, 1, 1,
                row_names=Ks,
                column_names=tensors["measurement_keys"],
                rule="databar")
    ws = wb.create_sheet("Rh")
    writeMatrix(ws, accuracies_rh, 1, 1,
                row_names=Ks,
                column_names=tensors["habit_keys"],
                rule="databar")
    wb.save(pathResult.joinpath("accuracy.xlsx"))


def _make_data_1(pathDocs, pathTensors, include_medicine=False):
    with open(str(pathDocs), "r", encoding="utf_8_sig") as f:
        documents = json.load(f)
    with open(str(pathTensors), 'rb') as f:
        tensors = pickle.load(f)

    morphomesKeys = tensors["tensor_keys"]
    measurementKeysHC = tensors["measurement_keys"]
    habitKeysHC = tensors["habit_keys"]
    tensors["habit_levels"] = habitLevels2
    if(include_medicine):
        habitKeysHC += tensors["medicine_keys"]
        tensors["habit_levels"] += medicineLevels

    docs = [[] for _ in range(len(morphomesKeys))]
    measurements = [[] for _ in range(len(measurementKeysHC))]
    habits = [[] for _ in range(len(habitKeysHC))]

    for doc in documents:
        pid = int(doc["p_seq"])
        if(pid in tensors["ids"]):
            idx = tensors["ids"].index(pid)
            if(tensors["ＧＯＴ（ＡＳＴ）"][idx] > 200):
                continue
            for rt, key in enumerate(morphomesKeys):
                docs[rt].append(doc[key + "_morphomes"])
            for rm, key in enumerate(measurementKeysHC):
                measurements[rm].append(tensors[key][idx])
            for rh, key in enumerate(habitKeysHC):
                habits[rh].append(tensors[key][idx])
    data = [docs, np.array(measurements), np.array(habits)]
    # data = [docs, np.empty([0, 0]), np.empty([0, 0])]
    # data = [[], np.array(measurements), np.empty([0, 0])]
    # data = [[], np.empty([0, 0]), np.array(habits)]

    return documents, tensors, data


def _excute(modelClass, args, data, pathResult, summary_args,
            testset=None, from_pickle=False, continue_from_pickle=False, do_hist_analysis=False):

    print(f"Model: {args.modelType}  (to {pathResult}) (from pickle: {from_pickle}) (continue from pickle: {continue_from_pickle})")

    args.device = DEVICE
    np.random.seed(args.random_seed)
    torch.manual_seed(args.random_seed)
    summary_args.summary_path = pathResult
    pathResult.mkdir(exist_ok=True, parents=True)

    """learning step or existing model"""
    if(from_pickle):
        model = torch.load(pathResult.joinpath("model.pickle"), args.device)
        history = torch.load(pathResult.joinpath("history.pickle"))
    else:
        if(not continue_from_pickle):
            model = modelClass(args, data)
            history = []
        else:
            model = torch.load(pathResult.joinpath("model.pickle"), args.device)
            history = torch.load(pathResult.joinpath("history.pickle"))
            model.args.num_steps += args.num_steps

        for n in range(args.num_steps):
            probability = model.step(args.step_subsample, deterministic_coef=args.deterministic_coefs[n])
            if((n + 1) % 10 == 0):
                print("i:{:<5d} loss:{:<f}".format(n + 1, probability))

            hist = {}
            hist["log_probs"] = probability
            if(do_hist_analysis):
                hist.update(_make_variables_summary_dict(model, summary_args.habitWorstLevels))
            history.append(hist)

        print("Saving the result")
        torch.save(model, pathResult.joinpath("model.pickle"))
        torch.save(history, pathResult.joinpath("history.pickle"))

    print("Saving the summary")
    model.summary(summary_args)
    pathResult.joinpath("figs").mkdir(exist_ok=True, parents=True)
    plt.plot([hist["log_probs"] for hist in history])
    plt.savefig(pathResult.joinpath("figs", "probability.png"))
    plt.close()

    """accuracy"""
    if(testset is not None):
        print("calcurating accuracy")
        model.set_testset(testset)
        accuracy = model.calc_all_mean_accuracy_from_testset(args.step_subsample, max_iter=100, min_iter=10)
        # accuracy = model.calc_mean_accuracy_from_testset({"rm": [0]}, args.step_subsample, max_iter=5)

        with open(str(pathResult.joinpath("accuracy.json")), "w", encoding="utf_8_sig") as output:
            text = json.dumps(accuracy, ensure_ascii=False, indent=4)
            output.write(text)

    """history analysis"""
    if(do_hist_analysis):
        print("history analysis")
        pathResult.joinpath("figs").mkdir(exist_ok=True, parents=True)
        _make_colormap_video_from_history(model, history, pathResult.joinpath("figs", "colormaps.mp4"))
        _make_step_hist_figs(model, history, pathResult.joinpath("figs"), window_size=50)


def _make_variables_summary_dict(model, habitWorstLevels):
    d = {}

    d["theta"] = model.theta()[:100, :]

    phis = []
    for rt in range(model.Rt):
        phis.append(model.phi(rt)[:, :100])
    d["phis"] = phis

    mus = []
    for rm in range(model.Rm):
        mus.append(model.mu(rm))
    d["mus"] = np.array(mus)

    sigmas = []
    for rm in range(model.Rm):
        sigmas.append(model.sigma(rm))
    d["sigmas"] = np.array(sigmas)

    d["rhos"] = model.getWorstAnswerProbs(habitWorstLevels)

    return d


def _make_colormap_video_from_history(final_model, history, path):
    rm_boundary = _get_rm_boundary(final_model)
    out = cv2.VideoWriter(str(path), cv2.VideoWriter_fourcc(*'mp4v'), 2.0, (500, 500))
    variables = history[0]
    hist_conv = []

    for hist in history[1:]:
        img = _make_colormap(final_model, hist, variables, hist_conv, rm_boundary)
        out.write(img)
        variables = hist

    out.release()


def _get_rm_boundary(model):
    mean = torch.mean(model.x_rm, 1).cpu().detach().numpy().copy()
    std = torch.std(model.x_rm, 1).cpu().detach().numpy().copy()
    lower = np.repeat((mean - std)[None, :], model.K, axis=0)
    upper = np.repeat((mean + std)[None, :], model.K, axis=0)
    return [lower, upper]


def _make_colormap(final_model, now_variables, past_variables, hist_conv, rm_boundary):
    canvas = np.zeros((500, 500, 3), dtype=np.uint8)
    conv1 = [None, None, None]
    conv2 = [None, None, None]

    if(final_model.Rt != 0):
        phi = now_variables["phis"][0][:, :10]
        conv1[0] = np.isclose(phi, past_variables["phis"][0][:, :10], rtol=5e-02, atol=0e-08)
        if(len(hist_conv) >= 4):
            conv2[0] = conv1[0]
            for i in range(4):
                conv2[0] &= hist_conv[-1 - i][0]
        img_phi = makeColorMap(phi, 40, 10, color_func=DEFAULT_COLOR_FUNC, axis=0)  # K*40 x 100
        # img_phi = makeColorMap(phi, 40, 10, color_func=DEFAULT_COLOR_FUNC2,
        #                        boundary=[0., 0.05], border_mask=conv2[0])  # K*40 x 100
        addImage(canvas, img_phi, 0, 0)

    if(final_model.Rm != 0):
        mu = now_variables["mus"].T
        conv1[1] = np.isclose(mu, past_variables["mus"].T, rtol=5e-02, atol=0e-08)
        if(len(hist_conv) >= 4):
            conv2[1] = conv1[1]
            for i in range(4):
                conv2[1] &= hist_conv[-1 - i][1]
        img_mu = makeColorMap(mu, 40, 10, color_func=DEFAULT_COLOR_FUNC, axis=0)  # K*40 x Rm*10
        # img_mu = makeColorMap(mu, 40, 10, color_func=DEFAULT_COLOR_FUNC2,
        #                       boundary=rm_boundary, border_mask=conv2[1])  # K*40 x Rm*10
        addImage(canvas, img_mu, 0, 120)

    if(final_model.Rh != 0):
        rho = now_variables["rhos"].T
        conv1[2] = np.isclose(rho, past_variables["rhos"].T, rtol=5e-02, atol=0e-08)
        if(len(hist_conv) >= 4):
            conv2[2] = conv1[2]
            for i in range(4):
                conv2[2] &= hist_conv[-1 - i][2]
        img_rho = makeColorMap(rho, 40, 10, color_func=DEFAULT_COLOR_FUNC, axis=0)  # K*40 x Rh*10
        # img_rho = makeColorMap(rho, 40, 10, color_func=DEFAULT_COLOR_FUNC2,
        #                        boundary=[0., 1.], border_mask=conv2[2])  # K*40 x Rh*10
        addImage(canvas, img_rho, 0, 140 + final_model.Rm * 10)

    hist_conv.append(conv1)
    canvas = np.transpose(canvas, [1, 0, 2])
    return canvas


def _make_step_hist_figs(final_model, history, pathFigs, window_size=50):
    path_step_hist_for_each_r = pathFigs.joinpath("step_hist_each_r")
    path_step_hist_for_each_k = pathFigs.joinpath("step_hist_each_k")
    path_step_hist_for_each_r.mkdir(exist_ok=True, parents=True)
    path_step_hist_for_each_k.mkdir(exist_ok=True, parents=True)

    phis = np.array([hist["phis"] for hist in history])  # [step, rt, k, v]
    mus = np.array([hist["mus"] for hist in history])  # [step, rm, k]
    rhos = np.array([hist["rhos"] for hist in history])  # [step, rh, k]

    # moving average
    window = np.ones(window_size) / window_size
    phis = simple_moving_average(phis, window)
    mus = simple_moving_average(mus, window)
    rhos = simple_moving_average(rhos, window)

    for rt in range(final_model.Rt):

        # phi, r固定, v固定, k間比較
        fig = plt.figure(figsize=[25, 10])
        for v in range(10):
            ax = fig.add_subplot(2, 5, v + 1)
            for k in range(final_model.K):
                ax.set_title(f"vocab{v+1}")
                ax.plot(phis[:, rt, k, v], label=f"topic{k+1}")
        fig.legend([f"topic{k+1}" for k in range(final_model.K)])
        fig.savefig(path_step_hist_for_each_r.joinpath(f"phis{rt+1}_hist.png"))
        plt.close(fig)

        # phi, rt固定, k固定, v間比較
        fig = plt.figure(figsize=[25, 10])
        phis_n = min_max_normalize(phis[:, rt, :, :], axis=(0, 1))
        for k in range(final_model.K):
            ax = fig.add_subplot((final_model.K - 1) // 5 + 1, 5, k + 1)
            ax.set_ylim((0, 1))
            for v in range(10):
                ax.set_title(f"topic{k+1}")
                ax.plot(phis_n[:, k, v], label=f"vocab{v+1}")
        fig.legend([f"vocab{v+1}" for v in range(10)])
        fig.savefig(path_step_hist_for_each_k.joinpath(f"phis{rt+1}_hist.png"))
        plt.close(fig)

    # mu, r固定, k間比較
    fig = plt.figure(figsize=[25, 10])
    for rm in range(final_model.Rm):
        ax = fig.add_subplot((final_model.Rm - 1) // 5 + 1, 5, rm + 1)
        for k in range(final_model.K):
            ax.set_title(f"rm={rm+1}")
            ax.plot(mus[:, rm, k])
    fig.legend([f"topic{k+1}" for k in range(final_model.K)])
    fig.savefig(path_step_hist_for_each_r.joinpath("mus_hist.png"))
    plt.close(fig)

    # mu, k固定, r間比較 (step, rに関して正規化)
    fig = plt.figure(figsize=[25, 10])
    mus_n = min_max_normalize(mus, axis=(0, 2))
    for k in range(final_model.K):
        ax = fig.add_subplot((final_model.K - 1) // 5 + 1, 5, k + 1)
        ax.set_ylim((0., 1.))
        for rm in range(final_model.Rm):
            ax.set_title(f"topic{k+1}")
            ax.plot(mus_n[:, rm, k])
    fig.legend([f"rm={rm+1}" for rm in range(final_model.Rm)])
    fig.savefig(path_step_hist_for_each_k.joinpath("mus_hist.png"))
    plt.close(fig)

    # rho, r固定, k間比較
    fig = plt.figure(figsize=[25, 10])
    for rh in range(final_model.Rh):
        ax = fig.add_subplot((final_model.Rh - 1) // 5 + 1, 5, rh + 1)
        for k in range(final_model.K):
            ax.set_title(f"rh={rh+1}")
            ax.plot(rhos[:, rh, k])
    fig.legend([f"topic{k+1}" for k in range(final_model.K)])
    fig.savefig(path_step_hist_for_each_r.joinpath("rhos_hist.png"))
    plt.close(fig)

    # rho, k固定, r間比較 (step, rに関して正規化)
    fig = plt.figure(figsize=[25, 10])
    rhos_n = min_max_normalize(rhos, axis=(0, 2))
    for k in range(final_model.K):
        ax = fig.add_subplot((final_model.K - 1) // 5 + 1, 5, k + 1)
        ax.set_ylim((0., 1.))
        for rh in range(final_model.Rh):
            ax.set_title(f"topic{k+1}")
            ax.plot(rhos_n[:, rh, k])
    fig.legend([f"rh={rh+1}" for rh in range(final_model.Rh)])
    fig.savefig(path_step_hist_for_each_k.joinpath("rhos_hist.png"))
    plt.close(fig)
