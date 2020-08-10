import random
import pickle
import numpy as np
import scipy
import torch
import pyro.distributions as dist
import openpyxl
from sklearn.manifold import MDS, TSNE
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

from LDA_pytorch.ModelBase import LDABase
from LDA_pytorch.MCLDA_infer import MCLDA_infer, _mask_validate
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix, addColorScaleRules, addBorderToMaxCell
from utils.wordcloud_util import create_wordcloud
from utils.general_util import cycleArray


class MCLDA(LDABase):

    def __init__(self, args, data):
        """
        Parameters
        ----------
        args : object
            K,
            n_rh,  # number of categories for record "rh"
            auto_beta, auto_alpha,
            coef_beta, coef_alpha,
            nu_h,
            device

        data : object
            [[doc[sentence[word, word, ...], ...], ...], [doc], ...],  # docs
            [Rm, D],  # measurements(continueous value)
            [Rh, D]]  # habits(categorical value)
        """
        super().__init__()

        docs = data[0]
        measurements = torch.tensor(data[1], device=args.device, dtype=torch.float64)  # [Rm, D]
        habits = torch.tensor(data[2], device=args.device, dtype=torch.int64)        # [Rh, D]

        # fields
        self.args = args
        # self.data = data
        self.device = args.device

        self.K = K = args.K
        self.D = D = len(docs[0]) if len(docs) != 0 else (len(measurements[0]) if len(measurements) != 0 else (len(habits[0]) if len(habits) != 0 else 0))
        self.Rt = len(docs)
        self.Rm = len(measurements)
        self.Rh = len(habits)
        self.V_rt = [0 for _ in range(self.Rt)]            # [Rt]
        self.totalN_rt = [0 for _ in range(self.Rt)]       # [Rt]
        self.n_rh = args.n_rh                              # [Rh]
        self.word_count_rt = [{} for _ in range(self.Rt)]  # [Rt][V_rt]
        self.word_dict_rt = [[] for _ in range(self.Rt)]   # [Rt][V_rt]
        self.wordids_rt = [[] for _ in range(self.Rt)]     # [Rt][totalN_rt]
        self.docids_rt = [[] for _ in range(self.Rt)]      # [Rt][totalN_rt]
        self.z_rt = [None for _ in range(self.Rt)]         # [Rt][totalN_rt]
        self.x_rm = measurements                           # [Rm, D]
        self.x_rh = habits                                 # [Rh, D]
        self.z_rm = None                                   # [Rm, D]
        self.z_rh = None                                   # [Rh, D]

        self.alpha = torch.full([K], args.coef_alpha, device=self.device, dtype=torch.float64)                           # [K]
        self.beta_rt = [None for _ in range(self.Rt)]                                                                    # [Rt][V_rt]
        self.mu_h_rm = torch.mean(measurements, 1)                                                                       # [Rm]
        self.nu_h_rm = torch.tensor(args.nu_h, device=self.device, dtype=torch.float64)
        self.sigma2_h_rm = torch.var(measurements, 1)                                                                    # [Rm]
        self.rho_h_rh = [torch.ones([self.n_rh[rh]], device=self.device, dtype=torch.float64) for rh in range(self.Rh)]  # [Rh][n_rh]

        self._tpd = torch.zeros((D, K), device=self.device, dtype=torch.int64)              # [D, K]
        self._nd = torch.zeros(D, device=self.device, dtype=torch.int64)                    # [D]
        self._wpt_rt = [None for _ in range(self.Rt)]                                       # [Rt][K, V_rt]
        self._wt_rt = [None for _ in range(self.Rt)]                                        # [Rt][K]
        self._mean_rm = torch.zeros((self.Rm, K), device=self.device, dtype=torch.float64)  # [Rm, K]
        self._std_rm = torch.zeros((self.Rm, K), device=self.device, dtype=torch.float64)   # [Rm, K]
        self._xt_rm = torch.zeros((self.Rm, K), device=self.device, dtype=torch.int64)      # [Rm, K]
        self._xpt_rh = [None for _ in range(self.Rh)]                                       # [Rh][K, n_rh]
        self._xt_rh = [None for _ in range(self.Rh)]                                        # [Rh][K]

        # init wordcount...
        for rt, documents in enumerate(docs):
            for d, doc in enumerate(documents):
                for sentence in doc:
                    for word in sentence:
                        if(word not in self.word_count_rt[rt]):
                            self.word_count_rt[rt][word] = 0
                        self.word_count_rt[rt][word] += 1
            self.word_dict_rt[rt] = [a[0] for a in sorted(self.word_count_rt[rt].items(), key=lambda x: -x[1])]
            self.V_rt[rt] = len(self.word_dict_rt[rt])
            self.beta_rt[rt] = torch.full([self.V_rt[rt]], args.coef_beta, device=self.device, dtype=torch.float64)

        # init wordids_rt...
        for rt, documents in enumerate(docs):
            for d, doc in enumerate(documents):
                for sentence in doc:
                    for word in sentence:
                        self.wordids_rt[rt].append(self.word_dict_rt[rt].index(word))
                        self.docids_rt[rt].append(d)
            self.totalN_rt[rt] = len(self.wordids_rt[rt])
            self.wordids_rt[rt] = torch.tensor(self.wordids_rt[rt], device=self.device, dtype=torch.int64)
            self.docids_rt[rt] = torch.tensor(self.docids_rt[rt], device=self.device, dtype=torch.int64)

        # init latent variables...
        for rt in range(self.Rt):
            self.z_rt[rt] = torch.randint(0, K, (self.totalN_rt[rt],), device=self.device, dtype=torch.int64)
            ones = torch.ones(self.totalN_rt[rt], device=self.device, dtype=torch.int64)
            self._tpd.index_put_([self.docids_rt[rt], self.z_rt[rt]], ones, accumulate=True)
            self._wpt_rt[rt] = torch.zeros((K, self.V_rt[rt]), device=self.device, dtype=torch.int64)
            self._wpt_rt[rt].index_put_([self.z_rt[rt], self.wordids_rt[rt]], ones, accumulate=True)
            self._wt_rt[rt] = torch.sum(self._wpt_rt[rt], 1)
        self.z_rm = torch.randint(0, K, [self.Rm, D], device=self.device, dtype=torch.int64)
        ones = torch.ones(self.D, device=self.device, dtype=torch.int64)
        idx = torch.arange(0, self.D, device=self.device, dtype=torch.int64)
        for rm in range(self.Rm):
            self._tpd.index_put_([idx, self.z_rm[rm]], ones, accumulate=True)
            for k in range(K):
                self._mean_rm[rm, k] = torch.mean(self.x_rm[rm][self.z_rm[rm] == k])
                self._std_rm[rm, k] = torch.std(self.x_rm[rm][self.z_rm[rm] == k])
                self._xt_rm[rm, k] = torch.sum(self.z_rm[rm] == k)
        self.z_rh = torch.randint(0, K, [self.Rh, D], device=self.device, dtype=torch.int64)
        for rh in range(self.Rh):
            self._tpd.index_put_([idx, self.z_rh[rh]], ones, accumulate=True)
            self._xpt_rh[rh] = torch.zeros((K, self.n_rh[rh]), device=self.device, dtype=torch.int64)
            self._xpt_rh[rh].index_put_([self.z_rh[rh], self.x_rh[rh]], ones, accumulate=True)
            self._xt_rh[rh] = torch.sum(self._xpt_rh[rh], 1)
        self._nd = torch.sum(self._tpd, 1)


    def step(self, num_subsample_partitions, parameter_update=False, deterministic_coef=None):
        """
        num_subsample_partitions: 何回に分けてzのサンプリングを行うか
        parameter_update: パラメータをアップデートするか(未実装)
        deterministic_coef: 学習率に対応するようなパラメータ，0で通常のMCMC，1でzを点推定
                            Noneなら0に相当する動きをして，これに関わる計算をしないので少し早い

        1. update z, wpt, ...
        2. update parameters if neseccary
        """

        # generate sabsampling batches
        rand_perm_rt = []
        rand_perm_rm = []
        rand_perm_rh = []
        for rt in range(self.Rt):
            rand_perm_rt.append(torch.randperm(self.totalN_rt[rt], device=self.device))
        for rm in range(self.Rm):
            rand_perm_rm.append(torch.randperm(self.D, device=self.device))
        for rh in range(self.Rh):
            rand_perm_rh.append(torch.randperm(self.D, device=self.device))

        # sampling
        for n in range(num_subsample_partitions):
            # texts
            for rt in range(self.Rt):
                s = (self.totalN_rt[rt] // num_subsample_partitions + 1) * n
                e = (self.totalN_rt[rt] // num_subsample_partitions + 1) * (n + 1) - 1
                e = e if e < self.totalN_rt[rt] else self.totalN_rt[rt] - 1
                self._sampling_rt(rt, rand_perm_rt[rt][s:e], deterministic_coef)
            # measurements
            for rm in range(self.Rm):
                s = (self.D // num_subsample_partitions + 1) * n
                e = (self.D // num_subsample_partitions + 1) * (n + 1) - 1
                e = e if e < self.D else self.D - 1
                self._sampling_rm(rm, rand_perm_rm[rm][s:e], deterministic_coef)
            # habits
            for rh in range(self.Rh):
                s = (self.D // num_subsample_partitions + 1) * n
                e = (self.D // num_subsample_partitions + 1) * (n + 1) - 1
                e = e if e < self.D else self.D - 1
                self._sampling_rh(rh, rand_perm_rh[rh][s:e], deterministic_coef)

        # update parameters
        if(parameter_update):
            self._update_parameters()

        """ calculation checking """
        # for i in range(100):
        #     k = random.randrange(self.K)
        #     d = random.randrange(self.D)
        #     s = torch.zeros(1, device=self.device, dtype=torch.int64)
        #     for rt in range(self.Rt):
        #         s += torch.sum(torch.logical_and(self.docids_rt[rt] == d, self.z_rt[rt] == k))
        #     s += torch.sum(self.z_rm[:, d] == k)
        #     s += torch.sum(self.z_rh[:, d] == k)
        #     assert self._tpd[d, k] == s

        return self.log_probability()


    def _sampling_rt(self, rt, idx, deterministic_coef=None):
        """
        1. sampleing z
        2. update wpt, tpd, wt
        """

        before = self.z_rt[rt][idx].clone()

        # zをサンプリング
        probs = self._get_z_rt_sampling_probs(rt, idx, deterministic_coef)
        self.z_rt[rt][idx] = dist.Categorical(probs).sample()

        # wptとか更新
        ones = torch.ones(len(idx), device=self.device, dtype=torch.int64)
        self._wpt_rt[rt].index_put_([before, self.wordids_rt[rt][idx]], -ones, accumulate=True)
        self._wpt_rt[rt].index_put_([self.z_rt[rt][idx], self.wordids_rt[rt][idx]], ones, accumulate=True)
        self._tpd.index_put_([self.docids_rt[rt][idx], before], -ones, accumulate=True)
        self._tpd.index_put_([self.docids_rt[rt][idx], self.z_rt[rt][idx]], ones, accumulate=True)
        self._wt_rt[rt] = torch.sum(self._wpt_rt[rt], 1)

        """ calculation checking """
        # for i in range(100):
        #     k = random.randrange(self.K)
        #     v = random.randrange(self.V_rt[rt])
        #     assert self._wpt_rt[rt][k, v] == torch.sum(torch.logical_and(self.z_rt[rt] == k, self.wordids_rt[rt] == v))


    def _get_z_rt_sampling_probs(self, rt, idx, deterministic_coef=None, without_theta=False):
        subsample_size = len(idx)
        probs = torch.ones((subsample_size, self.K), device=self.device, dtype=torch.float64)  # [subsample_size, K]
        a = torch.zeros((subsample_size, self.K), device=self.device, dtype=torch.float64)     # [subsample_size, K]
        a[torch.arange(0, subsample_size, device=self.device, dtype=torch.int64), self.z_rt[rt][idx]] = 1

        if(not without_theta):
            probs *= self._tpd[self.docids_rt[rt][idx], :] + self.alpha[None, :] - a
            probs /= self._nd[self.docids_rt[rt][idx]][:, None] + self.alpha.sum() - a
        probs *= self._wpt_rt[rt][:, self.wordids_rt[rt][idx]].T + self.beta_rt[rt][self.wordids_rt[rt][idx]][:, None] - a
        probs /= self._wt_rt[rt][None, :] + self.beta_rt[rt].sum() - a

        """ calculation checking """
        # for i in range(100):
        #     n = random.randrange(subsample_size)  # any
        #     k = random.randrange(self.K)  # any
        #     z = self.z_rt[rt][idx[n]]
        #     v = self.wordids_rt[rt][idx[n]]
        #     d = self.docids_rt[rt][idx[n]]
        #     a0 = (1 if z == k else 0)
        #     assert a[n, k] == a0
        #     assert(probs[n, k] == (self._wpt_rt[rt][k, v] - a0 + self.beta_rt[rt][v]) / (self._wt_rt[rt][k] + self.beta_rt[rt].sum() - a0)
        #            * (self._tpd[d, k] - a0 + self.alpha[k]) / (self._nd[d] + self.alpha.sum() - a0))

        probs /= torch.sum(probs, dim=1, keepdim=True)

        if(deterministic_coef is not None):
            probs = self._to_deterministic_probs(probs, deterministic_coef)

        return probs


    def _sampling_rm(self, rm, idx, deterministic_coef=None):
        """
        1. sampleing z
        2. update tpd, mean_rm, std_rm
        """

        before = self.z_rm[rm][idx].clone()

        # zをサンプリング
        probs = self._get_z_rm_sampling_probs(rm, idx, deterministic_coef)
        self.z_rm[rm][idx] = dist.Categorical(probs).sample()

        # tpdとか更新
        ones = torch.ones(len(idx), device=self.device, dtype=torch.int64)
        self._tpd.index_put_([idx, self.z_rm[rm][idx]], ones, accumulate=True)
        self._tpd.index_put_([idx, before], -ones, accumulate=True)
        for k in range(self.K):
            self._mean_rm[rm, k] = torch.mean(self.x_rm[rm][self.z_rm[rm] == k])
            self._std_rm[rm, k] = torch.std(self.x_rm[rm][self.z_rm[rm] == k])
            self._xt_rm[rm, k] = torch.sum(self.z_rm[rm] == k)

        """
        例外的な状況の処理について
        self._xt_rm[rm,k] == 0 のとき  (self._mean_rm[rm,k] == nan, self._std_rm[rm,k] == nan):
            self._mean_rm[rm,k] = self.mu_h_rm[rm,k], self._std_rm[rm,k] = 0 とする
            (_get_z_rm_sampling_probs()で対応)
        self._xt_rm[rm,k] == 1 の場合  (self._std_rm[rm,k] == nan):
            事前分布設定で解決?
        z=kのxがすべて一定の場合        (self._std_rm[rm,k] == 0):
            事前分布設定で解決
        """
        self._std_rm[rm][torch.isnan(self._std_rm[rm])] = 0


    def _get_z_rm_sampling_probs(self, rm, idx, deterministic_coef=None, without_theta=False):
        """
        本来_mean_rmと_std_rmは自分の影響を除いて計算するべきだけど，近似的にかわらないとした

        # 事後分布の平均は，事後分布の分散が既知(=self._std_rm[rm, k])として求め，
        # 事後分布の分散は，事後分布の平均が既知(=self._mean_rm[rm, k])として求めた．
        事後分布の分散は，全体の分散と線形結合した
        事前分布：
        mean ~ Normal(self.mu_h_rm[rm, k], self._std_rm[rm, k])
        var ~ Scaled-inv-chi-squared(self.mu_h_rm[rm, k], self.sigma2_h_rm[rm, k])
        本来は事後分布の平均も分散も未知として，事後分布の平均と分散を同時に推定するべき
        """

        subsample_size = len(idx)
        probs = torch.ones((subsample_size, self.K), device=self.device, dtype=torch.float64)                    # [subsample_size, K]
        x = torch.zeros([subsample_size, self.K], device=self.device, dtype=torch.float64) + self.x_rm[rm, idx][:, None]  # [subsample_size, K]

        if(not without_theta):
            a = torch.zeros((subsample_size, self.K), device=self.device, dtype=torch.float64)                       # [subsample_size, K]
            a[torch.arange(0, subsample_size, device=self.device, dtype=torch.int64), self.z_rm[rm][idx]] = 1
            probs *= self._tpd[idx, :] + self.alpha[None, :] - a
            probs /= self._nd[idx][:, None] + self.alpha.sum() - a

        # probs *= dist.Normal(self._mean_rm[rm], self._std_rm[rm]).log_prob(x).exp()

        std2 = self._std_rm[rm] ** 2
        mean = self._xt_rm[rm] * self.sigma2_h_rm[rm] * self._mean_rm[rm] + std2 * self.mu_h_rm[rm]  # [K]
        mean /= self._xt_rm[rm] * self.sigma2_h_rm[rm] + std2
        # mean = self._mean_rm[rm]
        mean[torch.isnan(mean)] = self.mu_h_rm[rm]
        # var = (self.nu_h_rm * self.sigma2_h_rm[rm] + self._xt_rm[rm] * std2) / (self.nu_h_rm + self._xt_rm[rm] - 2)  # [K]
        var = (self.nu_h_rm * self.sigma2_h_rm[rm] + self._xt_rm[rm] * std2) / (self.nu_h_rm + self._xt_rm[rm])  # [K]
        probs *= dist.Normal(mean, torch.pow(var, 0.5)).log_prob(x).exp()

        """ calculation checking """
        # for i in range(100):
        #     n = random.randrange(subsample_size)  # any
        #     k = random.randrange(self.K)  # any
        #     d = idx[n]
        #     z = self.z_rm[rm][idx[n]]
        #     a0 = (1 if z == k else 0)
        #     x = self.x_rm[rm, d]
        #     assert torch.allclose(probs[n, k], (self._tpd[d, k] - a0 + self.alpha[k]) / (self._nd[d] + self.alpha.sum() - a0)
        #                           * dist.Normal(self._mean_rm[rm, k], self._std_rm[rm, k]).log_prob(x).exp())

        probs /= torch.sum(probs, dim=1, keepdim=True)

        if(deterministic_coef is not None):
            probs = self._to_deterministic_probs(probs, deterministic_coef)

        return probs


    def _sampling_rh(self, rh, idx, deterministic_coef=None):
        """
        1. sampleing z
        2. update tpd, xpt, xt
        """

        before = self.z_rh[rh][idx].clone()

        # zをサンプリング
        probs = self._get_z_rh_sampling_probs(rh, idx, deterministic_coef)
        self.z_rh[rh][idx] = dist.Categorical(probs).sample()

        # tpdとか更新
        ones = torch.ones(len(idx), device=self.device, dtype=torch.int64)
        self._tpd.index_put_([idx, before], -ones, accumulate=True)
        self._tpd.index_put_([idx, self.z_rh[rh][idx]], ones, accumulate=True)
        self._xpt_rh[rh].index_put_([before, self.x_rh[rh][idx]], -ones, accumulate=True)
        self._xpt_rh[rh].index_put_([self.z_rh[rh][idx], self.x_rh[rh][idx]], ones, accumulate=True)
        self._xt_rh[rh] = torch.sum(self._xpt_rh[rh], 1)

        """ calculation checking """
        # for i in range(100):
        #     k = random.randrange(self.K)
        #     x = random.randrange(self.n_rh[rh])
        #     assert self._xpt_rh[rh][k, x] == torch.sum(torch.logical_and(self.z_rh[rh] == k, self.x_rh[rh] == x))


    def _get_z_rh_sampling_probs(self, rh, idx, deterministic_coef=None, without_theta=False):
        subsample_size = len(idx)
        probs = torch.ones((subsample_size, self.K), device=self.device, dtype=torch.float64)  # [subsample_size, K]
        a = torch.zeros((subsample_size, self.K), device=self.device, dtype=torch.float64)     # [subsample_size, K]
        a[torch.arange(0, subsample_size, device=self.device, dtype=torch.int64), self.z_rh[rh][idx]] = 1

        if(not without_theta):
            probs *= self._tpd[idx, :] + self.alpha[None, :] - a
            probs /= self._nd[idx][:, None] + self.alpha.sum() - a
        probs *= self._xpt_rh[rh][:, self.x_rh[rh, idx]].T + self.rho_h_rh[rh][self.x_rh[rh][idx]][:, None] - a
        probs /= self._xt_rh[rh][None, :] + self.rho_h_rh[rh].sum() - a

        """ calculation checking """
        # for i in range(100):
        #     n = random.randrange(subsample_size)  # any
        #     k = random.randrange(self.K)  # any
        #     d = idx[n]
        #     z = self.z_rh[rh, d]
        #     x = self.x_rh[rh, d]
        #     a0 = (1 if z == k else 0)
        #     assert a[n, k] == a0
        #     assert(probs[n, k] == (self._xpt_rh[rh][k, x] + self.rho_h_rh[rh][x] - a0) / (self._xt_rh[rh][k] + self.rho_h_rh[rh].sum() - a0)
        #            * (self._tpd[d, k] - a0 + self.alpha[k]) / (self._nd[d] + self.alpha.sum() - a0))

        probs /= torch.sum(probs, dim=1, keepdim=True)

        if(deterministic_coef is not None):
            probs = self._to_deterministic_probs(probs, deterministic_coef)

        return probs


    def _to_deterministic_probs(self, probs, deterministic_coef):
        deterministic_probs = torch.full_like(probs, 0.0, device=self.device, dtype=torch.float64)
        idx0 = torch.arange(probs.shape[0], device=self.device, dtype=torch.int64)
        idx1 = torch.argmax(probs, dim=1)
        deterministic_probs[idx0, idx1] = 1.
        return (1 - deterministic_coef) * probs + deterministic_coef * deterministic_probs


    def _update_parameters(self):
        # TODO
        pass


    def theta(self, to_cpu=True):
        """
        毎回計算するから何度も呼び出さないこと
        """
        theta = (self._tpd + self.alpha[None, :]) / (self._nd[:, None] + self.alpha.sum())  # [D, K]
        return theta.cpu().detach().numpy().copy() if to_cpu else theta


    def phi(self, rt, to_cpu=True):
        """
        毎回計算するから何度も呼び出さないこと
        """
        phi = (self._wpt_rt[rt] + self.beta_rt[rt][None, :]) / (self._wt_rt[rt][:, None] + self.beta_rt[rt].sum())  # [K, V]
        return phi.cpu().detach().numpy().copy() if to_cpu else phi


    def mu(self, rm, to_cpu=True):
        # mu = self._mean_rm[rm]
        mu = self._xt_rm[rm] * self.sigma2_h_rm[rm] * self._mean_rm[rm] + (self._std_rm[rm] ** 2) * self.mu_h_rm[rm]
        mu /= self._xt_rm[rm] * self.sigma2_h_rm[rm] + self._std_rm[rm] ** 2
        mu[torch.isnan(mu)] = self.mu_h_rm[rm]
        return mu.cpu().detach().numpy().copy() if to_cpu else mu


    def sigma(self, rm, to_cpu=True):
        # sigma = self._std_rm[rm]
        # var = (self.nu_h_rm * self.sigma2_h_rm[rm] + self._xt_rm[rm] * self._std_rm[rm] ** 2) / (self.nu_h_rm + self._xt_rm[rm] - 2)
        var = (self.nu_h_rm * self.sigma2_h_rm[rm] + self._xt_rm[rm] * self._std_rm[rm] ** 2) / (self.nu_h_rm + self._xt_rm[rm])
        sigma = torch.pow(var, 0.5)
        return sigma.cpu().detach().numpy().copy() if to_cpu else sigma


    def rho(self, rh, to_cpu=True):
        """
        毎回計算するから何度も呼び出さないこと
        """
        rho = (self._xpt_rh[rh] + self.rho_h_rh[rh][None, :]) / (self._xt_rh[rh][:, None] + self.rho_h_rh[rh].sum())  # [K, n_rh]
        return rho.cpu().detach().numpy().copy() if to_cpu else rho


    def getWorstAnswerProbs(self, habitWorstLevels, rhos=None):
        """
        Rhoから最悪の回答をする確率の表に変換する(CPU) [Rh, K]
        """
        rhos = rhos if rhos is not None else [self.rho(rh) for rh in range(self.Rh)]

        table = np.zeros((self.Rh, self.K))
        for r in range(self.Rh):
            table[r, :] = rhos[r][:, habitWorstLevels[r]]

        return table


    def log_probability(self, testset=None):
        if(testset is None):
            p = 0.
            theta = self.theta(to_cpu=False)
            for rt in range(self.Rt):
                dv = torch.mm(theta, self.phi(rt, to_cpu=False))
                p += dv[self.docids_rt[rt], self.wordids_rt[rt]].log().sum().item()
            idx = torch.arange(0, self.D, device=self.device, dtype=torch.int64)
            for rm in range(self.Rm):
                mu = self.mu(rm, to_cpu=False)
                sigma = self.sigma(rm, to_cpu=False)
                x = torch.zeros([self.D, self.K], device=self.device, dtype=torch.float64) + self.x_rm[rm][:, None]
                dk = dist.Normal(mu, sigma).log_prob(x)
                p += dk[idx, self.z_rm[rm]].sum().item()
                # p += (theta.log() + dk).sum().item()
            for rh in range(self.Rh):
                dx = torch.mm(theta, self.rho(rh, to_cpu=False))
                p += dx[idx, self.x_rh[rh]].log().sum().item()
            return p
        else:
            # TODO
            return None


    def perplexity(self, testset):
        pass


    def set_testset(self, testset):
        self.model_infer = MCLDA_infer(self, testset, {})


    def calc_mean_accuracy_from_testset(self, masked_records, num_subsample_partitions, max_iter=100, min_iter=10, return_iter=False):
        """
        未知のデータに対するこのモデルの平均正解率を計算する
        事前にset_testsetでテストセットを登録する必要がある
        Parametera
        ----------
        masked_records: {"rt": [masked_record_id, ...], "rm": [...], "rh": [...]}
        num_subsample_partitions: 説明はstepを参照
        Return
        ------
        mean_accuracy: {"rt": [mean_accuracy, ...], "rm": [...], "rh": [...]}
                       the same shape of test_records argument
        """

        self.model_infer.change_mask(masked_records)

        # learn step
        losses = []
        for n in range(max_iter):
            probability = self.model_infer.step(num_subsample_partitions)
            losses.append(probability)
            if((n + 1) >= min_iter and np.isclose(sum(losses[-5:]) / 5, losses[-1], rtol=1e-05)):
                break

        if(return_iter):
            return self.model_infer.calc_mean_accuracy(), n + 1
        else:
            return self.model_infer.calc_mean_accuracy()


    def calc_all_mean_accuracy_from_testset(self, num_subsample_partitions, max_iter=100, min_iter=10):
        mean_accuracy = _mask_validate({})
        mean_accuracy["rt_iter"] = []
        mean_accuracy["rm_iter"] = []
        mean_accuracy["rh_iter"] = []

        for rt in range(self.Rt):
            a, n = self.calc_mean_accuracy_from_testset({"rt": [rt]}, num_subsample_partitions, max_iter, min_iter, return_iter=True)
            mean_accuracy["rt"].append(a["rt"][0])
            mean_accuracy["rt_iter"].append(n)

        for rm in range(self.Rm):
            a, n = self.calc_mean_accuracy_from_testset({"rm": [rm]}, num_subsample_partitions, max_iter, min_iter, return_iter=True)
            mean_accuracy["rm"].append(a["rm"][0])
            mean_accuracy["rm_iter"].append(n)

        for rh in range(self.Rh):
            a, n = self.calc_mean_accuracy_from_testset({"rh": [rh]}, num_subsample_partitions, max_iter, min_iter, return_iter=True)
            mean_accuracy["rh"].append(a["rh"][0])
            mean_accuracy["rh_iter"].append(n)
        return mean_accuracy


    def coherence(self, rt, k, w2v_model):
        pass


    def _summary_print(self, summary_args):
        pass


    def _sammary_figs(self, summary_args):
        p = summary_args.summary_path.joinpath("figs", "mapping")
        p.mkdir(exist_ok=True, parents=True)

        theta = self.theta()
        x_rm = self.x_rm.cpu().detach().numpy().copy()

        clist = cycleArray(np.array(['tab:blue', 'tab:orange', 'tab:green', 'tab:red', 'tab:purple',
                                     'tab:brown', 'tab:pink', 'tab:gray', 'tab:olive', 'tab:cyan']))
        deta_colors = clist[np.argmax(theta, axis=1)]
        handles = [mpatches.Patch(color=clist[k], label=f"topic{k+1}") for k in range(self.K)]

        # mds_metric = MDS(n_components=2, metric=True, dissimilarity="precomputed", random_state=1)
        # mds_nonmetric = MDS(n_components=2, metric=False, dissimilarity="precomputed", random_state=1)
        pca = PCA(n_components=2)
        tsne = TSNE(n_components=2, random_state=1)

        # self._make_map(pca, theta, p.joinpath("docs_mapping_by_pca.png"), deta_colors=deta_colors)
        self._make_map(tsne, theta, p.joinpath("docs_mapping_by_tsne.png"), deta_colors=deta_colors, handles=handles)
        self._make_map(tsne, x_rm.T, p.joinpath("num_data_mapping_by_tsne.png"))
        self._make_map(pca, x_rm.T, p.joinpath("num_data_mapping_by_pca.png"))


    def _make_map(self, model, x, path, data_names=None, deta_colors=None, title=None, handles=None):
        pos = model.fit_transform(x)
        plt.scatter(pos[:, 0], pos[:, 1], c=deta_colors)
        if(data_names is not None):
            for n in range(pos.shape[0]):
                plt.annotate(data_names[n], xy=(pos[n, 0], pos[n, 1]))
        plt.title(title)
        if(handles is not None):
            plt.legend(handles=handles)
        plt.savefig(path)
        plt.clf()


    def _summary_to_excel(self, summary_args, wb):

        theta = self.theta()
        alpha = self.alpha.cpu().detach().numpy().copy()
        phi = [self.phi(rt) for rt in range(self.Rt)]
        beta = [self.beta_rt[rt].cpu().detach().numpy().copy() for rt in range(self.Rt)]
        mu = [self.mu(rm) for rm in range(self.Rm)]
        sigma = [self.sigma(rm) for rm in range(self.Rm)]
        rho = [self.rho(rh) for rh in range(self.Rh)]
        rho_h = [self.rho_h_rh[rh].cpu().detach().numpy().copy() for rh in range(self.Rh)]

        args_dict = {k: self.args.__dict__[k] for k in self.args.__dict__ if not k.startswith("__")}
        args_dict_str = args_dict.copy()
        for k in args_dict_str:
            if(not isinstance(args_dict_str[k], (int, float, complex, bool))):
                args_dict_str[k] = str(args_dict_str[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict_str.values()), axis="row", names=list(args_dict_str.keys()))

        ws = wb.create_sheet("num latent topics")
        writeMatrix(ws, [[self._wt_rt[r][k].item() for k in range(self.K)] for r in range(self.Rt)],
                    1, 1,
                    row_names=summary_args.full_tensors["tensor_keys"],
                    column_names=[f"topic{k+1}" for k in range(self.K)],
                    rule="databar")
        writeMatrix(ws, [[self._xt_rm[r][k].item() for k in range(self.K)] for r in range(self.Rm)],
                    self.Rt + 3, 1,
                    row_names=summary_args.full_tensors["measurement_keys"],
                    rule="databar")
        writeMatrix(ws, [[self._xt_rh[r][k].item() for k in range(self.K)] for r in range(self.Rh)],
                    self.Rt + self.Rm + 4, 1,
                    row_names=summary_args.full_tensors["habit_keys"],
                    rule="databar")
        writeVector(ws, [torch.sum(self._tpd[:, k]).item() for k in range(self.K)],
                    self.Rt + self.Rm + self.Rh + 5, 2,
                    rule="databar")

        ws = wb.create_sheet("core_values")
        rtn_rt, rtn_rm, rtn_rh = self._get_topic_core_values()
        writeMatrix(ws, rtn_rt, 1, 1,
                    row_names=summary_args.full_tensors["tensor_keys"],
                    column_names=[f"topic{k+1}" for k in range(self.K)],
                    rule="colorscale", ruleAxis="row")
        writeMatrix(ws, rtn_rm, self.Rt + 3, 1,
                    row_names=summary_args.full_tensors["measurement_keys"],
                    rule="colorscale", ruleAxis="row")
        writeMatrix(ws, rtn_rh, self.Rt + self.Rm + 4, 1,
                    row_names=summary_args.full_tensors["habit_keys"],
                    rule="colorscale", ruleAxis="row")

        ws = wb.create_sheet("mu_sigma")
        writeMatrix(ws, mu, 1, 1,
                    row_names=summary_args.full_tensors["measurement_keys"],
                    column_names=[f"topic{k+1}" for k in range(self.K)])
        writeVector(ws, self.mu_h_rm.cpu().detach().numpy(), 2, self.K + 3, axis="row")
        addColorScaleRules(ws, 2, self.Rm + 1, 2, self.K + 3, axis="column")
        addBorderToMaxCell(ws, 2, self.Rm + 1, 2, self.K + 1, axis="column")
        writeMatrix(ws, sigma, 1, self.K + 5,
                    row_names=summary_args.full_tensors["measurement_keys"],
                    column_names=[f"topic{k+1}" for k in range(self.K)],
                    rule="colorscale", ruleAxis="column")
        writeVector(ws, self.sigma2_h_rm.pow(0.5).cpu().detach().numpy(), 2, 2 * self.K + 7, axis="row")
        addColorScaleRules(ws, 2, self.Rm + 1, self.K + 6, 2 * self.K + 7, axis="column")

        ws = wb.create_sheet("rho")
        row = 1
        for r in range(self.Rh):
            ws.cell(row, 1, summary_args.full_tensors["habit_keys"][r])
            writeMatrix(ws, rho[r].T, row, 1,
                        row_names=summary_args.full_tensors["habit_levels"][r],
                        column_names=[f"topic{k+1}" for k in range(self.K)],
                        rule="databar", ruleBoundary=[0., 1.])
            _, counts = torch.unique(self.x_rh[r], return_counts=True)
            counts = counts.cpu().detach().numpy()
            counts = counts / np.sum(counts)
            writeMatrix(ws, counts[:, None], row, self.K + 3, column_names=["data distribution"],
                        rule="databar", ruleBoundary=[0., 1.])
            row += self.n_rh[r] + 2
        row += 1
        for r in range(self.Rh):
            wl = summary_args.habitWorstLevels[r]
            ws.cell(row + r, 1, f'{summary_args.full_tensors["habit_keys"][r]} -> {summary_args.full_tensors["habit_levels"][r][wl]}')
            writeVector(ws, rho[r][:, wl], row + r, 2, axis="column")
        addColorScaleRules(ws, row, row + self.Rh - 1, 2, self.K + 1, axis="column")
        addBorderToMaxCell(ws, row, row + self.Rh - 1, 2, self.K + 1, axis="column")

        ws = wb.create_sheet("alpha_hyper")
        writeVector(ws, alpha, axis="row", names=[f"topic{k+1}" for k in range(self.K)],
                    rule="databar")

        ws = wb.create_sheet("beta_hyper")
        for r in range(self.Rt):
            writeVector(ws, beta[r], column=r * 3 + 1,
                        axis="row", names=self.word_dict_rt[r],
                        rule="databar")

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(self.D)],
                    column_names=[f"topic{k+1}" for k in range(self.K)],
                    rule="databar", ruleBoundary=[0., 1.])

        for r in range(self.Rt):
            ws = wb.create_sheet(f"phi_r{r}")
            writeMatrix(ws, phi[r].T, 1, 1,
                        row_names=self.word_dict_rt[r],
                        column_names=[f"topic{k+1}" for k in range(self.K)],
                        rule="databar")

            ws = wb.create_sheet(f"phi_r{r}_sorted")
            writeSortedMatrix(ws, phi[r].T, axis=0, row=1, column=1,
                              row_names=self.word_dict_rt[r], column_names=[f"topic{k+1}" for k in range(self.K)],
                              maxwrite=100, order="higher")
            writeSortedMatrix(ws, phi[r].T, axis=0, row=1, column=self.K + 3,
                              row_names=None, column_names=[f"topic{k+1}" for k in range(self.K)],
                              maxwrite=100, order="higher",
                              rule="databar")

        # topics file
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        for k in range(self.K):
            ws = wb.create_sheet(f"topic_{k+1}")
            for r in range(self.Rt):
                ws.cell(1, r + 1, summary_args.full_tensors["tensor_keys"][r])
                idx = np.argsort(phi[r][k])[::-1]
                dat = np.array(self.word_dict_rt[r])[idx].tolist()
                writeVector(ws, dat, 2, r + 1, axis="row", names=None)

        wb.remove_sheet(tmp_ws)
        wb.save(summary_args.summary_path.joinpath("topics.xlsx"))


    def _get_topic_core_values(self):
        rtn_rt = np.zeros((self.Rt, self.K))
        rtn_rm = np.zeros((self.Rm, self.K))
        rtn_rh = np.zeros((self.Rh, self.K))

        for k in range(self.K):
            for rt in range(self.Rt):
                idx = torch.nonzero(self.z_rt[rt] == k, as_tuple=False)[:, 0]
                probs = self._get_z_rt_sampling_probs(rt, idx, without_theta=True)
                rtn_rt[rt, k] = probs[:, k].mean().item()

            for rm in range(self.Rm):
                idx = torch.nonzero(self.z_rm[rm] == k, as_tuple=False)[:, 0]
                probs = self._get_z_rm_sampling_probs(rm, idx, without_theta=True)
                rtn_rm[rm, k] = probs[:, k].mean().item()

            for rh in range(self.Rh):
                idx = torch.nonzero(self.z_rh[rh] == k, as_tuple=False)[:, 0]
                probs = self._get_z_rh_sampling_probs(rh, idx, without_theta=True)
                rtn_rh[rh, k] = probs[:, k].mean().item()

        return rtn_rt, rtn_rm, rtn_rh
