import random
import pickle
import numpy as np
import torch
import pyro.distributions as dist
import openpyxl

from LDA_pytorch.ModelBase import LDABase
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix
from utils.wordcloud_util import create_wordcloud


class LDA(LDABase):

    def __init__(self, args, data):
        """
        pytorchでgibbs sampling

        args:
            K,
            auto_beta, auto_alpha,  # bool
            coef_beta, coef_alpha,  # float
            device
        data:
            [doc[sentence[word, word, ...], ...], ...]
        """
        super().__init__()

        # fields
        self.args = args
        self.data = data
        self.device = args.device

        self.K = K = args.K
        self.D = D = len(data)
        self.V = 0
        self.totalN = 0
        self.word_count = {}  # [V]
        self.word_dict = []  # [V]
        self.wordids = []  # [totalN]
        self.docids = []  # [totalN]
        self.z = None  # [totalN]

        self.alpha = torch.full([K], args.coef_alpha, device=self.device, dtype=torch.float64)  # [K]
        self.beta = None  # [V]

        self._wpt = None  # [K, V]
        self._tpd = torch.zeros((D, K), device=self.device, dtype=torch.int64)  # [D, K]
        self._wt = torch.zeros(K, device=self.device, dtype=torch.int64)  # [K]
        self._nd = torch.zeros(D, device=self.device, dtype=torch.int64)  # [D]

        # init wordcount...
        for d, doc in enumerate(data):
            for sentence in doc:
                for word in sentence:
                    if(word not in self.word_count):
                        self.word_count[word] = 0
                    self.word_count[word] += 1
        self.word_dict = [a[0] for a in sorted(self.word_count.items(), key=lambda x: -x[1])]
        self.V = V = len(self.word_dict)
        self.beta = torch.full([V], args.coef_beta, device=self.device, dtype=torch.float64)  # [V]

        # init wordids...
        for d, doc in enumerate(data):
            for sentence in doc:
                for word in sentence:
                    self.wordids.append(self.word_dict.index(word))
                    self.docids.append(d)
        self.totalN = len(self.wordids)
        self.wordids = torch.tensor(self.wordids, device=self.device, dtype=torch.int64)
        self.docids = torch.tensor(self.docids, device=self.device, dtype=torch.int64)
        self.z = torch.randint(0, K, (self.totalN,), device=self.device, dtype=torch.int64)  # [totalN]
        self._wpt = torch.zeros((K, V), device=self.device, dtype=torch.int64)  # [K, V]

        # init wpt...
        ones = torch.ones(self.totalN, device=self.device, dtype=torch.int64)
        self._wpt.index_put_([self.z, self.wordids], ones, accumulate=True)
        self._tpd.index_put_([self.docids, self.z], ones, accumulate=True)
        self._wt = torch.sum(self._wpt, 1)
        self._nd = torch.sum(self._tpd, 1)


    def step(self, subsample_size, parameter_update=False):
        """
        subsample_sizeはgpuのコア数の整数倍が良さそう
        1. update z, wpt, ...
        2. update parameters if neseccary
        """

        rand_perm = torch.randperm(self.totalN, device=self.device)
        for n in range(self.totalN // subsample_size + 1):
            end = self.totalN if n == (self.totalN // subsample_size) else subsample_size * (n + 1)
            idx = rand_perm[subsample_size * n: end]
            self._sampling(idx)

        if(parameter_update):
            self._update_parameters()

        return self.log_perplexity()


    def _sampling(self, idx):
        """
        1. sampleing z
        2. update wpt, tpd, wt
        """

        before = self.z[idx].clone()

        # zをサンプリング
        probs = self._get_z_sampling_probs(idx)
        self.z[idx] = dist.Categorical(probs).sample()

        # wptとか更新
        ones = torch.ones(len(idx), device=self.device, dtype=torch.int64)
        self._wpt.index_put_([before, self.wordids[idx]], -ones, accumulate=True)
        self._wpt.index_put_([self.z[idx], self.wordids[idx]], ones, accumulate=True)
        self._tpd.index_put_([self.docids[idx], before], -ones, accumulate=True)
        self._tpd.index_put_([self.docids[idx], self.z[idx]], ones, accumulate=True)
        self._wt = torch.sum(self._wpt, 1)

        """ calculation checking """
        # for i in range(10):
        #     k = random.randrange(self.K)
        #     v = random.randrange(self.V)
        #     d = random.randrange(self.D)
        #     assert self._wpt[k, v] == torch.sum(torch.logical_and(self.z == k, self.wordids == v))
        #     assert self._tpd[d, k] == torch.sum(torch.logical_and(self.docids == d, self.z == k))


    def _get_z_sampling_probs(self, idx):
        subsample_size = len(idx)
        probs = torch.ones((subsample_size, self.K), device=self.device, dtype=torch.float64)  # [subsample_size, K]
        a = torch.zeros((subsample_size, self.K), device=self.device, dtype=torch.float64)  # [subsample_size, K]
        a[torch.arange(0, subsample_size, device=self.device, dtype=torch.int64), self.z[idx]] = 1

        probs *= self._wpt[:, self.wordids[idx]].T + self.beta[self.wordids[idx]][:, None] - a
        probs /= self._wt[None, :] + self.beta.sum() - a
        probs *= self._tpd[self.docids[idx], :] + self.alpha[None, :] - a
        probs /= self._nd[self.docids[idx]][:, None] + self.alpha.sum() - a

        """ calculation checking """
        # for i in range(100):
        #     n = random.randrange(subsample_size)  # any
        #     k = random.randrange(self.K)  # any
        #     z = self.z[idx[n]]
        #     v = self.wordids[idx[n]]
        #     d = self.docids[idx[n]]
        #     a0 = (1 if z == k else 0)
        #     assert a[n, k] == a0
        #     assert torch.allclose(probs[n, k], (self._wpt[k, v] - a0 + self.beta[v]) / (self._wt[k] + self.beta.sum() - a0)
        #                           * (self._tpd[d, k] - a0 + self.alpha[k]) / (self._nd[d] + self.alpha.sum() - a0))

        probs /= torch.sum(probs, dim=1, keepdim=True)
        return probs


    def _update_parameters(self):
        # TODO
        pass


    def theta(self, to_cpu=True):
        """
        毎回計算するから何度も呼び出さないこと
        """
        theta = (self._tpd + self.alpha[None, :]) / (self._nd[:, None] + self.alpha.sum())  # [D, K]
        if(to_cpu):
            return theta.cpu().detach().numpy().copy()
        else:
            return theta


    def phi(self, to_cpu=True):
        """
        毎回計算するから何度も呼び出さないこと
        """
        phi = (self._wpt + self.beta[None, :]) / (self._wt[:, None] + self.beta.sum())  # [K, V]
        if(to_cpu):
            return phi.cpu().detach().numpy().copy()
        else:
            return phi


    def log_perplexity(self, testset=None):
        if(testset is None):
            theta = self.theta(to_cpu=False)
            phi = self.phi(to_cpu=False)
            p = torch.mm(theta, phi)
            return p[self.docids, self.wordids].log().sum().item()
        else:
            # TODO
            return None


    def _summary_print(self, summary_args):
        D = self.D
        V = self.V
        K = self.K

        alpha = self.alpha.cpu().detach().numpy()
        beta = self.beta.cpu().detach().numpy()
        phi = self.phi()
        theta = self.theta()

        # print summary
        for k in range(K):
            msg = "Topic {:2d}: ".format(k)
            ind = np.argsort(phi[k, :])[::-1]
            for i in range(10):
                msg += "{} ".format(self.word_dict[ind[i]])
            print(msg)
        for k in range(K):
            msg = "Topic {:2d}: ".format(k)
            ind = np.argsort(phi[k, :])[::-1]
            for i in range(10):
                msg += "{:6f} ".format(phi[k, ind[i]])
            print(msg)
        for d in range(3):
            msg = "Documents {:2d}: ".format(d)
            for k in range(K):
                msg += "{:6f} ".format(theta[d, k])
            print(msg)


    def _sammary_wordcloud(self, summary_args):

        p = summary_args.summary_path.joinpath("wordcloud")
        p.mkdir(exist_ok=True, parents=True)

        phi = self.phi()

        for k in range(self.K):
            ind = np.argsort(phi[k, :])[::-1]
            top = [self.word_dict[ind[i]] for i in range(20)]
            create_wordcloud(top, p.joinpath(f"topic{k+1}.png"))


    def _summary_to_excel(self, summary_args, wb):
        D = self.D
        V = self.V
        K = self.K

        alpha = self.alpha.cpu().detach().numpy()
        beta = self.beta.cpu().detach().numpy()
        phi = self.phi()
        theta = self.theta()

        args_dict = {k: self.args.__dict__[k] for k in self.args.__dict__ if not k.startswith("__")}
        for k in args_dict:
            if(not isinstance(args_dict[k], (int, float, complex, bool))):
                args_dict[k] = str(args_dict[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict.values()), axis="row", names=list(args_dict.keys()))

        ws = wb.create_sheet("alpha")
        writeVector(ws, alpha, 1, 1, axis="row",
                    names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("beta")
        writeVector(ws, beta, 1, 1, axis="row",
                    names=self.word_dict,
                    addDataBar=True)

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("phi")
        writeMatrix(ws, phi.T, 1, 1,
                    row_names=self.word_dict,
                    column_names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("phi_sorted")
        writeSortedMatrix(ws, phi.T, axis=0, row=1, column=1,
                          row_names=self.word_dict, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher")
        writeSortedMatrix(ws, phi.T, axis=0, row=1, column=K + 3,
                          row_names=None, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher",
                          addDataBar=True)
