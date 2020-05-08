import pickle
import numpy as np
import openpyxl
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix

import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist


def model(data, args):
    """
    data:
        tensor[D, V]
    args:
        K,
        eps
        autoHyperParam
        coef_beta, coef_alpha
        device
    """
    D = data.shape[0]
    V = data.shape[1]
    K = args.K

    if(args.autoHyperParam):
        beta_hyper = pyro.param("beta_hyper", torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta, constraint=constraints.positive)
        alpha_hyper = pyro.param("alpha_hyper", torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha, constraint=constraints.positive)
    else:
        beta_hyper = torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta
        alpha_hyper = torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha

    with pyro.plate("topics", K):
        phi = pyro.sample("phi", dist.Dirichlet(beta_hyper))

    with pyro.plate("documents", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(alpha_hyper))

        w_d = data[d] + args.eps
        w_d = w_d / torch.sum(w_d, dim=1, keepdim=True)

        p = torch.mm(theta, phi)  # [D, V]
        N = torch.sum(data[d], dim=1)  # [D]
        p = p * N[:, None] + 1  # Dirichlet(p)は最頻値(極大値)がp(w)になる分布
        w = pyro.sample("w", dist.Dirichlet(p), obs=w_d)

        # p = torch.mm(theta, phi)  # [D, V]
        # w = pyro.sample("w", dist.Delta(p, event_dim=1), obs=w_d)

    return phi, theta, w


def guide(data, args):
    D = data.shape[0]
    V = data.shape[1]
    K = args.K

    beta_model = pyro.param("beta_model", torch.ones([K, V], device=args.device, dtype=torch.float64), constraint=constraints.positive)
    alpha_model = pyro.param("alpha_model", torch.ones([D, K], device=args.device, dtype=torch.float64), constraint=constraints.positive)

    with pyro.plate("topics", K) as k:
        phi = pyro.sample("phi", dist.Dirichlet(beta_model[k]))
    with pyro.plate("documents", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(alpha_model[d]))


def summary(data, args, words, reviews, pathResultFolder=None, counts=None):
    beta_model = pyro.param("beta_model")
    alpha_model = pyro.param("alpha_model")
    phi = dist.Dirichlet(beta_model).independent(1).mean
    theta = dist.Dirichlet(alpha_model).independent(1).mean

    beta_model_ = beta_model.cpu().detach().numpy()
    alpha_model_ = alpha_model.cpu().detach().numpy()
    phi_ = phi.cpu().detach().numpy()
    theta_ = theta.cpu().detach().numpy()

    if(args.autoHyperParam):
        beta_hyper = pyro.param("beta_hyper")
        alpha_hyper = pyro.param("alpha_hyper")
        beta_hyper_ = beta_hyper.cpu().detach().numpy()
        alpha_hyper_ = alpha_hyper.cpu().detach().numpy()

    D = data.shape[0]
    V = data.shape[1]
    K = args.K

    # print summary
    for k in range(K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{} ".format(words[ind[i]])
        print(msg)
    for k in range(K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{:6f} ".format(phi_[k, ind[i]])
        print(msg)
    for d in range(3):
        msg = "Documents {:2d}: ".format(d)
        for k in range(K):
            msg += "{:6f} ".format(theta_[d, k])
        print(msg)

    if(pathResultFolder is not None):
        print("saving models")

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}

        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["beta_model"] = beta_model_
        variables["alpha_model"] = alpha_model_
        variables["phi"] = phi_
        variables["theta"] = theta_
        if(args.autoHyperParam):
            variables["beta_hyper"] = beta_hyper_
            variables["alpha_hyper"] = alpha_hyper_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        args_dict_str = args_dict.copy()
        for k in args_dict_str:
            if(not isinstance(args_dict_str[k], (int, float, complex, bool))):
                args_dict_str[k] = str(args_dict_str[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict_str.values()), axis="row", names=list(args_dict_str.keys()))

        ws = wb.create_sheet("alpha_model")
        writeMatrix(ws, alpha_model_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("beta_model")
        writeMatrix(ws, beta_model_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("phi")
        writeMatrix(ws, phi_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        if(args.autoHyperParam):
            ws = wb.create_sheet("beta_hyper")
            writeVector(ws, beta_hyper_, axis="row", names=words,
                        addDataBar=True)

            ws = wb.create_sheet("alpha_hyper")
            writeVector(ws, alpha_hyper_, axis="row", names=[f"topic{d+1}" for d in range(K)],
                        addDataBar=True)

        ws = wb.create_sheet("phi_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=words, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher")

        ws = wb.create_sheet("phi_value_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=None, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher",
                          addDataBar=True)

        # if(counts is not None):
        #     phi2_ = phi_ / counts[None, :]
        #     ws = wb.create_sheet("phi(per num)_sorted")
        #     writeSortedMatrix(ws, phi2_.T, axis=0, row=1, column=1,
        #                       row_names=words, column_names=[f"topic{d+1}" for d in range(K)],
        #                       maxwrite=100, order="higher",)

        #     ws = wb.create_sheet("phi(per num)_value_sorted")
        #     writeSortedMatrix(ws, phi2_.T, axis=0, row=1, column=1,
        #                       row_names=None, column_names=[f"topic{d+1}" for d in range(K)],
        #                       maxwrite=100, order="higher",)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))
