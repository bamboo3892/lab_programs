import pickle
import numpy as np
import openpyxl
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix

import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist


def model(data, args):
    """
    zを生成しないでxを生成
    x ~ N(theta_dk mu_k, sigma)

    data:
        [[tensor[D, V_1], tensor[D, V_2], ... , tensor[D, V_Rt]],  # bow tensors
         tensor[Rm, D],  # measurements(continueous value)
         tensor[Rh, D]]  # habits(categorical value)
    args:
        K,
        n_h,  # number of categories for record "rh"
        eps,
        auto_beta, auto_alpha,
        coef_beta, coef_alpha,
        device
    """
    bows = data[0]
    measurements = data[1]
    habits = data[2]
    Rt = len(bows)
    Rm = len(measurements)
    Rh = len(habits)
    D = measurements[0].shape[0]
    V = [bows[r].shape[1] for r in range(Rt)]  # list[Rt]
    K = args.K
    n_h = args.n_h

    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper", torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha, constraint=constraints.positive)
    else:
        alpha_hyper = torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha

    beta_hyper = [0] * Rt
    for r in pyro.plate("records_rt1", Rt):
        if(args.auto_beta):
            beta_hyper[r] = pyro.param(f"beta_hyper_r{r}", torch.ones([V[r]], device=args.device, dtype=torch.float64) * args.coef_beta, constraint=constraints.positive)
        else:
            beta_hyper[r] = torch.ones([V[r]], device=args.device, dtype=torch.float64) * args.coef_beta

    with pyro.plate("documents", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(alpha_hyper))

    # words
    for r in pyro.plate("records_rt2", Rt):
        with pyro.plate(f"topics_rt{r}", K):
            phi = pyro.sample(f"phi_rt{r}", dist.Dirichlet(beta_hyper[r]))  # [K, V]

        with pyro.plate(f"documents_rt{r}", D) as d:
            w_d = bows[r][d] + args.eps
            w_d = w_d / torch.sum(w_d, dim=1, keepdim=True)

            p = torch.mm(theta, phi)  # [D, V]
            N = torch.sum(bows[r][d], dim=1)  # [D]
            p = p * N[:, None] + 1  # Dirichlet(p)は最頻値(極大値)がp(w)になる分布
            pyro.sample(f"w_rt{r}", dist.Dirichlet(p), obs=w_d)  # [D, V]

    # measurements
    # mu_m_hyper = pyro.param("mu_m_hyper", torch.mean(measurements, 1))  # [Rm]
    # mu_s_hyper = pyro.param("mu_s_hyper", torch.ones([Rm], device=args.device, dtype=torch.float64), constraint=constraints.positive)  # [Rm]
    mu_m_hyper = torch.mean(measurements, 1)   # [Rm]
    mu_s_hyper = torch.std(measurements, 1)   # [Rm]
    # mu_s_hyper = torch.ones([Rm], device=args.device, dtype=torch.float64) * 10  # [Rm]
    for r in pyro.plate("records_rm", Rm):
        # sigma = pyro.param(f"sigma_rm{r}", torch.rand([K], device=args.device, dtype=torch.float64) * torch.std(measurements[r]), constraint=constraints.positive)  # [K]
        sigma = pyro.param(f"sigma_rm{r}", torch.ones([K], device=args.device, dtype=torch.float64) * torch.std(measurements[r]), constraint=constraints.positive)  # [K]
        # sigma = pyro.param(f"sigma_rm{r}", torch.ones([K], device=args.device, dtype=torch.float64), constraint=constraints.positive)  # [K]
        with pyro.plate(f"topics_rm{r}", K) as k:
            mu = pyro.sample(f"mu_rm{r}", dist.Normal(mu_m_hyper[r], mu_s_hyper[r]))   # [K]
        with pyro.plate(f"documents_rm{r}", D) as d:
            mean = torch.mv(theta[d], mu)  # [D]
            std = torch.mv(theta[d], sigma)  # [D]
            x = pyro.sample(f"x_rm{r}", dist.Normal(mean, std), obs=measurements[r])  # [D]

    # habits
    for r in pyro.plate("records_rh", Rh):
        rho_hyper = torch.ones([n_h[r]], device=args.device, dtype=torch.float64)  # [n_h[r]]
        with pyro.plate(f"topics_rh{r}", K) as k:
            rho = pyro.sample(f"rho_rh{r}", dist.Dirichlet(rho_hyper))  # [K, n_h[r]]
        with pyro.plate(f"documents_rh{r}", D) as d:
            p = torch.mm(theta[d], rho)  # [D, n_h[r]]
            x = pyro.sample(f"x_rh{r}", dist.Categorical(p), obs=habits[r])  # [D]

    return phi, theta


def guide(data, args):
    bows = data[0]
    measurements = data[1]
    habits = data[2]
    Rt = len(bows)
    Rm = len(measurements)
    Rh = len(habits)
    D = measurements[0].shape[0]
    V = [bows[r].shape[1] for r in range(Rt)]  # list[Rt]
    K = args.K
    n_h = args.n_h

    theta_guide = pyro.param("theta_guide", torch.ones([D, K], device=args.device, dtype=torch.float64), constraint=constraints.positive)
    with pyro.plate("documents", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(theta_guide[d]))

    phi_guide = [0] * Rt
    for r in pyro.plate("records_rt1", Rt):
        phi_guide[r] = pyro.param(f"phi_guide_r{r}", torch.ones([K, V[r]], device=args.device, dtype=torch.float64), constraint=constraints.positive)

    # words
    for r in pyro.plate("records_rt2", Rt):
        with pyro.plate(f"topics_rt{r}", K) as k:
            phi = pyro.sample(f"phi_rt{r}", dist.Dirichlet(phi_guide[r][k]))  # [K, V]

    # measurements
    mu_m_guide = pyro.param("mu_m_guide", (torch.rand([Rm, K], device=args.device, dtype=torch.float64) + 0.5) * torch.mean(measurements, 1)[:, None])  # [Rm, K]
    mu_s_guide = pyro.param("mu_s_guide", torch.ones([Rm, K], device=args.device, dtype=torch.float64) * 10, constraint=constraints.positive)  # [Rm, K]
    for r in pyro.plate("records_rm", Rm):
        with pyro.plate(f"topics_rm{r}", K) as k:
            mu = pyro.sample(f"mu_rm{r}", dist.Normal(mu_m_guide[r, k], mu_s_guide[r, k]))   # [K]

    # habits
    for r in pyro.plate("records_rh", Rh):
        rho_guide = pyro.param(f"rho_guide_rh{r}", torch.ones([K, n_h[r]], device=args.device, dtype=torch.float64), constraint=constraints.positive)  # [K, n_h[r]]
        with pyro.plate(f"topics_rh{r}", K) as k:
            rho = pyro.sample(f"rho_rh{r}", dist.Dirichlet(rho_guide))  # [K, n_h[r]]


def summary(data, args, words, reviews, pathResultFolder=None, counts=None):
    bows = data[0]
    measurements = data[1]
    habits = data[2]
    Rt = len(bows)
    Rm = len(measurements)
    Rh = len(habits)
    D = bows[0].shape[0]
    V = [bows[r].shape[1] for r in range(Rt)]  # list[Rt]
    K = args.K
    n_h = args.n_h

    # theta
    theta_guide = pyro.param("theta_guide")
    theta_guide_ = theta_guide.cpu().detach().numpy()
    theta = dist.Dirichlet(theta_guide).independent(1).mean
    theta_ = theta.cpu().detach().numpy()

    # phi
    phi_guide = [0] * Rt
    phi_guide_ = [0] * Rt
    phi = [0] * Rt
    phi_ = [0] * Rt
    for r in range(Rt):
        phi_guide[r] = pyro.param(f"phi_guide_r{r}")
        phi_guide_[r] = phi_guide[r].cpu().detach().numpy()
        phi[r] = dist.Dirichlet(phi_guide[r]).independent(1).mean
        phi_[r] = phi[r].cpu().detach().numpy()

    # alpha
    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper")
    else:
        alpha_hyper = torch.ones([K]) * args.coef_alpha
    alpha_hyper_ = alpha_hyper.cpu().detach().numpy()

    # beta
    beta_hyper = [0] * Rt
    beta_hyper_ = [0] * Rt
    for r in range(Rt):
        if(args.auto_beta):
            beta_hyper[r] = pyro.param(f"beta_hyper_r{r}")
        else:
            beta_hyper[r] = torch.ones([V[r]]) * args.coef_beta
        beta_hyper_[r] = beta_hyper[r].cpu().detach().numpy()

    # mu, sigma
    mu_m_guide = pyro.param("mu_m_guide")
    mu = mu_m_guide
    mu_ = mu.cpu().detach().numpy()
    sigma = [0] * Rm
    sigma_ = [0] * Rm
    for r in range(Rm):
        sigma[r] = pyro.param(f"sigma_rm{r}")
        # sigma[r] = torch.ones([K], dtype=torch.float64) * torch.std(measurements[r]) * 0.5
        sigma_[r] = sigma[r].cpu().detach().numpy()
    sigma_ = np.array(sigma_)

    # rho
    rho = [0] * Rh
    rho_ = [0] * Rh
    for r in range(Rh):
        rho_guide = pyro.param(f"rho_guide_rh{r}")
        rho[r] = dist.Dirichlet(rho_guide).independent(1).mean
        # rho[r] = pyro.param(f"rho_rh{r}")
        rho_[r] = rho[r].cpu().detach().numpy()

    if(pathResultFolder is not None):
        print("saving models")

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}

        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["phi_guide"] = phi_guide_
        variables["theta_guide"] = theta_guide_
        variables["phi"] = phi_
        variables["theta"] = theta_
        variables["beta_hyper"] = beta_hyper_
        variables["alpha_hyper"] = alpha_hyper_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        args_dict_str = args_dict.copy()
        for k in args_dict_str:
            if(not isinstance(args_dict_str[k], (int, float, complex, bool))):
                args_dict_str[k] = str(args_dict_str[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict_str.values()), axis="row", names=list(args_dict_str.keys()))

        ws = wb.create_sheet("mu_sigma")
        writeMatrix(ws, mu_, 1, 1,
                    row_names=[f"record_m{r+1}" for r in range(Rm)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)
        writeMatrix(ws, sigma_, 1, K + 3,
                    row_names=[f"record_m{r+1}" for r in range(Rm)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("rho")
        row = 1
        for r in range(Rh):
            writeMatrix(ws, rho_[r].T, row, 1,
                        row_names=[f"category{c+1}" for c in range(n_h[r])],
                        column_names=[f"topic{k+1}" for k in range(K)],
                        addDataBar=True)
            _, counts = torch.unique(habits[r], return_counts=True)
            counts = counts.cpu().detach().numpy()
            counts = counts / np.sum(counts)
            writeMatrix(ws, counts[:, None], row, K + 3, column_names=["data distribution"], addDataBar=True)
            row += n_h[r] + 2

        ws = wb.create_sheet("alpha_hyper")
        writeVector(ws, alpha_hyper_, axis="row", names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("beta_hyper")
        for r in range(Rt):
            writeVector(ws, beta_hyper_[r], column=r * 3 + 1,
                        axis="row", names=words[r],
                        addDataBar=True)

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        for r in range(Rt):
            ws = wb.create_sheet(f"phi_r{r}")
            writeMatrix(ws, phi_[r].T, 1, 1,
                        row_names=words[r],
                        column_names=[f"topic{k+1}" for k in range(K)],
                        addDataBar=True)

            ws = wb.create_sheet(f"phi_r{r}_sorted")
            writeSortedMatrix(ws, phi_[r].T, axis=0, row=1, column=1,
                              row_names=words[r], column_names=[f"topic{k+1}" for k in range(K)],
                              maxwrite=100, order="higher")
            writeSortedMatrix(ws, phi_[r].T, axis=0, row=1, column=K + 3,
                              row_names=None, column_names=[f"topic{k+1}" for k in range(K)],
                              maxwrite=100, order="higher",
                              addDataBar=True)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))

        # topics file
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        for k in range(K):
            ws = wb.create_sheet(f"topic_{k}")
            for r in range(Rt):
                ws.cell(1, r + 1, f"r{r}")
                idx = np.argsort(phi_[r][k])[::-1]
                dat = np.array(words[r])[idx].tolist()
                writeVector(ws, dat, 2, r + 1, axis="row", names=None)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("topics.xlsx"))
