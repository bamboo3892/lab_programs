import pickle
import openpyxl
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix

import numpy as np
import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist


def model(data, args):
    """
    data: [tensor[D, V], attributions[S]]
    """

    phi = pyro.sample("phi", dist.Dirichlet(torch.ones([args.K, args.V]) * args.coef_phi).independent(1))
    sigma = pyro.sample("sigma", dist.Dirichlet(torch.ones([args.S, args.V]) * args.coef_sigma).independent(1))

    with pyro.plate("documents", args.D) as d:

        theta = pyro.sample("theta", dist.Dirichlet(torch.ones([args.K]) * args.coef_theta))

        w_d = data[0][d] + args.eps
        w_d = w_d / torch.sum(w_d, dim=1, keepdim=True)

        p1 = torch.mm(theta, phi)  # [D, V]
        p2 = sigma[data[1]]  # [D, V]
        p = p1 * p2
        p = p / torch.sum(p, dim=1, keepdim=True)  # w ~ Categorical(p)

        N = torch.sum(data[0][d], dim=1)  # [D]
        p = p * N[:, None] + 1  # Dirichlet(p)は最頻値(極大値)がp(w)になる分布

        w = pyro.sample("w", dist.Dirichlet(p), obs=w_d)

    return phi, theta, w


def guide(data, args):
    alpha = pyro.param("alpha", torch.ones([args.D, args.K]), constraint=constraints.positive)
    beta = pyro.param("beta", torch.ones([args.K, args.V]), constraint=constraints.positive)
    gamma = pyro.param("gamma", torch.ones([args.S, args.V]), constraint=constraints.positive)

    phi = pyro.sample("phi", dist.Dirichlet(beta).independent(1))
    sigma = pyro.sample("sigma", dist.Dirichlet(gamma).independent(1))

    with pyro.plate("documents", args.D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(alpha[d]))


def summary(data, args, words, reviews, pathResultFolder=None):
    beta = pyro.param("beta")
    alpha = pyro.param("alpha")
    phi = dist.Dirichlet(beta).independent(1).mean
    theta = dist.Dirichlet(alpha).independent(1).mean

    beta_ = beta.detach().numpy()
    alpha_ = alpha.detach().numpy()
    phi_ = phi.detach().numpy()
    theta_ = theta.detach().numpy()
    for k in range(args.K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{} ".format(words[ind[i]])
        print(msg)
    for k in range(args.K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{:6f} ".format(phi_[k, ind[i]])
        print(msg)
    for d in range(3):
        msg = "Documents {:2d}: ".format(d)
        for k in range(args.K):
            msg += "{:6f} ".format(theta_[d, k])
        print(msg)

    gamma = pyro.param("gamma")
    sigma = dist.Dirichlet(gamma).independent(1).mean
    gamma_ = gamma.detach().numpy()
    sigma_ = sigma.detach().numpy()
    for s in range(args.S):
        msg = "Topic {:2d}: ".format(s)
        ind = np.argsort(sigma_[s, :])[::-1]
        for i in range(10):
            msg += "{} ".format(words[ind[i]])
        print(msg)
    for s in range(args.S):
        msg = "Topic {:2d}: ".format(s)
        ind = np.argsort(sigma_[s, :])[::-1]
        for i in range(10):
            msg += "{:6f} ".format(sigma_[s, ind[i]])
        print(msg)

    # p0 = phi_ * sigma_[0, :][None, :]  # [K, S]
    # p0 = p0 / np.sum(p0, axis=1, keepdims=True)

    if(pathResultFolder is not None):

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}
        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["beta"] = beta_
        variables["alpha"] = alpha_
        variables["phi"] = phi_
        variables["theta"] = theta_
        variables["gamma"] = gamma_
        variables["sigma"] = sigma_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        tmp_ws = wb.get_active_sheet()

        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict.values()), axis="row", names=list(args_dict.keys()))

        ws = wb.create_sheet("alpha")
        writeMatrix(ws, alpha_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(args.D)],
                    column_names=[f"topic{k+1}" for k in range(args.K)])

        ws = wb.create_sheet("beta")
        writeMatrix(ws, beta_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(args.K)])

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(args.D)],
                    column_names=[f"topic{k+1}" for k in range(args.K)])

        ws = wb.create_sheet("phi")
        writeMatrix(ws, phi_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(args.K)])

        ws = wb.create_sheet("gamma")
        writeMatrix(ws, gamma_.T, 1, 1,
                    row_names=words,
                    column_names=[f"attr{s+1}" for s in range(args.S)])

        ws = wb.create_sheet("sigma")
        writeMatrix(ws, sigma_.T, 1, 1,
                    row_names=words,
                    column_names=[f"attr{s+1}" for s in range(args.S)])

        ws = wb.create_sheet("phi_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=words, column_names=[f"topic{d+1}" for d in range(args.K)],
                          maxwrite=100, order="higher",)

        ws = wb.create_sheet("phi_value_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=None, column_names=[f"topic{d+1}" for d in range(args.K)],
                          maxwrite=100, order="higher",)

        ws = wb.create_sheet("sigma_sorted")
        writeSortedMatrix(ws, sigma_.T, axis=0, row=1, column=1,
                          row_names=words, column_names=[f"attr{s+1}" for s in range(args.S)],
                          maxwrite=100, order="higher",)

        ws = wb.create_sheet("sigma_value_sorted")
        writeSortedMatrix(ws, sigma_.T, axis=0, row=1, column=1,
                          row_names=None, column_names=[f"attr{s+1}" for s in range(args.S)],
                          maxwrite=100, order="higher",)

        for s in range(args.S):
            ws = wb.create_sheet(f"phi-sigma{s}_sorted")
            p0 = phi_ * sigma_[s, :][None, :]  # [K, S]
            p0 = p0 / np.sum(p0, axis=1, keepdims=True)
            writeSortedMatrix(ws, p0.T, axis=0, row=1, column=1,
                              row_names=words, column_names=[f"topic{d+1}" for d in range(args.K)],
                              maxwrite=100, order="higher",)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))
