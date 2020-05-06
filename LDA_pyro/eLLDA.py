import pickle
import numpy as np
import openpyxl
from utils.openpyxl_util import writeVector, writeMatrix, writeSortedMatrix

import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist


def model(data, args):

    phi = pyro.sample("phi", dist.Dirichlet(torch.ones([args.K, args.V]) * args.coef_phi).independent(1))

    with pyro.plate("documents", args.D) as d:

        theta = pyro.sample("theta", dist.Dirichlet(data[1][d]))

        w_d = data[0][d] + args.eps
        w_d = w_d / torch.sum(w_d, dim=1, keepdim=True)

        p = torch.mm(theta, phi)  # [D, V]
        N = torch.sum(data[0][d], dim=1)  # [D]
        p = p * N[:, None] + 1  # Dirichlet(p)は最頻値(極大値)がp(w)になる分布
        w = pyro.sample("w", dist.Dirichlet(p), obs=w_d)

        # p = torch.mm(theta, phi)  # [D, V]
        # w = pyro.sample("w", dist.Delta(p, event_dim=1), obs=w_d)

    return phi, theta, w


def guide(data, args):
    beta = pyro.param("beta", torch.ones([args.K, args.V]), constraint=constraints.positive)
    alpha = pyro.param("alpha", torch.ones([args.D, args.K]), constraint=constraints.positive)

    phi = pyro.sample("phi", dist.Dirichlet(beta).independent(1))
    with pyro.plate("documents", args.D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(alpha[d]))


def summary(data, args, words, reviews, pathResultFolder=None):
    print("saving models")

    beta = pyro.param("beta")
    alpha = pyro.param("alpha")
    phi = dist.Dirichlet(beta).independent(1).mean
    theta = dist.Dirichlet(alpha).independent(1).mean

    beta_ = beta.detach().numpy()
    alpha_ = alpha.detach().numpy()
    phi_ = phi.detach().numpy()
    theta_ = theta.detach().numpy()
    for k in range(args.K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{} ".format(words[ind[i]])
        print(msg)
    for k in range(args.K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{:6f} ".format(phi_[k, ind[i]])
        print(msg)
    for d in range(3):
        msg = "Documents {:2d}: ".format(d)
        for k in range(args.K):
            msg += "{:6f} ".format(theta_[d, k])
        print(msg)

    if(pathResultFolder is not None):

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}
        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["beta"] = beta_
        variables["alpha"] = alpha_
        variables["phi"] = phi_
        variables["theta"] = theta_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        # tmp_ws = wb.get_active_sheet()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict.values()), axis="row", names=list(args_dict.keys()))

        ws = wb.create_sheet("alpha")
        writeMatrix(ws, alpha_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(args.D)],
                    column_names=[f"topic{k+1}" for k in range(args.K)])

        ws = wb.create_sheet("beta")
        writeMatrix(ws, beta_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(args.K)])

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(args.D)],
                    column_names=[f"topic{k+1}" for k in range(args.K)])

        ws = wb.create_sheet("phi")
        writeMatrix(ws, phi_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(args.K)])

        ws = wb.create_sheet("phi_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=words, column_names=[f"topic{d+1}" for d in range(args.K)],
                          maxwrite=100, order="higher",)

        ws = wb.create_sheet("phi_value_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=None, column_names=[f"topic{d+1}" for d in range(args.K)],
                          maxwrite=100, order="higher",)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))
