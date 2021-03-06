import pickle
import numpy as np
import openpyxl
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix

import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist


def model(data, args):
    """
    data:
        [tensor[D, V1], tensor[D, V2], ...]
    args:
        K,
        eps,
        auto_beta, auto_alpha,
        coef_beta, coef_alpha,
        device
    """
    R = len(data)
    D = data[0].shape[0]
    V = [data[r].shape[1] for r in range(R)]
    K = args.K

    # hyper params
    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper", torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha, constraint=constraints.positive)
    else:
        alpha_hyper = torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha

    beta_hyper = [0] * R
    for r in pyro.plate("records1", R):
        if(args.auto_beta):
            beta_hyper[r] = pyro.param(f"beta_hyper_r{r}", torch.ones([V[r]], device=args.device, dtype=torch.float64) * args.coef_beta, constraint=constraints.positive)
        else:
            beta_hyper[r] = torch.ones([V[r]], device=args.device, dtype=torch.float64) * args.coef_beta

    # theta
    with pyro.plate("documents1", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(alpha_hyper))

    for r in pyro.plate("records2", R):
        with pyro.plate(f"topics_r{r}", K):
            phi = pyro.sample(f"phi_r{r}", dist.Dirichlet(beta_hyper[r]))
        with pyro.plate(f"documents2_r{r}", D) as d:
            w_d = data[r][d] + args.eps
            w_d = w_d / torch.sum(w_d, dim=1, keepdim=True)

            p = torch.mm(theta, phi)  # [D, V]
            N = torch.sum(data[r][d], dim=1)  # [D]
            p = p * N[:, None] + 1  # Dirichlet(p)は最頻値(極大値)がp(w)になる分布
            w = pyro.sample(f"w_r{r}", dist.Dirichlet(p), obs=w_d)

    return phi, theta, w


def guide(data, args):
    R = len(data)
    D = data[0].shape[0]
    V = [data[r].shape[1] for r in range(R)]
    K = args.K

    theta_guide = pyro.param("theta_guide", torch.ones([D, K], device=args.device, dtype=torch.float64), constraint=constraints.positive)
    phi_guide = [0] * R
    for r in pyro.plate("records1", R):
        phi_guide[r] = pyro.param(f"phi_guide_r{r}", torch.ones([K, V[r]], device=args.device, dtype=torch.float64), constraint=constraints.positive)

    with pyro.plate("documents1", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(theta_guide[d]))

    for r in pyro.plate("records2", R):
        with pyro.plate(f"topics_r{r}", K) as k:
            phi = pyro.sample(f"phi_r{r}", dist.Dirichlet(phi_guide[r][k]))


def summary(data, args, words, reviews, pathResultFolder=None, counts=None):
    R = len(data)
    D = data[0].shape[0]
    V = [data[r].shape[1] for r in range(R)]
    K = args.K

    theta_guide = pyro.param("theta_guide")
    theta_guide_ = theta_guide.cpu().detach().numpy()
    theta = dist.Dirichlet(theta_guide).independent(1).mean
    theta_ = theta.cpu().detach().numpy()

    phi_guide = [0] * R
    phi_guide_ = [0] * R
    phi = [0] * R
    phi_ = [0] * R
    for r in range(R):
        phi_guide[r] = pyro.param(f"phi_guide_r{r}")
        phi_guide_[r] = phi_guide[r].cpu().detach().numpy()
        phi[r] = dist.Dirichlet(phi_guide[r]).independent(1).mean
        phi_[r] = phi[r].cpu().detach().numpy()

    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper")
    else:
        alpha_hyper = torch.ones([K]) * args.coef_alpha
    alpha_hyper_ = alpha_hyper.cpu().detach().numpy()

    beta_hyper = [0] * R
    beta_hyper_ = [0] * R
    for r in pyro.plate("records1", R):
        if(args.auto_beta):
            beta_hyper[r] = pyro.param(f"beta_hyper_r{r}")
        else:
            beta_hyper[r] = torch.ones([V[r]]) * args.coef_beta
        beta_hyper_[r] = beta_hyper[r].cpu().detach().numpy()

    if(pathResultFolder is not None):
        print("saving models")

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}

        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["phi_guide"] = phi_guide_
        variables["theta_guide"] = theta_guide_
        variables["phi"] = phi_
        variables["theta"] = theta_
        variables["beta_hyper"] = beta_hyper_
        variables["alpha_hyper"] = alpha_hyper_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        args_dict_str = args_dict.copy()
        for k in args_dict_str:
            if(not isinstance(args_dict_str[k], (int, float, complex, bool))):
                args_dict_str[k] = str(args_dict_str[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict_str.values()), axis="row", names=list(args_dict_str.keys()))

        ws = wb.create_sheet("alpha_hyper")
        writeVector(ws, alpha_hyper_, axis="row", names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("beta_hyper")
        for r in range(R):
            writeVector(ws, beta_hyper_[r], column=r * 3 + 1,
                        axis="row", names=words[r],
                        addDataBar=True)

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        for r in range(R):
            ws = wb.create_sheet(f"phi_r{r}")
            writeMatrix(ws, phi_[r].T, 1, 1,
                        row_names=words[r],
                        column_names=[f"topic{k+1}" for k in range(K)],
                        addDataBar=True)

            ws = wb.create_sheet(f"phi_r{r}_sorted")
            writeSortedMatrix(ws, phi_[r].T, axis=0, row=1, column=1,
                              row_names=words[r], column_names=[f"topic{k+1}" for k in range(K)],
                              maxwrite=100, order="higher")
            writeSortedMatrix(ws, phi_[r].T, axis=0, row=1, column=K + 3,
                              row_names=None, column_names=[f"topic{k+1}" for k in range(K)],
                              maxwrite=100, order="higher",
                              addDataBar=True)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))


        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        for k in range(K):
            ws = wb.create_sheet(f"topic_{k}")
            for r in range(R):
                ws.cell(1, r + 1, f"r{r}")
                idx = np.argsort(phi_[r][k])[::-1]
                dat = np.array(words[r])[idx].tolist()
                writeVector(ws, dat, 2, r + 1, axis="row", names=None)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("topics.xlsx"))
