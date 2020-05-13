import pickle
import numpy as np
import openpyxl
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix

import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist



def model(data, args):
    """
    data:
        [tensor[totalN] (word id),
         tensor[totalN] (doc id)]
    args:
        K,
        auto_beta, auto_alpha
        coef_beta, coef_alpha
        device
    """
    D = args.D
    V = args.V
    totalN = len(data[0])
    K = args.K

    if(args.auto_beta):
        beta_hyper = pyro.param("beta_hyper", torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta, constraint=constraints.positive)
    else:
        beta_hyper = torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta

    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper", torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha, constraint=constraints.positive)
    else:
        alpha_hyper = torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha

    with pyro.plate("topics", K):
        phi = pyro.sample("phi", dist.Dirichlet(beta_hyper))

    with pyro.plate("documents", D):
        theta = pyro.sample("theta", dist.Dirichlet(alpha_hyper))

    with pyro.plate("words", totalN, subsample_size=10000) as n:
        data0 = data[0][n].long()
        data1 = data[1][n].long()
        z = pyro.sample("z", dist.Categorical(theta[data1]))
        w = pyro.sample("w", dist.Categorical(phi[z]), obs=data0)

    return phi, theta, z, w


def guide(data, args):
    D = args.D
    V = args.V
    totalN = len(data[0])
    K = args.K

    beta_model = pyro.param("beta_model", torch.ones([K, V], device=args.device, dtype=torch.float64), constraint=constraints.positive)
    with pyro.plate("topics", K):
        phi = pyro.sample("phi", dist.Dirichlet(beta_model))

    alpha_model = pyro.param("alpha_model", torch.ones([D, K], device=args.device, dtype=torch.float64), constraint=constraints.positive)
    with pyro.plate("documents", D):
        theta = pyro.sample("theta", dist.Dirichlet(alpha_model))

    # q_model = pyro.param("q_model", torch.ones([totalN, K], device=args.device, dtype=torch.float64) / K, constraint=constraints.simplex)
    q_model = pyro.param("q_model", torch.randn([totalN, K], device=args.device, dtype=torch.float64).exp(), constraint=constraints.simplex)
    with pyro.plate("words", totalN, subsample_size=10000) as n:
        z = pyro.sample("z", dist.Categorical(q_model[n]))


def summary(data, args, words, reviews, pathResultFolder=None, counts=None):

    wordIDs = data[0].cpu().detach().numpy()
    docIDs = data[1].cpu().detach().numpy()

    D = args.D
    V = args.V
    K = args.K

    beta_model = pyro.param("beta_model")
    alpha_model = pyro.param("alpha_model")
    q_model = pyro.param("q_model")
    phi = dist.Dirichlet(beta_model).independent(1).mean
    theta = dist.Dirichlet(alpha_model).independent(1).mean

    beta_model_ = beta_model.cpu().detach().numpy()
    alpha_model_ = alpha_model.cpu().detach().numpy()
    q_model_ = q_model.cpu().detach().numpy()
    phi_ = phi.cpu().detach().numpy()
    theta_ = theta.cpu().detach().numpy()


    if(args.auto_beta):
        beta_hyper = pyro.param("beta_hyper")
        beta_hyper_ = beta_hyper.cpu().detach().numpy()
    else:
        beta_hyper_ = np.ones(V) * args.coef_beta
    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper")
        alpha_hyper_ = alpha_hyper.cpu().detach().numpy()
    else:
        alpha_hyper_ = np.ones(K) * args.coef_alpha

    # print summary
    for k in range(K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{} ".format(words[ind[i]])
        print(msg)
    for k in range(K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{:6f} ".format(phi_[k, ind[i]])
        print(msg)
    for d in range(3):
        msg = "Documents {:2d}: ".format(d)
        for k in range(K):
            msg += "{:6f} ".format(theta_[d, k])
        print(msg)

    if(pathResultFolder is not None):
        print("saving models")

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}

        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["beta_model"] = beta_model_
        variables["alpha_model"] = alpha_model_
        variables["q_model"] = q_model_
        variables["phi"] = phi_
        variables["theta"] = theta_
        variables["beta_hyper"] = beta_hyper_
        variables["alpha_hyper"] = alpha_hyper_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        args_dict_str = args_dict.copy()
        for k in args_dict_str:
            if(not isinstance(args_dict_str[k], (int, float, complex, bool))):
                args_dict_str[k] = str(args_dict_str[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict_str.values()), axis="row", names=list(args_dict_str.keys()))

        ws = wb.create_sheet("alpha_model")
        writeMatrix(ws, alpha_model_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("beta_model")
        writeMatrix(ws, beta_model_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("q_model")
        writeVector(ws, docIDs[:100], 2, 1, axis="row")
        writeVector(ws, np.array(words)[wordIDs[:100]], 2, 2, axis="row")
        writeMatrix(ws, q_model_[:100, :], 1, 3,
                    row_names=None,
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("phi")
        writeMatrix(ws, phi_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("alpha_hyper")
        writeVector(ws, alpha_hyper_, axis="row", names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("beta_hyper")
        writeVector(ws, beta_hyper_, axis="row", names=words,
                    addDataBar=True)

        ws = wb.create_sheet("phi_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=words, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher")

        ws = wb.create_sheet("phi_value_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=None, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher",
                          addDataBar=True)

        # if(counts is not None):
        #     phi2_ = phi_ / counts[None, :]
        #     ws = wb.create_sheet("phi(per num)_sorted")
        #     writeSortedMatrix(ws, phi2_.T, axis=0, row=1, column=1,
        #                       row_names=words, column_names=[f"topic{d+1}" for d in range(K)],
        #                       maxwrite=100, order="higher",)

        #     ws = wb.create_sheet("phi(per num)_value_sorted")
        #     writeSortedMatrix(ws, phi2_.T, axis=0, row=1, column=1,
        #                       row_names=None, column_names=[f"topic{d+1}" for d in range(K)],
        #                       maxwrite=100, order="higher",)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))
