import pickle
import numpy as np
import openpyxl
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix

import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist


def model(data, args):
    """
    zを生成しないでxを生成
    x ~ N(theta_dk mu_k, sigma)

    data:
        [tensor[Rm, D],  # measurements(continueous value)
         tensor[Rh, D]]  # habits(categorical value)
    args:
        K,
        n_h,  # number of categories for record "rh"
        eps,
        auto_beta, auto_alpha,
        coef_beta, coef_alpha,
        device
    """
    measurements = data[0]
    Rm = len(measurements)
    D = measurements[0].shape[0]
    K = args.K
    habits = data[1]
    Rh = len(habits)
    n_h = args.n_h

    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper", torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha, constraint=constraints.positive)
    else:
        alpha_hyper = torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha

    with pyro.plate("documents", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(alpha_hyper))  # [D, K]

    # measurements
    # mu_m_hyper = pyro.param("mu_m_hyper", torch.mean(measurements, 1))  # [Rm]
    # mu_s_hyper = pyro.param("mu_s_hyper", torch.ones([Rm], device=args.device, dtype=torch.float64), constraint=constraints.positive)  # [Rm]
    mu_m_hyper = torch.mean(measurements, 1)   # [Rm]
    mu_s_hyper = torch.std(measurements, 1)   # [Rm]
    # mu_s_hyper = torch.ones([Rm], device=args.device, dtype=torch.float64) * 10  # [Rm]
    for r in pyro.plate("records_rm", Rm):
        # sigma = pyro.param(f"sigma_rm{r}", torch.rand([K], device=args.device, dtype=torch.float64) * torch.std(measurements[r]), constraint=constraints.positive)  # [K]
        sigma = pyro.param(f"sigma_rm{r}", torch.ones([K], device=args.device, dtype=torch.float64) * torch.std(measurements[r]), constraint=constraints.positive)  # [K]
        # sigma = pyro.param(f"sigma_rm{r}", torch.ones([K], device=args.device, dtype=torch.float64), constraint=constraints.positive)  # [K]
        with pyro.plate(f"topics_rm{r}", K) as k:
            mu = pyro.sample(f"mu_rm{r}", dist.Normal(mu_m_hyper[r], mu_s_hyper[r]))   # [K]
        with pyro.plate(f"documents_rm{r}", D) as d:
            mean = torch.mv(theta[d], mu)  # [D]
            std = torch.mv(theta[d], sigma)  # [D]
            x = pyro.sample(f"x_rm{r}", dist.Normal(mean, std), obs=measurements[r])  # [D]

    # habits
    for r in pyro.plate("records_rh", Rh):
        rho_hyper = torch.ones([n_h[r]], device=args.device, dtype=torch.float64)  # [n_h[r]]
        with pyro.plate(f"topics_rh{r}", K) as k:
            rho = pyro.sample(f"rho_rh{r}", dist.Dirichlet(rho_hyper))  # [K, n_h[r]]
        # rho = pyro.param(f"rho_rh{r}", torch.ones([K, n_h[r]], device=args.device, dtype=torch.float64) / n_h[r], constraint=constraints.simplex)  # [K, n_h[r]]
        with pyro.plate(f"documents_rh{r}", D) as d:
            p = torch.mm(theta[d], rho)  # [D, n_h[r]]
            x = pyro.sample(f"x_rh{r}", dist.Categorical(p), obs=habits[r])  # [D]
            # z = pyro.sample(f"z_rh{r}", dist.Categorical(theta[d]))  # [D]
            # x = pyro.sample(f"x_rh{r}", dist.Categorical(rho[z]), obs=habits[r])  # [D]

    return


def guide(data, args):
    K = args.K
    measurements = data[0]
    Rm = len(measurements)
    D = measurements[0].shape[0]
    habits = data[1]
    Rh = len(habits)
    n_h = args.n_h

    # theta_guide = pyro.param("theta_guide", torch.rand([D, K], device=args.device, dtype=torch.float64), constraint=constraints.positive)
    theta_guide = pyro.param("theta_guide", torch.ones([D, K], device=args.device, dtype=torch.float64), constraint=constraints.positive)
    with pyro.plate("documents", D) as d:
        theta = pyro.sample("theta", dist.Dirichlet(theta_guide[d]))

    # measurements
    # mu_m_guide = pyro.param("mu_m_guide", torch.rand([Rm, K], device=args.device, dtype=torch.float64) * 100)  # [Rm, K]
    # mu_s_guide = pyro.param("mu_s_guide", torch.rand([Rm, K], device=args.device, dtype=torch.float64) * 10, constraint=constraints.positive)  # [Rm, K]
    mu_m_guide = pyro.param("mu_m_guide", (torch.rand([Rm, K], device=args.device, dtype=torch.float64) + 0.5) * torch.mean(measurements, 1)[:, None])  # [Rm, K]
    mu_s_guide = pyro.param("mu_s_guide", torch.ones([Rm, K], device=args.device, dtype=torch.float64) * 10, constraint=constraints.positive)  # [Rm, K]
    for r in pyro.plate("records_rm", Rm):
        with pyro.plate(f"topics_rm{r}", K) as k:
            mu = pyro.sample(f"mu_rm{r}", dist.Normal(mu_m_guide[r, k], mu_s_guide[r, k]))   # [K]

    # habits
    for r in pyro.plate("records_rh", Rh):
        rho_guide = pyro.param(f"rho_guide_rh{r}", torch.ones([K, n_h[r]], device=args.device, dtype=torch.float64), constraint=constraints.positive)  # [K, n_h[r]]
        with pyro.plate(f"topics_rh{r}", K) as k:
            rho = pyro.sample(f"rho_rh{r}", dist.Dirichlet(rho_guide))  # [K, n_h[r]]
        # z_guide = pyro.param(f"z_rh{r}_guide", torch.ones([D, K], device=args.device, dtype=torch.float64), constraint=constraints.positive)  # [D * K]
        # with pyro.plate(f"documents_rh{r}", D) as d:
        #     z = pyro.sample(f"z_rh{r}", dist.Categorical(z_guide[d]))  # [D]

    return


def summary(data, args, words, reviews, pathResultFolder=None, counts=None):
    K = args.K
    measurements = data[0]
    Rm = len(measurements)
    D = measurements[0].shape[0]
    habits = data[1]
    Rh = len(habits)
    n_h = args.n_h

    # theta
    theta_guide = pyro.param("theta_guide")
    theta_guide_ = theta_guide.cpu().detach().numpy()
    theta = dist.Dirichlet(theta_guide).independent(1).mean
    theta_ = theta.cpu().detach().numpy()

    # alpha
    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper")
    else:
        alpha_hyper = torch.ones([K]) * args.coef_alpha
    alpha_hyper_ = alpha_hyper.cpu().detach().numpy()

    # mu, sigma
    mu_m_guide = pyro.param("mu_m_guide")
    mu = mu_m_guide
    mu_ = mu.cpu().detach().numpy()
    sigma = [0] * Rm
    sigma_ = [0] * Rm
    for r in range(Rm):
        if(f"sigma_rm{r}" in pyro.get_param_store()):
            sigma[r] = pyro.param(f"sigma_rm{r}")
        else:
            sigma[r] = torch.ones([K], dtype=torch.float64) * torch.std(measurements[r]) * 0.5
        sigma_[r] = sigma[r].cpu().detach().numpy()
    sigma_ = np.array(sigma_)

    # rho
    rho = [0] * Rh
    rho_ = [0] * Rh
    for r in range(Rh):
        rho_guide = pyro.param(f"rho_guide_rh{r}")
        rho[r] = dist.Dirichlet(rho_guide).independent(1).mean
        # rho[r] = pyro.param(f"rho_rh{r}")
        rho_[r] = rho[r].cpu().detach().numpy()

    if(pathResultFolder is not None):
        print("saving models")

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}

        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["theta_guide"] = theta_guide_
        variables["theta"] = theta_
        variables["alpha_hyper"] = alpha_hyper_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        args_dict_str = args_dict.copy()
        for k in args_dict_str:
            if(not isinstance(args_dict_str[k], (int, float, complex, bool))):
                args_dict_str[k] = str(args_dict_str[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict_str.values()), axis="row", names=list(args_dict_str.keys()))

        ws = wb.create_sheet("mu_sigma")
        writeMatrix(ws, mu_, 1, 1,
                    row_names=[f"record_m{r+1}" for r in range(Rm)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)
        writeMatrix(ws, sigma_, 1, K + 3,
                    row_names=[f"record_m{r+1}" for r in range(Rm)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("rho")
        row = 1
        for r in range(Rh):
            writeMatrix(ws, rho_[r].T, row, 1,
                        row_names=[f"category{c+1}" for c in range(n_h[r])],
                        column_names=[f"topic{k+1}" for k in range(K)],
                        addDataBar=True)
            _, counts = torch.unique(habits[r], return_counts=True)
            counts = counts.cpu().detach().numpy()
            counts = counts / np.sum(counts)
            writeMatrix(ws, counts[:, None], row, K + 3, column_names=["data distribution"], addDataBar=True)
            row += n_h[r] + 2

        ws = wb.create_sheet("alpha_hyper")
        writeVector(ws, alpha_hyper_, axis="row", names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))
