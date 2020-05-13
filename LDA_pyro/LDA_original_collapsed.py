import pickle
import numpy as np
import openpyxl
from utils.openpyxl_util import writeMatrix, writeVector, writeSortedMatrix

import torch
from torch.distributions import constraints
import pyro
import pyro.poutine as poutine
from pyro import distributions as dist


def model(data, args):
    """
    data:
        [tensor[totalN] (word id),
         tensor[totalN] (doc id)]
    args:
        K,
        auto_beta, auto_alpha
        coef_beta, coef_alpha
        device
    """
    D = args.D
    V = args.V
    totalN = len(data[0])
    K = args.K

    if(args.auto_beta):
        beta_hyper = pyro.param("beta_hyper", torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta, constraint=constraints.positive)
    else:
        beta_hyper = torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta
    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper", torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha, constraint=constraints.positive)
    else:
        alpha_hyper = torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha

    # phi, theta = calcPhiAndTheta(data, D, V, K, alpha_hyper, beta_hyper, args.device)
    # theta, phi 生成
    with pyro.plate("topics", K):
        phi = pyro.sample("phi", dist.Dirichlet(beta_hyper))
    with pyro.plate("documents", D):
        theta = pyro.sample("theta", dist.Dirichlet(alpha_hyper))

    with pyro.plate("words", totalN, subsample_size=10000) as n:
        wordID = data[0][n].long()
        docID = data[1][n].long()
        z = pyro.sample("z", dist.Categorical(theta[docID]))
        w = pyro.sample("w", dist.Categorical(phi[z]), obs=wordID)

    return phi, theta, z, w


def guide(data, args):
    D = args.D
    V = args.V
    totalN = len(data[0])
    K = args.K

    if(args.auto_beta):
        beta_hyper = pyro.param("beta_hyper", torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta, constraint=constraints.positive)
    else:
        beta_hyper = torch.ones([V], device=args.device, dtype=torch.float64) * args.coef_beta
    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper", torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha, constraint=constraints.positive)
    else:
        alpha_hyper = torch.ones([K], device=args.device, dtype=torch.float64) * args.coef_alpha

    # theta, phi 近似分布を使わずに生成
    with pyro.plate("topics", K):
        phi = pyro.sample("phi", dist.Dirichlet(beta_hyper))
    with pyro.plate("documents", D):
        theta = pyro.sample("theta", dist.Dirichlet(alpha_hyper))

    q_model = pyro.param("q_model", torch.ones([totalN, K], device=args.device, dtype=torch.float32) / K, constraint=constraints.simplex)
    with pyro.plate("words", totalN, subsample_size=10000) as n:
        z = pyro.sample("z", dist.Categorical(q_model[n]))


def calcPhiAndTheta(data, D, V, K, alpha_hyper, beta_hyper, device):
    q_model = pyro.param("q_model")
    totalN = len(data[0])

    phi = torch.zeros((K, V), device=device)
    theta = torch.zeros((D, K), device=device)

    for v in range(V):
        phi[:, v] = torch.sum(q_model[data[0] == v, :], 0) + beta_hyper[v]
    for d in range(D):
        theta[d, :] = torch.sum(q_model[data[1] == d, :], 0) + alpha_hyper

    phi = phi / torch.sum(phi, 1, keepdim=True)
    theta = theta / torch.sum(theta, 1, keepdim=True)

    return phi, theta


def summary(data, args, words, reviews, pathResultFolder=None, counts=None):

    D = args.D
    V = args.V
    K = args.K

    if(args.auto_beta):
        beta_hyper = pyro.param("beta_hyper")
        beta_hyper_ = beta_hyper.cpu().detach().numpy()
    else:
        beta_hyper_ = np.ones(V) * args.coef_beta
    if(args.auto_alpha):
        alpha_hyper = pyro.param("alpha_hyper")
        alpha_hyper_ = alpha_hyper.cpu().detach().numpy()
    else:
        alpha_hyper_ = np.ones(K) * args.coef_alpha

    phi, theta = calcPhiAndTheta(data, D, V, K, alpha_hyper, beta_hyper, args.device)
    phi_ = phi.cpu().detach().numpy()
    theta_ = theta.cpu().detach().numpy()

    # print summary
    for k in range(K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{} ".format(words[ind[i]])
        print(msg)
    for k in range(K):
        msg = "Topic {:2d}: ".format(k)
        ind = np.argsort(phi_[k, :])[::-1]
        for i in range(10):
            msg += "{:6f} ".format(phi_[k, ind[i]])
        print(msg)
    for d in range(3):
        msg = "Documents {:2d}: ".format(d)
        for k in range(K):
            msg += "{:6f} ".format(theta_[d, k])
        print(msg)

    if(pathResultFolder is not None):
        print("saving models")

        # model_variables.pickle
        variables = {}
        args_dict = {k: args.__dict__[k] for k in args.__dict__ if not k.startswith("__")}

        variables["args"] = args_dict
        variables["words"] = words
        variables["reviews"] = reviews
        variables["phi"] = phi_
        variables["theta"] = theta_
        variables["beta_hyper"] = beta_hyper_
        variables["alpha_hyper"] = alpha_hyper_
        with open(pathResultFolder.joinpath("model_variables.pickle"), 'wb') as f:
            pickle.dump(variables, f)

        # result.xlsx
        wb = openpyxl.Workbook()
        tmp_ws = wb[wb.get_sheet_names()[0]]

        args_dict_str = args_dict.copy()
        for k in args_dict_str:
            if(not isinstance(args_dict_str[k], (int, float, complex, bool))):
                args_dict_str[k] = str(args_dict_str[k])
        ws = wb.create_sheet("args")
        writeVector(ws, list(args_dict_str.values()), axis="row", names=list(args_dict_str.keys()))

        ws = wb.create_sheet("theta")
        writeMatrix(ws, theta_, 1, 1,
                    row_names=[f"doc{d+1}" for d in range(D)],
                    column_names=[f"topic{k+1}" for k in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("phi")
        writeMatrix(ws, phi_.T, 1, 1,
                    row_names=words,
                    column_names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("beta_hyper")
        writeVector(ws, beta_hyper_, axis="row", names=words,
                    addDataBar=True)

        ws = wb.create_sheet("alpha_hyper")
        writeVector(ws, alpha_hyper_, axis="row", names=[f"topic{d+1}" for d in range(K)],
                    addDataBar=True)

        ws = wb.create_sheet("phi_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=words, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher")

        ws = wb.create_sheet("phi_value_sorted")
        writeSortedMatrix(ws, phi_.T, axis=0, row=1, column=1,
                          row_names=None, column_names=[f"topic{d+1}" for d in range(K)],
                          maxwrite=100, order="higher",
                          addDataBar=True)

        # if(counts is not None):
        #     phi2_ = phi_ / counts[None, :]
        #     ws = wb.create_sheet("phi(per num)_sorted")
        #     writeSortedMatrix(ws, phi2_.T, axis=0, row=1, column=1,
        #                       row_names=words, column_names=[f"topic{d+1}" for d in range(K)],
        #                       maxwrite=100, order="higher",)

        #     ws = wb.create_sheet("phi(per num)_value_sorted")
        #     writeSortedMatrix(ws, phi2_.T, axis=0, row=1, column=1,
        #                       row_names=None, column_names=[f"topic{d+1}" for d in range(K)],
        #                       maxwrite=100, order="higher",)

        wb.remove_sheet(tmp_ws)
        wb.save(pathResultFolder.joinpath("result.xlsx"))
