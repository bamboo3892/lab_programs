import numpy as np
import openpyxl
from openpyxl.styles import Font, Color, Border, Side
from openpyxl.formatting.rule import ColorScale, DataBar, FormatObject, Rule, DataBarRule, ColorScaleRule
from openpyxl.utils.cell import get_column_letter


FULL_BORDER = Border(
    left=Side(
        border_style="thin",
        color="000000"
    ),
    right=Side(
        border_style="thin",
        color="000000"
    ),
    top=Side(
        border_style="thin",
        color="000000"
    ),
    bottom=Side(
        border_style="thin",
        color="000000"

    )
)


def writeVector(ws, vector, row=1, column=1, axis="column", names=None,
                matrix_font=None,
                name_font=Font(color="006400"),
                rule=None, ruleBoundary=None):
    if(axis == "column"):
        writeMatrix(ws, [vector], row=row, column=column, column_names=names,
                    matrix_font=matrix_font, column_name_font=name_font,
                    rule=rule, ruleBoundary=ruleBoundary)
    elif(axis == "row"):
        writeMatrix(ws, [[a] for a in vector], row=row, column=column, row_names=names,
                    matrix_font=matrix_font, row_name_font=name_font,
                    rule=rule, ruleBoundary=ruleBoundary)
    else:
        raise Exception("irregal parameter")


def writeMatrix(ws, matrix, row=1, column=1, row_names=None, column_names=None,
                matrix_font=None,
                row_name_font=Font(color="006400"),
                column_name_font=Font(color="006400"),
                rule=None, ruleBoundary=None, ruleAxis=None, rule_color=None):

    offset_row = row if column_names is None else row + 1
    offset_column = column if row_names is None else column + 1

    if(row_names is not None):
        for i in range(len(row_names)):
            ws.cell(offset_row + i, column, row_names[i]).font = row_name_font

    if(column_names is not None):
        for i in range(len(column_names)):
            ws.cell(row, offset_column + i, column_names[i]).font = column_name_font

    maxVecLength = 0
    for i in range(len(matrix)):
        vec = matrix[i]
        maxVecLength = len(vec) if len(vec) > maxVecLength else maxVecLength
        for j in range(len(vec)):
            ws.cell(offset_row + i, offset_column + j, vec[j]).font = matrix_font

    if(rule == "databar"):
        addDataBarRules(ws, offset_row, offset_row + len(matrix), offset_column, offset_column + maxVecLength,
                        color=rule_color, boundary=ruleBoundary, axis=ruleAxis)
    elif(rule == "colorscale"):
        addColorScaleRules(ws, offset_row, offset_row + len(matrix), offset_column, offset_column + maxVecLength,
                           colors=rule_color, boundary=ruleBoundary, axis=ruleAxis)


def writeSortedMatrix(ws, matrix, axis=0, row=1, column=1, row_names=None, column_names=None,
                      maxwrite=None, order="higher",
                      matrix_font=None,
                      name_font=Font(color="006400"),
                      rule=None, ruleBoundary=None, ruleAxis=None):
    """
    names is none: write sorted matrix values
    names is not none: write sorted names
    """

    idx = np.argsort(matrix, axis)
    if(order == "higher"):
        if(axis == 0):
            idx = idx[::-1, :]
        elif(axis == 1):
            idx = idx[:, ::-1]

    if(maxwrite is not None):
        if(axis == 0):
            idx = idx[:min([idx.shape[0], maxwrite]), :]
        elif(axis == 1):
            idx = idx[:, :min([idx.shape[1], maxwrite])]

    if(axis == 0):
        if(row_names is None):
            data = np.zeros_like(idx, dtype=matrix.dtype)
            for i in range(data.shape[1]):
                data[:, i] = matrix[idx[:, i], i]
        else:
            data = np.array(row_names)[idx].tolist()
    elif(axis == 1):
        if(column_names is None):
            data = np.zeros_like(idx, dtype=matrix.dtype)
            for i in range(data.shape[0]):
                data[i, :] = matrix[i, idx[i, :]]
        else:
            data = np.array(column_names)[idx].tolist()

    if(axis == 0):
        writeMatrix(ws, data, row=row, column=column, column_names=column_names,
                    matrix_font=matrix_font,
                    row_name_font=name_font, rule=rule, ruleBoundary=ruleBoundary, ruleAxis=ruleAxis)
    elif(axis == 1):
        writeMatrix(ws, data, row=row, column=column, row_names=row_names,
                    matrix_font=matrix_font,
                    row_name_font=name_font, rule=rule, ruleBoundary=ruleBoundary, ruleAxis=ruleAxis)


def addDataBarRules(ws, row1, row2, column1, column2, *,
                    color=None, boundary=None, axis=None):
    if(row1 > row2):
        a = row1
        row1 = row2
        row2 = a
    if(column1 > column2):
        a = column1
        column1 = column2
        column2 = a
    color = color if color is not None else "00bfff"

    if(axis == "column"):
        for row in range(row1, row2 + 1):
            addDataBarRules(ws, row, row, column1, column2, color=color, boundary=boundary, axis=None)
    elif(axis == "row"):
        for column in range(column1, column2 + 1):
            addDataBarRules(ws, row1, row2, column, column, color=color, boundary=boundary, axis=None)

    area = getAreaLatter(row1, row2, column1, column2)
    if(boundary is None):
        rule = DataBarRule(start_type="min", start_value=None, end_type="max", end_value=None, color=color, showValue=None, minLength=0, maxLength=100)
    else:
        rule = DataBarRule(start_type="num", start_value=boundary[0], end_type="num", end_value=boundary[1], color=color, showValue=None, minLength=0, maxLength=100)
    ws.conditional_formatting.add(area, rule)


def addColorScaleRules(ws, row1, row2, column1, column2, *,
                       colors=[Color('FFFFFF'), Color('008000')], boundary=None, axis=None):
    """
    colors: list of colors. [2] or [3]
    boundary: list of values. the same shape as colors
    axis: axis to apply rule. 0 or 1
    """
    if(row1 > row2):
        a = row1
        row1 = row2
        row2 = a
    if(column1 > column2):
        a = column1
        column1 = column2
        column2 = a
    colors = colors if colors is not None else [Color('FFFFFF'), Color('008000')]

    if(axis == "column"):
        for row in range(row1, row2 + 1):
            addColorScaleRules(ws, row, row, column1, column2, colors=colors, boundary=boundary, axis=None)
    elif(axis == "row"):
        for column in range(column1, column2 + 1):
            addColorScaleRules(ws, row1, row2, column, column, colors=colors, boundary=boundary, axis=None)

    # make rule
    if(boundary is None):
        if(len(colors) == 2):
            rule = ColorScaleRule(start_type="min", start_value=None, start_color=colors[0],
                                  end_type="max", end_value=None, end_color=colors[1])
        else:
            rule = ColorScaleRule(start_type="min", start_value=None, start_color=colors[0],
                                  mid_type="percent", mid_value=50, mid_color=colors[1],
                                  end_type="max", end_value=None, end_color=colors[2])
    else:
        if(len(colors) == 2):
            rule = ColorScaleRule(start_type="num", start_value=boundary[0], start_color=colors[0],
                                  end_type="num", end_value=boundary[1], end_color=colors[1])
        elif(len(colors) == 3):
            rule = ColorScaleRule(start_type="num", start_value=boundary[0], start_color=colors[0],
                                  mid_type="num", mid_value=boundary[1], mid_color=colors[1],
                                  end_type="num", end_value=boundary[2], end_color=colors[2])

    # apply rule
    area = getAreaLatter(row1, row2, column1, column2)
    ws.conditional_formatting.add(area, rule)


def addBorderToMaxCell(ws, row1, row2, column1, column2, *,
                       border=FULL_BORDER, order="max", n=1, axis=None):
    def func(cell):
        cell.border = border
    doFuncToMaxCell(ws, row1, row2, column1, column2, func, order=order, n=n, axis=axis)


def doFuncToMaxCell(ws, row1, row2, column1, column2, func, *,
                    order="max", n=1, axis=None):

    if(row1 > row2):
        a = row1
        row1 = row2
        row2 = a
    if(column1 > column2):
        a = column1
        column1 = column2
        column2 = a

    if(axis == "column"):
        for row in range(row1, row2 + 1):
            doFuncToMaxCell(ws, row, row, column1, column2, func, order=order, n=n, axis=None)
    elif(axis == "row"):
        for column in range(column1, column2 + 1):
            doFuncToMaxCell(ws, row1, row2, column, column, func, order=order, n=n, axis=None)

    v = float("inf") if order == "min" else -float("inf")
    row = []
    column = []
    for r in range(row1, row2 + 1):
        for c in range(column1, column2 + 1):
            if(isinstance(ws.cell(r, c).value, (int, float))):
                if(ws.cell(r, c).value == v):
                    row.append(r)
                    column.append(c)
                elif((order == "min" and ws.cell(r, c).value < v) or ws.cell(r, c).value > v):
                    v = ws.cell(r, c).value
                    row = [r]
                    column = [c]
    for i in range(len(row)):
        func(ws.cell(row[i], column[i]))


def paintCells(ws, row, column, height, width, color="006400"):
    fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=color)
    for r in range(row, row + height):
        for c in range(column, column + width):
            ws.cell(r, c).fill = fill


def writePaintedText(ws, row, text, paintWidth=20, textFont=Font(color="ffffff"), paintColor="006400"):
    paintCells(ws, row, 1, 1, paintWidth, color=paintColor)
    ws.cell(row, 1, value=text).font = textFont


def getAreaLatter(row1, row2, column1, column2):
    return get_column_letter(column1) + str(row1) + ":" + get_column_letter(column2) + str(row2)


def drawImage(ws, imagePath, height, width, anker):
    """
    draw image in worksheet
    imagePath: path to the image. Do not delete image until worksheet is saved.
    height: height
    width: width
    anker: The position of the cell to be anchored (string e.g. "A1")
    """
    img = openpyxl.drawing.image.Image(str(imagePath))
    img.height = height
    img.width = width
    ws.add_image(img, anker)
