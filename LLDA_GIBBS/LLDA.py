# -*- coding: utf-8 -*-
"""
Created on NOV 21 2018\n
LLDA(simplified) with Collapsed Gibbs sampling + fixed-point iteration\n
@author: takeda masaki
"""

import numpy as np
import random
import csv
from scipy.special import digamma
import sys
import math
import openpyxl
from openpyxl.styles import Font

from LDA_GIBBS import LDA
from LLDA_GIBBS import LLDA_util


class LLDA_S(LDA.LDA):
    """
    LLDA class
    K(number of topics) will be set the number of unique label
    alphaFactor: alpha is initial parameter of theta
                 alphaFactor is the values that determine each document's alpha
                 depending on whether the doc's label contains the topic.
                 When the doc's label does not have the topic, the corresponing alpha will be set alphaFactor[0].
                 When the doc's label has the topic, the corresponing alpha will be set alphaFactor[1].
                 When the topic dones not have corresponding label, the corresponing alpha will be set alphaFactor[2].
                 [0.05, 0.1, 0.05]
    beta       : initial parameter of phi
                 same value between topics
    docs       : documents
                 [
                     [
                         [
                            [word], [word], ...
                         ], [sentence], [sentence], ...
                     ], [document], [document], ...
                 ]
    labels     : label list
    K          : number of topics
    scoreCalcRate: the rate score will be calclated from training data
                   Set negative to ban calclating score during learning
    """

    def __init__(self, alphaFactor, beta, docs, labels, *,
                 K=None, labelTypes=[], randomInit=False, scoreCalcRate=10, testDocs=None, testLabels=None,
                 w2vModel=None, coherenceLabel=None):

        self.alphaFactor = alphaFactor

        # init label
        self.labels = []
        self.testLabels = []
        self.labelTypes = labelTypes
        for docLabels in labels:
            a0 = []
            for label in docLabels:
                if(label not in self.labelTypes):
                    self.labelTypes.append(label)
                for k, labelType in enumerate(self.labelTypes):
                    if(label == labelType and k not in a0):
                        a0.append(k)
                # a0.append(self.labelTypes.index(label))
            self.labels.append(a0)
        if((testDocs is not None) and (testLabels is not None)):
            for docLabels in testLabels:
                a0 = []
                for label in docLabels:
                    if(label in self.labelTypes):
                        # a0.append(self.labelTypes.index(label))
                        for k, labelType in enumerate(self.labelTypes):
                            if(label == labelType and k not in a0):
                                a0.append(k)
                self.testLabels.append(a0)
        self.coherenceLabel = coherenceLabel

        if(K is None):
            K = len(self.labelTypes)

        super().__init__(K, alphaFactor[0], beta, docs, arrangeInSentece=False, randomInit=randomInit,
                         scoreCalcRate=scoreCalcRate, testDocs=testDocs, w2vModel=w2vModel)

    def sampleInitZ(self, randomInit):

        # disable alpha, init alpha2
        self.alpha = "Disabled in LLDA"
        a = np.full(self.K, self.alphaFactor[0])
        a[len(self.labelTypes):] = self.alphaFactor[2]
        self.alpha2 = np.resize(a, (self.D, self.K))
        for d, docLabels in enumerate(self.labels):
            for n, labelType in enumerate(self.labelTypes):
                if(n in self.labels[d]):
                    self.alpha2[d, n] = self.alphaFactor[1]

        # sample initial z
        if(not randomInit):
            for d, document in enumerate(self.documents):
                for n, v in enumerate(document):
                    p0 = self.nWord_dk[d] + self.alpha2[d]
                    p1 = self.nWord_kv[:, v] + self.beta[v]
                    p2 = self.nWord_k + self.beta.sum()
                    p = p0 * p1 / p2
                    p = p / p.sum()
                    z0 = np.random.multinomial(1, p).argmax()
                    self.z[d][n] = z0
                    self.nWord_k[z0] += 1
                    self.nWord_dk[d, z0] += 1
                    self.nWord_kv[z0, v] += 1
        else:
            for d, document in enumerate(self.documents):
                for n, v in enumerate(document):
                    z0 = np.random.randint(0, self.K)
                    self.z[d][n] = z0
                    self.nWord_k[z0] += 1
                    self.nWord_dk[d, z0] += 1
                    self.nWord_kv[z0, v] += 1

    def learn(self):
        """
        Execute single learning loop\n
        Make sure to use #calcThetaAndPhi after learning loop
        """

        """ 1. sampleing z (Collapsed Gibbs sampling) (bottleneck)"""
        LLDA_util.updateZ(self.documents,
                          self.alpha2, self.beta,
                          self.nWord_dk, self.nWord_kv, self.nWord_k,
                          self.z)

        self.countLearned += 1

    def updateAlphaAndBeta(self):
        b0 = digamma(self.nWord_kv + np.resize(self.beta, (self.K, self.V))).sum(axis=0)
        b0 -= self.K * digamma(self.beta)
        b1 = digamma(self.nWord_k + np.full(self.K, self.beta.sum())).sum()
        b1 -= self.K * digamma(self.beta.sum())
        self.beta *= b0 / b1

    def calcThetaAndPhi(self):
        t0 = self.nWord_dk + self.alphaFactor[0]
        t1 = np.resize(self.nWord_d + self.alphaFactor[0] * self.K, (self.K, self.D)).T
        self.theta = t0 / t1
        p0 = self.nWord_kv + np.resize(self.beta, (self.K, self.V))
        p1 = np.resize(self.nWord_k + self.beta.sum(), (self.V, self.K)).T
        self.phi = p0 / p1

        self.theta2 = np.zeros((self.D, self.K))
        self.theta3 = np.zeros((self.D, self.K))
        phi_ = self.phi / self.phi.sum(axis=0, keepdims=True)
        for d, document in enumerate(self.documents):
            for n, v in enumerate(document):
                self.theta2[d] += phi_[:, v]
                self.theta3[d] += self.phi[:, v]

    def addHistory(self):
        super().addHistory()

    def calcPerplexity2(self, testdocs, testRatio=0.2):

        self.calcThetaAndPhi()

        # init trainset
        trainset = []
        for document in testdocs:
            l0 = []
            if(not self.arrangeInSentece):
                for sentence in document:
                    for word in sentence:
                        if(word in self.words):
                            l0.append(self.words.index(word))
                trainset.append(l0)
            else:
                for sentence in document:
                    l0 = []
                    for word in sentence:
                        if(word in self.words):
                            l0.append(self.words.index(word))
                    trainset.append(l0)
        testD = len(trainset)

        # init testset by spliting trainset
        testset = []
        for document in trainset:
            l1 = []
            i0 = int(len(document) * testRatio)
            if(i0 > 0):
                for i1 in range(i0):
                    l1.append(document.pop(random.randrange(len(document))))
            testset.append(l1)

        nWord_d = []
        for document in trainset:
            nWord_d.append(len(document))
        nWord_dk = np.zeros((testD, self.K))
        phi_ = self.phi / self.phi.sum(axis=0, keepdims=True)
        for d, document in enumerate(trainset):
            for n, v in enumerate(document):
                nWord_dk[d] += phi_[:, v]

        # calc perplexity
        p = 0
        nTestWord = 0
        for d, document in enumerate(testset):
            nTestWord += len(document)
            for n, v in enumerate(document):
                p0 = nWord_dk[d] + self.alphaFactor[0]
                p1 = nWord_d[d] + self.alphaFactor[0] * self.K
                p2 = self.nWord_kv[:, v] + self.beta[v]
                p3 = self.nWord_k + self.beta.sum()
                p += np.log((p0 / p1 * p2 / p3).sum())
        perplexity = math.exp(-1 * p / nTestWord)

        return perplexity

    def getTopicName(self, k):
        if(k < len(self.labelTypes)):
            return self.labelTypes[k]
        else:
            return super().getTopicName(k)

    def writeModelToFolder(self, pathResultFolder):
        """
        Make files about model under param folder\n
        created files:
            model.pickle    : pickle of this model
            theta.csv       : theta
            phi.csv         : phi and historys
            phi.xlsx        : phi
            z.csv           : z
            alpha-beta.csv  : alpha and beta
            words.xlsx      : visualized z
        """

        super().writeModelToFolder(pathResultFolder)

        pathResultFolder.joinpath("{}".format(self.alphaFactor)).mkdir(parents=True, exist_ok=True)

        coherences = []
        coherences2 = []
        with open(str(pathResultFolder.joinpath("phi.csv")), mode='a', encoding="utf_8_sig") as f:
            writer = csv.writer(f, lineterminator='\n')
            writer.writerow("")
            writer.writerow(["label types"])
            writer.writerow(self.labelTypes)
            for k in range(self.K):
                coherences.append(self.calcCoherence(c=10, k=k, w2v=self.w2vModel))
            writer.writerow(coherences)
            if(self.coherenceLabel is not None):
                writer.writerow(["coherence", "", "mean"] + list(range(1, 6)))
                for k in range(len(self.coherenceLabel)):
                    # num = 0
                    # sum0 = 0
                    # l0 = [None] * 5
                    # argSorted = np.argsort(self.phi[k, :])[::-1]
                    # for i in range(5):
                    #     if(self.words[argSorted[i]] in self.w2vModel2.wv.vocab):
                    #         num += 1
                    #         sim = self.w2vModel2.wv.similarity(self.words[argSorted[i]], self.coherenceLabel[k])
                    #         sum0 += sim
                    #         l0[i] = sim
                    # coherences2.append(sum0 / num)
                    sum0 = 0
                    l0 = [None] * 5
                    argSorted = np.argsort(self.phi[k, :])[::-1]
                    for i in range(5):
                        sim = self.similarity[argSorted[i], self.words.index(self.coherenceLabel[k])]
                        sum0 += sim
                        l0[i] = sim
                    coherences2.append(sum0 / 5)
                    writer.writerow([coherences[k], self.coherenceLabel[k], coherences2[k]] + l0)
        with open(str(pathResultFolder.joinpath("alpha-beta.csv")), mode='a', encoding="utf_8_sig") as f:
            writer = csv.writer(f, lineterminator='\n')
            writer.writerow("")
            writer.writerow(["alpha2"])
            writer.writerows(self.alpha2)

        wb = openpyxl.Workbook()
        ws = wb.active
        fills = []
        colors1 = []
        sss = ["F0", "B0", "80", "40", "00"]
        for i1 in range(5):
            for i2 in range(5):
                for i3 in range(5):
                    colors1.append("88" + sss[i1] + sss[(i2 + 1) % 5] + sss[(i3 + 2) % 5])
        colors2 = []
        for i in range(125):
            colors2.append(colors1[(i * 49) % 125])
        for i in range(self.K):
            fills.append(openpyxl.styles.PatternFill(fill_type='solid', fgColor=colors2[i]))
        for k in range(len(self.labelTypes)):
            ws.cell(row=1, column=k + 1).value = self.labelTypes[k]
            ws.cell(row=1, column=k + 1).fill = fills[k]
        for d, document in enumerate(self.documents):
            for n, word in enumerate(document):
                ws.cell(row=d + 3, column=n + 1).value = self.words[word]
                ws.cell(row=d + 3, column=n + 1).fill = fills[self.z[d][n]]
                ws.cell(row=d + 3, column=n + 1).comment = openpyxl.comments.Comment(str(self.nWord_kv[:, word].tolist()), None)
            if(d > 100):
                break
        wb.save(str(pathResultFolder.joinpath("words.xlsx")))

        wb = openpyxl.load_workbook(str(pathResultFolder.joinpath("phi.xlsx")))
        ws = wb.get_sheet_by_name("phi")
        ws.cell(row=3, column=1, value="similarity(label)")
        if(self.labelTypes is not None):
            for k in range(len(self.labelTypes)):
                ws.cell(row=1, column=2 + k, value=self.labelTypes[k]).font = Font(color="dc143c")
                ws.cell(row=1, column=6 + self.K + k, value=self.labelTypes[k]).font = Font(color="dc143c")
        if(self.coherenceLabel is not None):
            for k in range(len(self.coherenceLabel)):
                ws.cell(row=3, column=2 + k, value=coherences2[k]).font = Font(color="dc143c")
                ws.cell(row=3, column=2 + k).comment = openpyxl.comments.Comment(self.coherenceLabel[k], None)
        wb.save(str(pathResultFolder.joinpath("phi.xlsx")))


class LLDA_weak(LDA.LDA):

    def __init__(self, alphaFactor, beta, docs, labels, *,
                 K=None, labelTypes=[], scoreCalcRate=10, testDocs=None, testLabels=None):

        self.alphaFactor = alphaFactor

        # init label
        self.labels = []
        self.testLabels = []
        self.labelTypes = labelTypes
        for docLabels in labels:
            a0 = []
            for label in docLabels:
                if(label not in self.labelTypes):
                    self.labelTypes.append(label)
                a0.append(self.labelTypes.index(label))
            self.labels.append(a0)
        if((testDocs is not None) and (testLabels is not None)):
            for docLabels in testLabels:
                a0 = []
                for label in docLabels:
                    if(label in self.labelTypes):
                        a0.append(self.labelTypes.index(label))
                self.testLabels.append(a0)

        if(K is None):
            K = len(self.labelTypes)

        super().__init__(K, alphaFactor[0], beta, docs, arrangeInSentece=False,
                         scoreCalcRate=scoreCalcRate, testDocs=testDocs)

    def sampleInitZ(self):

        # init alpha2
        self.alpha2 = np.full((self.D, self.K), self.alphaFactor[0])
        for d, docLabels in enumerate(self.labels):
            for label in docLabels:
                self.alpha2[d, label] = self.alphaFactor[1]

        # sample initial z
        for d, document in enumerate(self.documents):
            for n, v in enumerate(document):
                p0 = self.nWord_dk[d] + self.alpha2[d]
                p1 = self.nWord_kv[:, v] + self.beta[v]
                p2 = self.nWord_k + self.beta.sum()
                p = p0 * p1 / p2
                p = p / p.sum()
                z0 = np.random.multinomial(1, p).argmax()
                self.z[d][n] = z0
                self.nWord_k[z0] += 1
                self.nWord_dk[d, z0] += 1
                self.nWord_kv[z0, v] += 1

    def learn(self):
        if(self.countLearned < 10):
            LLDA_util.updateZ(self.documents,
                              self.alpha2, self.beta,
                              self.nWord_dk, self.nWord_kv, self.nWord_k,
                              self.z)
            self.countLearned += 1
        else:
            super().learn()

    def writeModelToFolder(self, pathResultFolder):
        super().writeModelToFolder(pathResultFolder)

        pathResultFolder.joinpath("{}".format(self.alphaFactor).replace(".", "")).mkdir(parents=True, exist_ok=True)

        with open(str(pathResultFolder.joinpath("phi.csv")), mode='a', encoding="utf_8_sig") as f:
            writer = csv.writer(f, lineterminator='\n')
            writer.writerow("")
            writer.writerow(["label types"])
            writer.writerow(self.labelTypes)
        with open(str(pathResultFolder.joinpath("alpha-beta.csv")), mode='a', encoding="utf_8_sig") as f:
            writer = csv.writer(f, lineterminator='\n')
            writer.writerow("")
            writer.writerow(["alpha2"])
            writer.writerows(self.alpha2)

        wb = openpyxl.Workbook()
        ws = wb.active
        fills = []
        colors1 = []
        sss = ["F0", "80", "08"]
        for i1 in range(3):
            for i2 in range(3):
                for i3 in range(3):
                    colors1.append("88" + sss[i1] + sss[(i2 + 1) % 3] + sss[(i3 + 2) % 3])
        colors2 = []
        for i in range(27):
            colors2.append(colors1[(i * 4) % 27])
        for i in range(self.K):
            fills.append(openpyxl.styles.PatternFill(fill_type='solid', fgColor=colors2[i]))
        for k in range(len(self.labelTypes)):
            ws.cell(row=1, column=k + 1).value = self.labelTypes[k]
            ws.cell(row=1, column=k + 1).fill = fills[k]
        for d, document in enumerate(self.documents):
            for n, word in enumerate(document):
                ws.cell(row=d + 3, column=n + 1).value = self.words[word]
                ws.cell(row=d + 3, column=n + 1).fill = fills[self.z[d][n]]
                ws.cell(row=d + 3, column=n + 1).comment = openpyxl.comments.Comment(str(self.nWord_kv[:, word].tolist()), None)
            if(d > 100):
                break
        wb.save(str(pathResultFolder.joinpath("words.xlsx")))
