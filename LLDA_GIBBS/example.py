# -*- coding: utf-8 -*-
"""
Created on NOV 21 2018\n
@author: takeda masaki
"""

import numpy as np
import json
import csv
import pickle
from joblib import Parallel, delayed
import openpyxl
from openpyxl.styles import Font

from LLDA_GIBBS import LLDA


def testLLDA(pathMorphomes, pathTestMorphomes, pathResultFolder, *,
             fromPickle=False, w2vModel=None):
    """ Example usage for LLDA class. """

    print("Start LLDA")

    # for golf
    def func0(path0, noLabelOk):
        # D1
        keys = ["rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy", "rating_fairway"]
        labalName = ["rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy", "rating_fairway"]
        # D2
        # keys = ["rating_meal", "rating_facility", "rating_length", "rating_staff", "rating_strategy", "rating_fairway"]
        # labalName = ["rating_meal", "rating_facility", "rating_length", "rating_staff", "rating_strategy", "rating_fairway"]
        # keys = ["rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy"]
        # labalName = ["rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy"]
        # D3
        # keys = ["rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy", "rating_fairway"]
        # labalName = ["rating_meal", "rating_facility", "rating_length", "rating_cf", "rating_staff", "rating_strategy", "rating_cf"]
        full = []
        docs = []
        lebels = []
        with open(str(path0), "r", encoding="utf_8_sig") as file0:
            reviews = json.load(file0)
            for review in reviews:
                label = []
                for i, key in enumerate(keys):
                    rating = review[key]
                    if(rating == 1 or rating == 2 or rating == 5):
                        label.append(labalName[i])
                if(len(label) == 0 and not noLabelOk):
                    continue
                full.append(review)
                docs.append(review["morphomes"])
                lebels.append(label)
        return full, docs, lebels

    # for sompo
    dict0 = {"1200": "モニタリング", "1210": "正しい病識", "1100": "適正受診", "1220": "適正服薬", "1230": "疾病の自己管理",
             "1300": "活動量",
             "1400": "飲料習慣", "1500": "間食習慣",
             "1720": "栄養バランス", "1730": "飽和脂肪酸の量", "1740": "主菜バランス", "1750": "野菜,海藻,きのこ類の量", "1760": "塩分量",
             "1600": "飲酒習慣",
             "1700": "夕食の時間と量", "1710": "食事リズム",
             "1800": "喫煙",
             "1900": "疲労回復,ストレス解消",
             #  "2100": "（就寝前の）歯磨き", "2200": "フッ素入り歯磨き剤", "2300": "歯間ブラシ・フロス", "2400": "ゆっくり噛む・食べる", "2500": "歯磨き指導", "2600": "定期健診（歯科）",
             "2000": "その他"
             }
    dict1 = {"1200": "疾病管理", "1210": "疾病管理", "1100": "疾病管理", "1220": "疾病管理", "1230": "疾病管理",
             "1300": "運動",
             "1400": "食材・バランス", "1500": "食材・バランス",
             "1720": "食材・バランス", "1730": "食材・バランス", "1740": "食材・バランス", "1750": "食材・バランス", "1760": "食材・バランス",
             "1600": "飲酒",
             "1700": "生活リズム", "1710": "生活リズム",
             "1800": "タバコ",
             #  "1900": "ストレス",
             #  "2100": "歯科", "2200": "歯科", "2300": "歯科", "2400": "歯科", "2500": "歯科", "2600": "歯科",
             #  "2000": "その他"
             }
    dict2 = {"1100": "適正受診",
             "1200": "モニタリング",
             "1300": "運動",
             "1400": "飲料",
             "1500": "間食",
             "1600": "飲酒",
             "1700": "生活リズム", "1710": "生活リズム",
             #  "1720": "食材・バランス", "1730": "食材・バランス", "1740": "食材・バランス", "1750": "食材・バランス", "1760": "食材・バランス",
             #  "1800": "タバコ",
             #  "1900": "ストレス",
             #  "2100": "歯科", "2200": "歯科", "2300": "歯科", "2400": "歯科", "2500": "歯科", "2600": "歯科",
             #  "2000": "その他"
             }

    def func1(path0, noLabelOk):
        full = []
        docs = []
        lebels = []
        dic = dict0
        with open(str(path0), "r", encoding="utf_8_sig") as file0:
            reviews = json.load(file0)
            for review in reviews:
                label = []
                if("p_r_tgtset_explan_seqs_id" in review.keys()):
                    l0 = review["p_r_tgtset_explan_seqs_id"].split(",")
                    for l1 in l0:
                        if(l1 in dic.keys()):
                            label.append(dic[l1])
                if(len(label) == 0 and not noLabelOk):
                    continue
                full.append(review)
                docs.append(review["morphomes"])
                lebels.append(label)
        return full, docs, lebels

    # read file
    # full, docs, lebels = func0(pathMorphomes, False)
    # testFull, testDocs, testLabels = func0(pathTestMorphomes, False)
    full, docs, lebels = func1(pathMorphomes, True)
    testFull, testDocs, testLabels = func1(pathTestMorphomes, True)

    if(not fromPickle):
        # start learning

        # D1
        # model = LLDA.LLDA_S([0.0, 0.1, 0.1], 0.01, docs, lebels, K=8,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel,
        #                     # labelTypes=["rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy", "rating_fairway"],
        #                     labelTypes=["rating_meal", "rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy", "rating_fairway"],
        #                     randomInit=True,
        #                     # coherenceLabel=["食事", "設備", "距離", "価格", "スタッフ", "戦略", "Fairway"]
        #                     )

        # D2
        # model = LLDA.LLDA_S([0.0, 0.1, 0.05], 0.01, docs, lebels, K=7,
        #                     labelTypes=["rating_meal", "rating_facility", "rating_length", "rating_staff", "rating_strategy", "rating_fairway"],
        #                     # labelTypes=["rating_meal", "rating_facility", "rating_length", "rating_cost", "rating_staff", "rating_strategy"],
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel,
        #                     randomInit=False)

        # D3
        # model = LLDA.LLDA_S([0.0, 0.1, 0.04], 0.01, docs, lebels, K=7,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel,
        #                     # labelTypes=["rating_meal", "rating_facility", "rating_length", "rating_staff", "rating_strategy", "rating_cf", "rating_cf"],
        #                     labelTypes=["rating_meal", "rating_facility", "rating_length", "rating_staff", "rating_strategy", "rating_cf"],
        #                     randomInit=False)

        # sompo17
        # model = LLDA.LLDA_S([0.0, 1.0, 0.5], 0.01, docs, lebels, K=30,
        #                     labelTypes=["モニタリング", "正しい病識", "適正受診", "適正服薬", "疾病の自己管理", "活動量", "飲料習慣", "間食習慣",
        #                                 "栄養バランス", "飽和脂肪酸の量", "主菜バランス", "野菜,海藻,きのこ類の量", "塩分量",
        #                                 "飲酒習慣", "夕食の時間と量", "食事リズム", "喫煙"],
        #                     randomInit=False,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)

        # sompo18
        # model = LLDA.LLDA_S([0.0, 1.0, 0.5], 0.01, docs, lebels, K=19,
        #                     labelTypes=["モニタリング", "正しい病識", "適正受診", "適正服薬", "疾病の自己管理", "活動量", "飲料習慣", "間食習慣",
        #                                 "栄養バランス", "飽和脂肪酸の量", "主菜バランス", "野菜,海藻,きのこ類の量", "塩分量",
        #                                 "飲酒習慣", "夕食の時間と量", "食事リズム", "喫煙", "疲労回復,ストレス解消"],
        #                     randomInit=False,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)
        # model = LLDA.LLDA_S([0.0, 1.0, 0.0], 0.01, docs, lebels, K=19,
        #                     labelTypes=["モニタリング", "正しい病識", "適正受診", "適正服薬", "疾病の自己管理", "活動量", "飲料習慣", "飲料習慣", "間食習慣",
        #                                 "栄養バランス", "飽和脂肪酸の量", "主菜バランス", "野菜,海藻,きのこ類の量", "塩分量",
        #                                 "飲酒習慣", "夕食の時間と量", "食事リズム", "喫煙", "疲労回復,ストレス解消"],
        #                     randomInit=False,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)
        # model = LLDA.LLDA_S([0.0, 1.0, 0.0], 0.01, docs, lebels, K=19,
        #                     labelTypes=["モニタリング", "正しい病識", "適正受診", "適正服薬", "疾病の自己管理", "活動量", "飲料習慣", "間食習慣",
        #                                 "栄養バランス", "飽和脂肪酸の量", "主菜バランス", "主菜バランス", "野菜,海藻,きのこ類の量", "塩分量",
        #                                 "飲酒習慣", "夕食の時間と量", "食事リズム", "喫煙", "疲労回復,ストレス解消"],
        #                     randomInit=False,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)

        # sompo19
        # model = LLDA.LLDA_S([0.0, 1.0, 0.5], 0.01, docs, lebels, K=30,
        #                     labelTypes=["モニタリング", "正しい病識", "適正受診", "適正服薬", "疾病の自己管理", "活動量", "飲料習慣", "間食習慣",
        #                                 "栄養バランス", "飽和脂肪酸の量", "主菜バランス", "野菜,海藻,きのこ類の量", "塩分量",
        #                                 "飲酒習慣", "夕食の時間と量", "食事リズム", "喫煙", "疲労回復,ストレス解消", "その他"],
        #                     randomInit=False,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)

        # sompo6
        # model = LLDA.LLDA_S([0.0, 0.1, 0.0], 0.01, docs, lebels, K=6,
        #                     labelTypes=["疾病管理", "運動", "食材・バランス", "飲酒", "生活リズム", "タバコ"],
        #                     # labelTypes=["疾病管理", "運動", "食材・バランス", "食材・バランス", "飲酒", "生活リズム", "タバコ"],
        #                     randomInit=False,
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)
        # model = LLDA.LLDA_S([0.0, 0.1, 0.1], 0.01, docs, lebels, K=7,
        #                     labelTypes=["適正受診", "モニタリング", "運動", "生活リズム", "飲料", "飲酒", "間食"],
        #                     testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)

        # model = LLDA.LLDA_weak([0.01, 0.1], 0.01, docs, lebels, K=None,
        #                        labelTypes=["適正受診", "モニタリング", "運動", "生活リズム", "飲料", "飲酒", "間食"],
        #                        testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)

        # sompo message
        model = LLDA.LLDA_S([0.0, 1.0, 0.5], 0.01, docs, lebels, K=30,
                            labelTypes=["モニタリング", "正しい病識", "適正受診", "適正服薬", "疾病の自己管理", "活動量", "飲料習慣", "間食習慣",
                                        "栄養バランス", "飽和脂肪酸の量", "主菜バランス", "野菜,海藻,きのこ類の量", "塩分量",
                                        "飲酒習慣", "夕食の時間と量", "食事リズム", "喫煙", "疲労回復,ストレス解消", "その他"],
                            randomInit=True,
                            testDocs=testDocs, testLabels=testLabels, w2vModel=w2vModel)

        model.startLearning(50)
    else:
        with open(str(pathResultFolder.joinpath("model.pickle")), mode='rb') as f:
            model = pickle.load(f)
            model.w2vModel = w2vModel

    # # write result
    LLDA.LLDA_S.writeModelToFolder(model, pathResultFolder)
    for d, doc in enumerate(full):
        doc["theta_d"] = model.theta[d].tolist()
        doc["nWord_dk"] = model.nWord_dk[d].tolist()
        doc["theta2_d"] = model.theta2[d].tolist()
        doc["theta3_d"] = model.theta3[d].tolist()
    with open(str(pathResultFolder.joinpath("documents.json")), "w", encoding="utf_8_sig") as file1:
        text = json.dumps(full, ensure_ascii=False)
        text = text.replace("},", "},\n")
        file1.write(text)

    # with open(str(pathResultFolder.joinpath("topic-sentence.csv")), "w", encoding="utf_8_sig") as f5:
    #     writer = csv.writer(f5, lineterminator='\n')
    #     n = 5
    #     writer.writerow(["{} sentences with highest each topic ratio".format(n)])
    #     writer.writerow([""])
    #     for k in range(model.K):
    #         writer.writerow(["###############################################################"])
    #         writer.writerow([model.getTopicName(k)])
    #         p0 = model.theta2[:, k]
    #         arg = np.argsort(p0)[::-1]
    #         for i in range(n):
    #             writer.writerow([full[arg[i]]["p_r_tgtset_explan"].replace("\n", "")])
    #             # writer.writerow([model.theta2[arg[i], :]])
    #             # label vector

    # number of labels
    labelSum = np.zeros(model.K)
    for lab in model.labels:
        labelSum[lab] += 1
    with open(str(pathResultFolder.joinpath("label.csv")), "w", encoding="utf_8_sig") as f2:
        writer = csv.writer(f2, lineterminator='\n')
        writer.writerow(["n document", model.D])
        writer.writerow([])
        writer.writerow(model.labelTypes)
        writer.writerow(labelSum)

    # for sompo data analysis
    def listToStr(a):
        s1 = ""
        for aa in a:
            s1 += "{:>5.2f}, ".format(aa)
        return s1[0:len(s1) - 2]

    # with open(str(pathResultFolder.joinpath("documents.json")), "r", encoding="utf_8_sig") as f0:
    #     reviews = json.load(f0)
    # reviews2 = []
    # for i, review in enumerate(reviews):
    #     if(int(review["p_num"]) == 1):
    #         review2 = []
    #         review2.append(review["p_r_tgtset_explan"])
    #         review2.append(listToStr(review["nWord_dk"]))
    #         review2.append(listToStr(review["theta2_d"]))
    #         for theta_dk in review["theta2_d"]:
    #             i0 = int(theta_dk / 0.5)
    #             s0 = ""
    #             for i1 in range(i0):
    #                 s0 += "#"
    #             review2.append(s0)
    #         review2.append(listToStr(review["theta3_d"]))
    #         for theta_dk in review["theta3_d"]:
    #             i0 = int(theta_dk / 0.02)
    #             s0 = ""
    #             for i1 in range(i0):
    #                 s0 += "#"
    #             review2.append(s0)
    #         if("p_r_tgtset_explan_seqs_id" in review.keys()):
    #             review2.append(review["p_r_tgtset_explan_seqs_id"])
    #         reviews2.append(review2)
    #         if(len(reviews2) > 100):
    #             break
    # with open(str(pathResultFolder.joinpath("p_r_tgtset_explan.json")), "w", encoding="utf_8_sig") as output:
    #     json.dump(reviews2, output, ensure_ascii=False, indent=2)

    print("Finish LLDA")


def testLLDAs(pathMorphomes, pathTestMorphomes, pathResultFolder, fileStr, *,
              fromPickle=False, w2vModel=None):
    """
    excute testLLDA with multiple processes
    please set following params manually:
        N:      number of model estimation
        n_jobs: number of cpus to use
    """

    def func(n):
        testLLDA(pathMorphomes,
                 pathTestMorphomes,
                 pathResultFolder.joinpath(fileStr + str(n)),
                 w2vModel=w2vModel,
                 fromPickle=False)

    N = 10
    if(not fromPickle):
        Parallel(n_jobs=2, verbose=0)([delayed(func)(n) for n in range(0, N)])

    model = None
    sumPhi = 0
    coherences5 = []
    coherences10 = []
    perplexities1 = []
    perplexities2 = []
    for n in range(N):
        with open(str(pathResultFolder.joinpath(fileStr + str(n), "model.pickle")), mode='rb') as f:
            model = pickle.load(f)
        sumPhi += model.phi
        c5 = [0] * model.K
        c10 = [0] * model.K
        for k in range(model.K):
            c5[k] = model.calcCoherence(c=5, k=k, w2v=w2vModel)
            c10[k] = model.calcCoherence(c=10, k=k, w2v=w2vModel)
        coherences5.append(c5)
        coherences10.append(c10)
        perplexities1.append(model.calcPerplexity2(model.docs, 0.2))
        perplexities2.append(model.calcPerplexity2(model.testDocs, 0.2))
    model.phi = sumPhi / N
    K = model.K
    coherences5 = np.array(coherences5).astype("float64")
    coherences10 = np.array(coherences10).astype("float64")

    LLDA.LLDA_S.writeModelToFolder(model, pathResultFolder)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "scores"
    # coherence(c=5)
    for n in range(N):
        ws.cell(row=1, column=3 + n, value="trial" + str(n)).font = Font(color="dc143c")
    ws.cell(row=1, column=3 + N, value="average score between models").font = Font(color="800000")
    ws.cell(row=1, column=5 + N, value="average model").font = Font(color="ff1493")
    ws.cell(row=2, column=1, value="coherence(c=5)").font = Font(color="dc143c")
    for k in range(K):
        if(k < len(model.labelTypes)):
            ws.cell(row=2 + k, column=2, value=model.labelTypes[k]).font = Font(color="1e90ff")
        else:
            ws.cell(row=2 + k, column=2, value="topic" + str(k)).font = Font(color="1e90ff")
        for n in range(N):
            ws.cell(row=2 + k, column=3 + n, value=coherences5[n, k])
        ws.cell(row=2 + k, column=3 + N, value=np.nanmean(coherences5[:, k]))
        ws.cell(row=2 + k, column=5 + N, value=model.calcCoherence(c=5, k=k, w2v=w2vModel))
    ws.cell(row=2 + K, column=2, value="mean").font = Font(color="0000ff")
    for n in range(N):
        ws.cell(row=2 + K, column=3 + n, value=np.nanmean(coherences5[n, :]))
    ws.cell(row=2 + K, column=3 + N, value=np.nanmean(coherences5))
    ws.cell(row=2 + K, column=5 + N, value=model.calcCoherence(c=5, w2v=w2vModel))
    # coherence(c=10)
    nowRow = K + 4
    ws.cell(row=nowRow, column=1, value="coherence(c=10)").font = Font(color="dc143c")
    for k in range(K):
        if(k < len(model.labelTypes)):
            ws.cell(row=nowRow + k, column=2, value=model.labelTypes[k]).font = Font(color="1e90ff")
        else:
            ws.cell(row=nowRow + k, column=2, value="topic" + str(k)).font = Font(color="1e90ff")
        for n in range(N):
            ws.cell(row=nowRow + k, column=3 + n, value=coherences10[n, k])
        ws.cell(row=nowRow + k, column=3 + N, value=np.nanmean(coherences10[:, k]))
        ws.cell(row=nowRow + k, column=5 + N, value=model.calcCoherence(c=10, k=k, w2v=w2vModel))
    ws.cell(row=nowRow + K, column=2, value="mean").font = Font(color="0000ff")
    for n in range(N):
        ws.cell(row=nowRow + K, column=3 + n, value=np.nanmean(coherences10[n, :]))
    ws.cell(row=nowRow + K, column=3 + N, value=np.nanmean(coherences10))
    ws.cell(row=nowRow + K, column=5 + N, value=model.calcCoherence(c=10, w2v=w2vModel))
    # perplexity
    nowRow = K * 2 + 6
    ws.cell(row=nowRow, column=1, value="perplexity(train)").font = Font(color="dc143c")
    ws.cell(row=nowRow + 1, column=1, value="perplexity(test)").font = Font(color="dc143c")
    for n in range(N):
        ws.cell(row=nowRow, column=3 + n, value=perplexities1[n])
        ws.cell(row=nowRow + 1, column=3 + n, value=perplexities2[n])
    ws.cell(row=nowRow, column=3 + N, value=np.nanmean(perplexities1))
    ws.cell(row=nowRow + 1, column=3 + N, value=np.nanmean(perplexities2))
    wb.save(str(pathResultFolder.joinpath("models.xlsx")))
